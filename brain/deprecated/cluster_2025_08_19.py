import os, sys
import csv
import json
import numpy as np
import pandas as pd
import itertools
import copy
import math
import matplotlib.pyplot as plt
import joblib  # import the joblib library
from ase import Atoms, io
import platform
import re
import shutil
import subprocess
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error


import numpy as np
from ase import Atoms
from ase.io import read, write

# Set default font properties
# plt.rc('font', size=18)  # Default text font size
# plt.rc('axes', titlesize=18)  # Title font size
# plt.rc('axes', labelsize=18)  # X and Y label font size
# plt.rc('xtick', labelsize=16)  # X tick label font size
# plt.rc('ytick', labelsize=16)  # Y tick label font size
# plt.rc('legend', fontsize=16)  # Legend font size
# plt.rc('figure', titlesize=20)  # Figure title font size
plt.rcParams['font.family'] = 'Times New Roman'
# plt.rcParams["font.family"] = "Times New Roman"
plt.rcParams["font.size"]   = 14

from matplotlib.patches import Patch
import matplotlib.patches as patches
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.collections import PatchCollection
from matplotlib.colors import ListedColormap
from matplotlib import cm
from matplotlib.colors import ListedColormap


from scipy.spatial import KDTree, ConvexHull
from scipy.spatial.transform import Rotation as R
from scipy.spatial.distance import euclidean
from scipy.integrate import odeint
from scipy.integrate import solve_ivp
from scipy.constants import h,  eV

from ase import Atoms
from ase.io import read, write
from ase.neighborlist import NeighborList, natural_cutoffs, neighbor_list
from ase.io.vasp import read_vasp, write_vasp
from ase.visualize import view

from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import cross_val_score, KFold
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, BayesianRidge, LogisticRegression
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.metrics import make_scorer, mean_absolute_error, mean_squared_error, r2_score
from sklearn.neural_network import MLPRegressor
from sklearn.decomposition import PCA
from sklearn.cross_decomposition import PLSRegression
from sklearn.pipeline import make_pipeline
# from xgboost import XGBRegressor
from scipy.optimize import curve_fit
from scipy.stats import pearsonr
from sklearn.metrics import mean_absolute_error, mean_squared_error

from mkm import * 

def round_scientific(number, decimals):
    # Format the number in scientific notation with the desired number of decimals
    formatted_number = f"{number:.{decimals}e}"
    return formatted_number

covalent_radii = np.array([1,
                           0.32, 0.46, 1.20, 0.94, 0.77, 0.75, 0.71, 0.63, 0.64, 0.67,
                           1.40, 1.25, 1.13, 1.04, 1.10, 1.02, 0.99, 0.96, 1.76, 1.54,
                           1.33, 1.22, 1.21, 1.10, 1.07, 1.04, 1.00, 0.99, 1.01, 1.09,
                           1.12, 1.09, 1.15, 1.10, 1.14, 1.17, 1.89, 1.67, 1.47, 1.39,
                           1.32, 1.24, 1.15, 1.13, 1.13, 1.08, 1.15, 1.23, 1.28, 1.26,
                           1.26, 1.23, 1.32, 1.31, 2.09, 1.76, 1.62, 1.47, 1.58, 1.57,
                           1.56, 1.55, 1.51, 1.52, 1.51, 1.50, 1.49, 1.49, 1.48, 1.53,
                           1.46, 1.37, 1.31, 1.23, 1.18, 1.16, 1.11, 1.12, 1.13, 1.32,
                           1.30, 1.30, 1.36, 1.31, 1.38, 1.42, 2.01, 1.81, 1.67, 1.58,
                           1.52, 1.53, 1.54, 1.55])

cnmax_values = {
    'top': 12,
    'bridge': 18,
    'three_fold': 22,
    'four_fold': 26
}

gas_dict = {
    "CH": -6.05202894,
    "CH2": -11.78901352,
    "CH3": -18.19317746,
    "CO": -14.78130054,
    "H": -1.11671722,
    "H2": -6.76668776,
    "N": -3.12298738,
    "N2": -16.62922486,
    "NH": -8.10061060,
    "NH2": -13.53307777,
    "NH3": -19.53586573,
    "O": -1.90384114,
    "O2": -9.85600535,
    "OH": -7.72794651,
    "OH2": -14.21849996,
    "H2O": -14.21849996,
    "S": -1.08009627,
    "SH": -6.01226074,
    "SH2": -11.19981098,
    "H2S": -11.19981098,
    "thiol": -55.96808919,
    'thiolate_lateral':-50.67527044,
    "thiolate_upright": -50.67527044,
    "slab":-354.42635463
}

def calculate_triangle_center(list_coordinates):
    return np.mean(list_coordinates, axis=0)

def is_valid_bri(atoms_positions, max_length=3.0):
    if np.linalg.norm(atoms_positions[0] - atoms_positions[1]) > max_length:
        return False
    return True

def calculate_normal(v1, v2, v3):
    """Calculate the unit normal vector of a triangle defined by three points"""
    normal = np.cross(v2 - v1, v3 - v1)
    return normal / np.linalg.norm(normal)

def calculate_cn(structure, metal, cutoff=2.7):
    """Calculate the coordination number for each atom in the structure, considering only metal atoms."""
    distances = structure.get_all_distances(mic=True)
    cn = np.zeros(len(structure))
    for i, atom in enumerate(structure):
        if atom.symbol == metal:
            neighbors = np.where((distances[i] < cutoff) & (distances[i] > 0))[0]
            cn[i] = sum(1 for n in neighbors if structure[n].symbol == metal)
    return cn

def get_CN(atoms, mult=1.0):
    radii = natural_cutoffs(atoms, mult=mult)
    nl = NeighborList(radii, bothways=True, self_interaction=False)
    nl.update(atoms)
    CN_matrix = nl.get_connectivity_matrix(sparse=False)
    CN = CN_matrix.sum(axis=0)
    return CN, CN_matrix

def calculate_gcn(structure, cn, cnmax, metal='Ru'):
    """Calculate the GCN for each atom in the structure, considering only metal atoms."""
    distances = structure.get_all_distances(mic=True)
    gcn = np.zeros(len(structure))
    for i, atom in enumerate(structure):
        if atom.symbol == metal:
            neighbors = np.where((distances[i] < 3) & (distances[i] > 0))[0]
            cn_neighbors = [cn[n] for n in neighbors if structure[n].symbol == metal]
            gcn[i] = np.sum(cn_neighbors) / cnmax
    return np.round(gcn, 2)  # 保留两位小数

def get_GCN(CN, CN_matrix, cn_max: float=None):
    if cn_max is None:
        cn_max = CN_matrix.sum(axis=0).max()
    num_atoms = CN_matrix.shape[0]
    GCN = np.zeros(num_atoms)
    for A in range(num_atoms):
        for B in range(num_atoms):
            GCN[A] += CN_matrix[A, B] * CN[B]
        GCN[A] /=  cn_max
    return GCN

def calculate_gcn_for_sites(structure, site_indices, gcn):
    """Calculate GCN values for specified sites in a structure, considering only metal atoms."""
    gcn_values = []
    for sites in site_indices:
        total_gcn = 0
        for site in sites:
            total_gcn += gcn[site]
        gcn_values.append(round(total_gcn / len(sites), 2))  # 保留两位小数
    return gcn_values

def calculate_cluster_size(structure, metal='Ru'):
    """Calculate the cluster size based on the maximum distance between metal atoms."""
    metal_positions = [atom.position for atom in structure if atom.symbol == metal]
    max_distance = 0
    for i in range(len(metal_positions)):
        for j in range(i + 1, len(metal_positions)):
            distance = np.linalg.norm(metal_positions[i] - metal_positions[j])
            if distance > max_distance:
                max_distance = distance
    return max_distance


def get_top_sites(atoms_in, metal = 'Ru', mult=0.9):
    """Obtain the exposed  top sites"""   
    filtered_atoms = [atom for atom in atoms_in if atom.symbol  in [metal]]
    atoms = Atoms(filtered_atoms)
    radii = natural_cutoffs(atoms, mult=mult)
    nl = NeighborList(radii, bothways=True, self_interaction=False)
    nl.update(atoms)
    CN_matrix = nl.get_connectivity_matrix(sparse=False)
    CN = CN_matrix.sum(axis=0)
    exposed_top_sites = [i for i, cn in enumerate(CN) if cn <= 9]
    return exposed_top_sites

def get_connection(atoms_in, metal='Ru', mult=0.9):
    # atoms_in = read(path + '/POSCAR')

    filtered_atoms = [atom for atom in atoms_in if atom.symbol in [metal]]
    atoms = Atoms(filtered_atoms)
    radii = natural_cutoffs(atoms, mult=mult)
    nl = NeighborList(radii, bothways=True, self_interaction=False)
    nl.update(atoms)
    CN_matrix = nl.get_connectivity_matrix(sparse=False)
    CN = CN_matrix.sum(axis=0)
    exposed_top_sites = [i for i, cn in enumerate(CN) if cn <= 9]       # Exposed top sites (CN <= 9)

    connections = {}     # Get the connections of all atoms
    for i in range(len(atoms)):
        connections[i] = list(np.where(CN_matrix[i])[0])

    cn_of_connected_atoms = {}      # Request 2: Calculate the CN of the connected atoms
    for i in range(len(atoms)):
        connected_indices = connections[i]
        cn_of_connected_atoms[i] = [CN[j] for j in connected_indices]

    exposed_connections = {i: [j for j in connections[i] if j in exposed_top_sites] for i in exposed_top_sites}

    # Bridge sites (pairs of connected exposed top sites)
    bridge_sites = []
    for i in exposed_top_sites:
        for j in exposed_connections[i]:
            if i < j:  # Ensure each pair is considered only once
                # The CN < 9 condition is not strictly necessary for bridge detection,
                # but it is often used to focus on undercoordinated (surface) atoms.
                bridge_sites.append([i, j])

    # Hollow sites (triplets of closely connected exposed top sites)
    hollow_sites = []
    for i in exposed_top_sites:
        for j in exposed_connections[i]:
            for k in exposed_connections[j]:
                if i < j and j < k and i in exposed_connections[k]:  # Ensure each triplet is considered only once and forms a triangle
                    # The CN < 9 condition is not strictly necessary for hollow detection,
                    # but it is often used to focus on undercoordinated (surface) atoms.
                    hollow_sites.append([i, j, k])

    # Square sites (quartets of connected exposed top sites forming a closed loop)
    square_sites = []
    for i in exposed_top_sites:
        for j in exposed_connections[i]:
            if j <= i:
                continue
            for k in exposed_connections[j]:
                if k in (i, j) or k <= j:
                    continue
                for l in exposed_connections[k]:
                    if l in (i, j, k) or l <= k:
                        continue
                    # Check if l connects back to i to form a closed loop
                    if i in exposed_connections[l]:
                        # Ensure all connections exist: i-j, j-k, k-l, l-i
                        square = [i, j, k, l]
                        # The CN < 9 condition is not strictly necessary for square detection,
                        # but it is often used to focus on undercoordinated (surface) atoms.
                        square_sites.append(square)

    return connections, cn_of_connected_atoms, exposed_top_sites, bridge_sites, hollow_sites, square_sites

def number_to_letter(num):
    """Conver the CN to letters"""
    if num == 0:
        return '0'
    return chr(ord('a') + num - 1)


def get_CN_GA(path, mult=0.9, metal='Ru'):
    
    '''
    Important: it is the path for the clean cluster. 
    1) the exposed surface of the cluster will be scanned and all the top, bridge, and hollow sites will be stored.
    2) the sites will be stored as a dictionray 
    3) all the groups will be stored too to  fix the order of the  groups in the matrix
    '''    
    file_in = os.path.join(path,'POSCAR')
    atoms = read(file_in)
    print(metal)
    connections, cn_of_connected_atoms, exposed_top_sites, bridge_sites, hollow_sites, square_sites  = get_connection(atoms, metal=metal, mult=0.9)
    
    def get_one_site_string(site):
        """Obtain the text string representation for single site """
        cn_values = sorted(cn_of_connected_atoms[site])
        # print(cn_values)
        if len(cn_values) < 13:
            cn_values += [0] * (13 - len(cn_values))  # Pad with zeros if less than 12 elements
        else: 
            cn_values += [0] * (13 - 12)  # Pad with zeros if less than 12 elements
        
        cn_string = ''.join(number_to_letter(num) for num in cn_values)
            
        return cn_string
    
    def get_groups(site_indices, site_type):
        """
        Generic function to get group strings for a site (top, bridge, hollow, square).
        site_indices: int for top, list of ints for others.
        site_type: 'top', 'bridge', 'hollow', or 'square'
        """
        if site_type == 'top':
            site = site_indices
            groups = [get_one_site_string(site)]
            groups += [get_one_site_string(s) for s in connections[site]]
            return groups
        else:
            sites = site_indices
            strings = [get_one_site_string(i) for i in sites]
            groups = ['-'.join(sorted(strings))]
            # Collect all unique surrounding atoms except the site atoms themselves
            all_neighbors = set()
            for i in sites:
                all_neighbors.update(connections[i])
            surrounding_atoms = list(all_neighbors - set(sites))
            groups += [get_one_site_string(s) for s in surrounding_atoms]
            return groups

    GA_dict = {}
    for site in exposed_top_sites:
        key = site + 1
        GA_dict[key] = get_groups(site, 'top')

    for sites in bridge_sites:
        key = '_'.join(str(i + 1) for i in sites)
        GA_dict[key] = get_groups(sites, 'bridge')

    for sites in hollow_sites:
        key = '_'.join(str(i + 1) for i in sites)
        GA_dict[key] = get_groups(sites, 'hollow')

    for sites in square_sites:
        key = '_'.join(str(i + 1) for i in sites)
        GA_dict[key] = get_groups(sites, 'square')

    GA_file = os.path.join(path, 'GA_dict.txt')

    # Writing JSON data
    with open(GA_file, 'w') as f:
        json.dump(GA_dict, f)
        
    # Summarize and extract all unique groups from GA_dict
    groups = {group for sublist in GA_dict.values() for group in sublist}
    
    # Save groups to a text file, one per line
    groups_file = os.path.join(path, 'groups.txt')
    with open(groups_file, 'w') as f:
        for group in groups:
            f.write(f"{group}\n")


def func(x, a, b, c):
    return -0.5 * a * x**2 - b * x + c


def get_dipole_polarization(GA_matrix, fill_nans_with_fit=True, verbose=True):
    def func(x, a, b, c):
        return -0.5 * a * x**2 - b * x + c

    x_values = np.array([-0.6, -0.4, -0.2, 0.0, 0.2, 0.4, 0.6])
    eads_cols = ['E_ads_' + i for i in ["-0.6", "-0.4", "-0.2", "0.0", "0.2", "0.4", "0.6"] ]
    site_col = 'Site' if 'Site' in GA_matrix.columns else 'site'

    results_list = []
    GA_matrix = GA_matrix.copy()  # 避免修改原始 DataFrame

    for idx, row in GA_matrix.iterrows():
        site = row[site_col]
        y_values_orig = row[eads_cols].values.astype(float)
        valid_mask = ~np.isnan(y_values_orig)
        x_fit = x_values[valid_mask]
        y_fit = y_values_orig[valid_mask]

        if len(x_fit) < 3:
            if verbose:
                print(f"Skipping site '{site}' (index {idx}) due to insufficient data points.")
            continue

        try:
            # 第一次拟合
            params, _ = curve_fit(func, x_fit, y_fit)
            y_pred = func(x_fit, *params)
            deviation = np.abs(y_pred - y_fit)
            outlier_found = False

            if np.any(deviation > 0.2):
                outlier_found = True
                if verbose:
                    print(f"Site '{site}' (index {idx}) has deviations > 0.2 eV:")
                for xi, yi, ypi, dev in zip(x_fit, y_fit, y_pred, deviation):
                    if dev > 0.2:
                        if verbose:
                            print(f"  E={xi:.1f}: actual={yi:.3f}, predicted={ypi:.3f}, Δ={dev:.3f}")
                        GA_matrix.at[idx, f"{xi:.1f}"] = np.nan

                # 重新提取数据
                y_values_clean = GA_matrix.loc[idx, eads_cols].values.astype(float)
                valid_mask = ~np.isnan(y_values_clean)
                x_fit = x_values[valid_mask]
                y_fit = y_values_clean[valid_mask]

                if len(x_fit) < 3:
                    if verbose:
                        print(f"After removing outliers, not enough data for site '{site}' (index {idx})")
                    continue

                # 第二次拟合
                params, _ = curve_fit(func, x_fit, y_fit)
                y_pred = func(x_fit, *params)

            # 填补初始 NaN（如果没有出现 outlier）
            elif fill_nans_with_fit and np.any(np.isnan(y_values_orig)):
                full_pred = func(x_values, *params)
                for j, val in enumerate(y_values_orig):
                    if np.isnan(val):
                        if verbose:
                            print(f"  Filling NaN for site '{site}' at E={x_values[j]:.1f} with predicted value {full_pred[j]:.3f}")
                        GA_matrix.at[idx, f"{x_values[j]:.1f}"] = full_pred[j]

            a, b, c = params
            R2 = r2_score(y_fit, y_pred)
            mae = mean_absolute_error(y_fit, y_pred)
            rmse = np.sqrt(mean_squared_error(y_fit, y_pred))

            results_list.append({
                'site': site,
                'polarizability': a,
                'dipole': b,
                'c': c,
                # 'R2': R2,
                # 'MAE': mae,
                # 'RMSE': rmse
            })

        except Exception as e:
            if verbose:
                print(f"Error fitting site '{site}' (index {idx}): {e}")
            continue

    results_df = pd.DataFrame(results_list)
    return results_df#, GA_matrix




def load_ga_data(cluster_path):
    """
    Loads the GA dictionary and the groups list from files.

    Parameters:
        load_path (str): The directory where the files are stored.
                         Files are expected to be named 'GA_dict.txt' and 'groups.txt'.
    
    Returns:
        tuple: A tuple containing the GA dictionary and the groups list.
    """
    # Load GA_dict from JSON file
    ga_file = os.path.join(cluster_path, 'GA_dict.txt')
    with open(ga_file, 'r') as f:
        GA_dict = json.load(f)
    
    # Load groups from the text file
    groups_file = os.path.join(cluster_path, 'groups.txt')
    with open(groups_file, 'r') as f:
        groups = [line.strip() for line in f if line.strip()]
    
    return GA_dict, groups


def get_full_GA_matrix(cluster_path):
    """
    Generate the GA (Group Additivity) matrix for all the surface sites
    """
    GA_dict, groups = load_ga_data(cluster_path)
    # Prepare header for the GA matrix DataFrame
    header = ['site'] + groups
    GA_matrix = []

    for key,values in GA_dict.items():
        groups_geo = values
        row = [key]
        # Create the GA matrix row: first column is the site label, then counts for each group
        for group in groups:
            count = groups_geo.count(group)
            row.append(count)
        GA_matrix.append(row)

    # Create a DataFrame from the GA matrix
    df_GA = pd.DataFrame(GA_matrix, columns=header)

    
    # Save the GA matrix to CSV
    output_csv = cluster_path + "/GA_matrix_full.csv"
    
    df_GA.to_csv(output_csv, index=False)
    print("GA matrix saved as", output_csv)
    
    return df_GA


def generate_GA_matrix_species(cluster_path, list_EF):
    
    ga_matrix_file = os.path.join(cluster_path, 'GA_matrix_full.csv')
    print("Reading from:", ga_matrix_file)

    with open(ga_matrix_file, 'r') as f:
        first_line = f.readline()
        print("Actual header line in file:")
        print(first_line.strip())
        
    # GA_matrix_full = get_full_GA_matrix(cluster_path)
    GA_matrix_full = pd.read_csv(ga_matrix_file)
    print("Original GA_matrix_full columns from slab folder:")
    print(GA_matrix_full.columns.tolist())
    
    sites = GA_matrix_full['site'].tolist()
    GA_matrix_species = GA_matrix_full.copy()
    # Ensure results directory exists
    if not os.path.exists('results'):
        os.makedirs('results')
    
    data_file = './sorted_data.csv'
    df_data = pd.read_csv(data_file)

    species = get_species_from_cwd()  

    for ef_val in list_EF:
        eads_ef = []
        for site in sites:         
        # Retrieve the slab energy from df where Species equals 'slab'
            try:
                slab_energy = df_data.loc[df_data['Species'] == 'slab', ef_val].values[0]
                # print(slab_energy)
            except KeyError:
                raise KeyError(f"EF value '{ef_val}' not found in the data.")
    
            # Compute the adsorption energy: subtract slab energy and a gas-phase correction
    
            if species in ['H', 'N', 'O']:
                gas_species = species + '2'
                E_gas = float(gas_dict[gas_species]) / 2
            else:
                gas_species = species
                E_gas = float(gas_dict[gas_species])
                
            # print('Eslab', slab_energy)
            try:
                ef_value = df_data.loc[df_data['site'] == site, ef_val].values[0]
                eads_value = ef_value - slab_energy - E_gas
            except :
                eads_value = float('nan') 
            eads_ef.append(eads_value)
            # Add the computed energy column to the GA matrix DataFrame.
            # (Assumes the row order of df and df_GA is the same.)
        GA_matrix_species[f'E_ads_{ef_val}'] = eads_ef
    # print("Original GA_matrix_full columns:")
    # print(GA_matrix_full.columns.tolist())
    
    # print("\nModified GA_matrix_species columns:")
    # print(GA_matrix_species.columns.tolist())
    #     # Save the GA matrix to CSV
    output_csv = "./GA_matrix.csv"
    
    GA_matrix_species.to_csv(output_csv, index=False)
    print("GA matrix of the species is saved as", output_csv)
    
    return GA_matrix_species

def update_GA_matrix_with_dipole_polarization(csv_filepath='./GA_matrix.csv'):
    
    """
    Update the GA_matrix by adding 'polarizability' and 'dipole' columns based on curve fitting.
    
    This function reads the GA_matrix CSV file, computes the dipole polarization fits for each row 
    (using get_dipole_polarization, which returns columns 'polarizability' and 'dipole'), merges these 
    values into the original GA_matrix based on the 'site' column, and then overwrites the CSV file.
    
    Args:
        csv_filepath (str): Path to the GA_matrix CSV file (default: './GA_matrix.csv').
    
    Returns:
        updated_df (pandas.DataFrame): The updated GA_matrix DataFrame with the new columns.
        
            
    """
    # Read the current GA_matrix.
    df_GA = pd.read_csv(csv_filepath)
    
    # Get the dipole polarization results.
    # This function must be defined and should return a DataFrame with columns: 
    # 'site', 'polarizability', 'dipole', 'c', 'R2', 'MAE', 'RMSE'
    dipole_results = get_dipole_polarization(df_GA)
    
    # Merge the dipole polarization results into the GA_matrix based on the 'site' column.
    updated_df = df_GA.merge(dipole_results[['site', 'polarizability', 'dipole', 'c']], on='site', how='left')
    
    # Overwrite the original GA_matrix CSV file with the updated DataFrame.
    updated_df.to_csv(csv_filepath, index=False)
    print(f"Updated GA_matrix saved to {csv_filepath}.")
    
    return updated_df

    
def get_species_from_cwd():
    # cwd = os.getcwd()
    cwd = os.path.realpath(os.getcwd())
    # print(cwd, 'cwd')
    species = None
    try:
        # path_parts = cwd.replace(' ', '').split(os.path.sep)
        part = os.path.basename(cwd).strip()
        # print('part', part)
        if '_ads' in part:
            species = part.split('_')[0]
        
    except Exception as e:
        print(f"An error occurred: {e}")
    
    if species:
        return species
    else:
        print("No species found in the current working directory.")


def get_GA_matrix(cluster_path, EF):
    """
    Retrieve the GA matrix from the CSV file if it exists; otherwise, generate it.
    Then extract the feature matrix X and target vector Y.

    Args:
        EF (str): The EF value (as a string) used to label the adsorption energy column.
                  For example, if EF='0.6', the energy column is assumed to be 'E_ads_0.6'.

    Returns:
        tuple: (df_matrix, X, Y, n_group)
            - df_matrix: The full GA matrix DataFrame (including 'site' and adsorption energy columns).
            - X: The features as a NumPy array (with 'site' and all adsorption energy columns removed).
            - Y: The target adsorption energies (from the corresponding column).
            - n_group: The number of feature columns in X.
    """
    csv_filepath = './GA_matrix.csv'
    df_matrix = pd.read_csv(csv_filepath)
    
    prop_list = ["E_ads_-0.6", "E_ads_-0.4", "E_ads_-0.2", "E_ads_0.0", "E_ads_0.2", "E_ads_0.4", "E_ads_0.6", "polarizability", "dipole", "c"]
   
    # if any(col.startswith('Eads_') for col in df_matrix.columns):
    #     prefix = 'Eads_'
    # else:
    #     raise KeyError("No adsorption energy columns found in GA_matrix.")
    
    # Determine the target adsorption energy column name based on EF.
    # If EF is empty, we remove the trailing underscore from the prefix.
    # col_name = prefix[:-1] if EF == "" else prefix + EF

    # # Extract Y values (adsorption energies)
    # try:
    #     Y = df_matrix[col_name].values
    # except KeyError:
    #     raise KeyError(f"The expected adsorption energy column '{col_name}' is not found in the GA matrix.")
    
    # Extract Y values (adsorption energies)
    if EF not in  ["polarizability", "dipole", "c"]:
        prop = 'E_ads_' + str(EF) 
    else: 
        prop = EF
    try:
        Y = df_matrix[prop].values
    except KeyError:
        raise KeyError(f"The expected adsorption energy column '{EF}' is not found in the GA matrix.")
    
    # Drop the 'site' column and all adsorption energy columns (those starting with the prefix) to get X.
    cols_to_drop = ['site'] + [col for col in prop_list]
    df_features = df_matrix.drop(columns=cols_to_drop)
    X = df_features.values
    n_group = df_features.shape[1]
    
    return df_matrix, X, Y, n_group


def GA_prediction(X_training, Y_training):
    '''Run conventional GA approach '''
 
    model = PseudoInverseLinearRegression() 
    model.fit(X_training,Y_training)

    Y_pred = model.predict(X_training)
    
    # Calculate the MAE, RMSE, and R²
    mae = mean_absolute_error(Y_training, Y_pred)
    rmse = np.sqrt(mean_squared_error(Y_training, Y_pred))
    r2 = r2_score(Y_training, Y_pred)
    
    # Print the errors and R²
    print(f'Mean Absolute Error (MAE): {mae}')
    print(f'Root Mean Squared Error (RMSE): {rmse}')
    print(f'R-squared (R²): {r2}')
         
    # # Plot the data and regression line
    plt.figure(figsize=(10, 6))
    plt.scatter(Y_training, Y_pred, color='blue', label='Model')
    
    # Add text box for metrics
    textstr = f'MAE: {mae:.2f}\nRMSE: {rmse:.2f}\nR2: {r2:.2f}'
    plt.gca().text(0.65, 0.25, textstr, transform=plt.gca().transAxes,
                    fontsize=16, verticalalignment='top', bbox=dict(facecolor='white', alpha=0))
    
    # Customize the plot
    plt.xlabel('X / Unit', fontsize=16)
    plt.ylabel('Y / Unit', fontsize = 16)
    plt.xticks(fontsize=16)
    plt.yticks(fontsize=16)
    plt.legend(loc='upper left', frameon=False, fontsize=18)
    #plt.title('Linear Regression of De vs Ea')
    plt.tight_layout()
    # Save the figure
    plt.savefig('XY.png')
    return Y_training, Y_pred

    
def save_one_model_split(species, df, EF,  X, Y, models, model_name, ratio=0.95):
    model = models[model_name]
    n_samples = int(len(X) * ratio)  # 80% of the data for training
    indices = np.random.choice(len(X), size=len(X), replace=True)
    X_train, Y_train = X[indices[:n_samples]], Y[indices[:n_samples]]   
    model.fit(X_train,Y_train)
    
    joblib.dump(model, f'{species}_{model_name}_{EF}.pkl')


def test_one_model_split(df, EF,  X, Y, models, model_name, ratio):
    '''Run conventional GA approach '''
    model = models[model_name]
    
    n_samples = int(len(X) * ratio)  # 80% of the data for training
    
    indices = np.random.choice(len(X), size=len(X), replace=True)
    X_train, Y_train = X[indices[:n_samples]], Y[indices[:n_samples]]
    X_test, Y_test = X[indices[n_samples:]], Y[indices[n_samples:]]
    
    model.fit(X_train,Y_train)
    Y_pred = model.predict(X_test)
    
    # Calculate the MAE, RMSE, and R²
    mae = mean_absolute_error(Y_test, Y_pred)
    rmse = np.sqrt(mean_squared_error(Y_test, Y_pred))
    r2 = r2_score(Y_test, Y_pred)

    fig, ax  = plt.subplots(figsize=(10,8))  
    
    species = get_species_from_cwd()
    text = '%s adsorption' %(species)
    # ax.set_title(text)
    ax.scatter(Y_test, Y_pred, edgecolor='k', alpha=0.7, s=200, label = text, color='#1f77b4')
    ax.plot([min(Y), max(Y)], [min(Y), max(Y)], 'r--', linewidth=3)
    ax.set_xlabel(f'DFT Eads (eV) at EF = {EF} eV/A', fontsize=36)
    ax.set_ylabel('Pred. Eads(eV)', fontsize=36)

    textstr = f'MAE: {mae:.2f}  \nRMSE: {rmse:.2f}  \nR2: {r2:.2f}'
    plt.gca().text(0.65, 0.25, textstr, transform=plt.gca().transAxes,
                    fontsize=20, verticalalignment='top', bbox=dict(facecolor='white', alpha=0))


    for i in range(len(Y_test)):
        if abs(Y_test[i] - Y_pred[i]) > 0.15: # 0.15 eV difference
            ax.annotate(df['site'][i], (Y_test[i], Y_pred[i]), textcoords="offset points", xytext=(0,10), ha='center')
            ax.scatter(Y_test[i], Y_pred[i], color='red', s=100, edgecolor='k')
    
    
    plt.legend(loc='upper left', frameon=False, fontsize=40)

    plt.grid(False)
    ax = plt.gca()
    for spine in ax.spines.values():
        spine.set_linewidth(3)
        
    #plt.title('Linear Regression of De vs Ea')
    ax.tick_params(axis='both', which='major', width=2, length=8, labelsize=28)
    plt.tight_layout()
    plt.savefig(f'results/XY_%s_testing_%s_{EF}.png' %(model_name, str(1-ratio)), dpi=600)
    
    return mae, rmse, r2 
        
 
class PseudoInverseLinearRegression:
    def __init__(self, fit_intercept=True):
        self.beta_ = None
        self.fit_intercept = fit_intercept

    def fit(self, X, Y):
        if self.fit_intercept:
            # Add a column of ones to X to account for the intercept
            X = np.hstack([np.ones((X.shape[0], 1)), X])
        
        # Compute the beta coefficients using the Moore-Penrose pseudoinverse
        self.beta_ = np.linalg.pinv(X).dot(Y)

    def predict(self, X):
        if self.fit_intercept:
            # Add a column of ones to X for predictions
            X = np.hstack([np.ones((X.shape[0], 1)), X])
        
        # Make predictions using the computed beta coefficients
        return X.dot(self.beta_)
    

# # Now include this custom model in the models dictionary
def refine_models(n_group):
    models = {
        "Ridge": Ridge(alpha=1.0),
        "Lasso": Lasso(alpha=1/n_group),
        "EN": ElasticNet(alpha=0.1, l1_ratio=0.5),
        "DT": DecisionTreeRegressor(max_depth=n_group),
        "RF": RandomForestRegressor(n_estimators=n_group*5, max_depth=n_group*5),
        "GB": GradientBoostingRegressor(n_estimators=n_group*5, max_depth=n_group),
        "SVR": SVR(kernel='linear', C=1.0),
        "GA": PseudoInverseLinearRegression(),
        # "XGB": XGBRegressor(n_estimators=11*5, max_depth=11, learning_rate=0.1),
        }
    return models


def bootstrap_cross_validation_one_model(models, model_name, EF, X, Y, ratio, n_bootstrap=100):
    """
    Perform bootstrapped cross-validation for a single model and save results.
    
    Args:
        models (dict): Dictionary of models with their names as keys.
        model_name (str): The name of the model to use for cross-validation.
        X (np.array): Features dataset.
        Y (np.array): Target values.
        ratio (float): Proportion of data to use for training (e.g., 0.8 for 80%).
        n_bootstrap (int): Number of bootstrap iterations.
    
    Returns:
        dict: Dictionary containing metric results for the model.
    """
    model = models[model_name]
    results = {metric: [] for metric in ['MAE', 'MSE', 'RMSE', 'R2', 'MPD', 'MND']}
    n_samples = int(len(X) * ratio)  # Number of training samples

    for i in range(n_bootstrap):
        print(f'Iteration: {i + 1}/{n_bootstrap}')
        
        # Generate bootstrap indices
        indices = np.random.choice(len(X), size=len(X), replace=True)
        X_train, Y_train = X[indices[:n_samples]], Y[indices[:n_samples]]
        X_test, Y_test = X[indices[n_samples:]], Y[indices[n_samples:]]
        
        # Train the model and predict
        model.fit(X_train, Y_train)
        predictions = model.predict(X_test)
        
        # Calculate metrics
        results['MAE'].append(mean_absolute_error(Y_test, predictions))
        results['MSE'].append(mean_squared_error(Y_test, predictions))
        results['RMSE'].append(np.sqrt(mean_squared_error(Y_test, predictions)))
        results['R2'].append(r2_score(Y_test, predictions))
        results['MPD'].append(np.max(predictions - Y_test))
        results['MND'].append(np.min(predictions - Y_test))
    
    # Save results to CSV files
    results_dir = 'results'
    os.makedirs(results_dir, exist_ok=True)  # Ensure the results directory exists
    
    for metric, values in results.items():
        metric_file = f'{results_dir}/{metric.lower()}_{ratio}_{EF}.csv'
        averages_file = f'{results_dir}/{metric.lower()}_averages_{ratio}_{EF}.csv'
        
        # Save all values
        pd.DataFrame({model_name: values}).to_csv(metric_file, index=False)
        print(f"Saved all {metric} values to {metric_file}_{EF}")
        
        # Save averages
        pd.DataFrame({model_name: [np.mean(values)]}).to_csv(averages_file, index=False)
        print(f"Saved average {metric} value to {averages_file}_{EF}")

    return results


def clean_ga_matrix(df):
    """
    Cleans a GA matrix DataFrame by:
    1. Removing the first row (assumed to be an extra header),
    2. Dropping the 'site' column,
    3. Converting all remaining values to numeric.
    
    Returns a cleaned DataFrame.
    """
    return df.iloc[1:].drop(columns=['site']).reset_index(drop=True).apply(pd.to_numeric, errors='coerce')


def active_learning_one_model(models, model_name, EF, X, Y, sites, 
                              initial_ratio=0.4, increment=0.05, max_ratio=0.95,
                              error_threshold=0.5, results_dir='results'):
    """
    Perform an active learning process for a single model, saving both training
    and testing predictions and metrics each iteration, and finally output a
    combined CSV of train/test metrics.

    Returns:
        list of dict: metrics per iteration.
    """
    from sklearn.pipeline import Pipeline
    from sklearn.preprocessing import StandardScaler

    # Step 1: Remove NaNs before anything else
    valid_mask = ~np.isnan(Y)
    X = X[valid_mask]
    Y = Y[valid_mask]
    sites = np.array(sites)[valid_mask]

    total_samples = len(X)
    idx = np.arange(total_samples)
    np.random.shuffle(idx)

    # Initial train/test split
    n_train = int(total_samples * initial_ratio)
    train_idx = idx[:n_train].tolist()
    test_idx  = idx[n_train:].tolist()

    os.makedirs(results_dir, exist_ok=True)
    results = []
    iteration = 0

    while len(train_idx)/total_samples < max_ratio and test_idx:
        
        iteration += 1

        # Prepare data
        X_tr, Y_tr = X[train_idx], Y[train_idx]
        X_te, Y_te = X[test_idx], Y[test_idx]
        sites_tr = np.array(sites)[train_idx]
        sites_te = np.array(sites)[test_idx]

        # Fit model
        model = models[model_name]
        model.fit(X_tr, Y_tr)

        # Predict
        Y_tr_pred = model.predict(X_tr)
        Y_te_pred = model.predict(X_te)
        Y_all_pred = model.predict(X)

        # Save model
        species = get_species_from_cwd()
        pipe = Pipeline([('scaler', StandardScaler()),('model', model) ])
        pipe.fit(X_tr, Y_tr)
        joblib.dump(pipe, f'{species}_{model_name}_{EF}.pkl')

        # Compute train metrics
        train_mae  = mean_absolute_error(Y_tr, Y_tr_pred)
        train_mse  = mean_squared_error(Y_tr, Y_tr_pred)
        train_rmse = np.sqrt(train_mse)
        train_r2   = r2_score(Y_tr, Y_tr_pred)
        tr_dev     = Y_tr_pred - Y_tr
        train_mpd  = tr_dev.max()
        train_mnd  = tr_dev.min()

        # Compute test metrics
        mae  = mean_absolute_error(Y_te, Y_te_pred)
        mse  = mean_squared_error(Y_te, Y_te_pred)
        rmse = np.sqrt(mse)
        r2   = r2_score(Y_te, Y_te_pred)
        dev  = Y_te_pred - Y_te
        mpd  = dev.max()
        mnd  = dev.min()

        # Save predictions
        pd.DataFrame({
            'site': sites_tr,
            'Y_true': Y_tr,
            'Y_pred': Y_tr_pred
        }).to_csv(
            os.path.join(results_dir, f'train_preds_iter{iteration}_{EF}.csv'),
            index=False
        )

        pd.DataFrame({
            'site': sites_te,
            'Y_true': Y_te,
            'Y_pred': Y_te_pred
        }).to_csv(
            os.path.join(results_dir, f'test_preds_iter{iteration}_{EF}.csv'),
            index=False
        )
            
        #Save ALL predictions
        pd.DataFrame({
            'site': sites,
            'Y_true_all': Y,
            'Y_pred_all': Y_all_pred
        }).to_csv(
            os.path.join(results_dir, f'all_preds_iter{iteration}_{EF}.csv'),
            index=False
        )

        # Log iteration
        print(f"Iteration {iteration}: train MAE={train_mae:.4f}, test MAE={mae:.4f}")

        # Outlier analysis
        errs = np.abs(dev)
        mask = errs > error_threshold
        if mask.any():
            outs = np.where(mask)[0]
            for i in outs[np.argsort(-errs[outs])]:
                print(f"  Outlier: site={sites_te[i]}, err={errs[i]:.3f}")
        else:
            print("  No outliers above threshold.")

        # Record metrics
        results.append({
            'iteration':        iteration,
            'training_ratio':   len(train_idx)/total_samples,
            'n_train':          len(train_idx),
            'n_test':           len(test_idx),

            'train_MAE':        train_mae,
            'train_MSE':        train_mse,
            'train_RMSE':       train_rmse,
            'train_R2':         train_r2,
            'train_MPD':        train_mpd,
            'train_MND':        train_mnd,

            'test_MAE':         mae,
            'test_MSE':         mse,
            'test_RMSE':        rmse,
            'test_R2':          r2,
            'test_MPD':         mpd,
            'test_MND':         mnd,
        })

        # Active learning: add top-error test samples to train
        n_add = max(1, int(total_samples * increment))
        n_add = min(n_add, len(test_idx))
        top_idxs = np.argsort(-errs)[:n_add]
        selected = [test_idx[i] for i in top_idxs]
        train_idx.extend(selected)
        test_idx = [i for i in test_idx if i not in selected]

    # Save all iteration metrics
    df = pd.DataFrame(results)
    base_cols = ['iteration','training_ratio','n_train','n_test']
    metric_cols = ['MAE','MSE','RMSE','R2','MPD','MND']
    cols = base_cols + [f'train_{m}' for m in metric_cols] + [f'test_{m}' for m in metric_cols]
    out_file = os.path.join(results_dir, f'active_learning_metrics_{EF}.csv')
    df[cols].to_csv(out_file, index=False)
    print(f"✅ Saved combined metrics to {out_file}")

    return results



def get_matrix_to_be_predicted(cluster_path, site):
    
    site = str(site)
    
    ga_matrix_file = os.path.join(cluster_path, 'GA_matrix_full.csv')
    GA_matrix_full = pd.read_csv(ga_matrix_file)
    
    def get_feature_by_site(df, site):
        """
        根据 site 值获取对应的一行特征（去掉 'site' 列）。
    
        参数：
            df (pd.DataFrame): 原始DataFrame，包含'site'列和多个特征列
            site (int): 要查询的 site 编号
    
        返回：
            np.ndarray: shape 为 (1, n_features) 的 NumPy 数组
        """
        row = df[df['site'] == site]
        if row.empty:
            raise ValueError(f"Site {site} not found in the DataFrame.")
        return row.drop(columns=['site']).values
    
    matrix_site = get_feature_by_site(GA_matrix_full, site)
    # # Load or generate GA data
    # try: 
    #     GA_dict, groups = load_ga_data(cluster_path)
    # except Exception as e:
    #     print("Error loading GA data:", e)
    #     get_CN_GA(cluster_path, mult=0.9)
    #     GA_dict, groups = load_ga_data(cluster_path)
    # # print(GA_dict)
    # """ Get the group matrix of one site"""
    # try:
    #     groups_site = GA_dict[site]
    # except:
    #     site = str(site)
    #     groups_site = GA_dict[site]
    # matrix_site = []
    # for group in groups:
    #     num_group = 0
    #     for group_site in groups_site:
    #         if group_site == group:
    #             num_group += 1
    #     matrix_site.append(num_group)
    # matrix_site = np.array([matrix_site])

    return matrix_site

def predict_Eads_site(cluster_path, species, site, Prop):
    '''Print the adsorption energy at the specific site, 
    the data in the data_path is applied to train the GA model to predict the 
    same species at the provided site.
    Prop values: -0.6, -0.4, -0.2, 0.0, 0.2, 0.4, 0.6, polarizability, dipole, all are strings.
    '''
    preidct_matrix =  get_matrix_to_be_predicted(cluster_path, site)
    
    # try: 
    pkl_file = os.path.join(cluster_path,f'{species}_GA_{Prop}.pkl')
    # print(pkl_file)
    GA_model = joblib.load(pkl_file)  
    # GA_model = joblib.load(f'{species}_GA_{Prop}.pkl')  
    
#     # print('Good3',preidct_matrix )
#     ### predict the energies of species at the site    
    E_species_ef = GA_model.predict(preidct_matrix)[0]
        
    #     if species == 'NH3':
    #         print('E_NH3', E_species_ef, 'site', site)
        
    # except: 
    # try:
        # pkl_file = os.path.join(cluster_path,f'{species}_GA_0.0.pkl')
        # print(pkl_file)
        # GA_model = joblib.load(pkl_file)  
    
        # GA_model = joblib.load(f'{species}_GA_{Prop}.pkl')  
        # preidct_matrix =  get_matrix_to_be_predicted(cluster_path, site)
        # print('Good3',preidct_matrix )
        ### predict the energies of species at the site    
        # E_species = GA_model.predict(preidct_matrix)[0]
        # print('E', E_species)
    
    # # Prediction via Taylor Expansion
    # pkl_polarizability = os.path.join(cluster_path,f'{species}_GA_polarizability.pkl')
    # pkl_dipole = os.path.join(cluster_path,f'{species}_GA_dipole.pkl')
    # pkl_constant = os.path.join(cluster_path,f'{species}_GA_c.pkl')
    
    # polarizability_model = joblib.load(pkl_polarizability)  
    # dipole_model = joblib.load(pkl_dipole)  
    # constant_model = joblib.load(pkl_constant)  
    
    # polarizability  = polarizability_model.predict(preidct_matrix)[0]
    # dipole  = dipole_model.predict(preidct_matrix)[0]
    # constant = constant_model.predict(preidct_matrix)[0]
    
    # E_species_ef = func(float(Prop), polarizability, dipole, constant)  
        
        # def func(x, a, b, c):
    #     #     return -0.5 * a * x**2 - b * x + c
        
    #     print(E_species_ef)
    # except:
    #     print('failed')
    return E_species_ef


def plot_active_learning_from_csv(csv_filepath, EF, results_dir='results'):
    """
    Read the combined active learning results CSV and plot both training and testing
    performance metrics vs. training ratio.

    The CSV is expected to have columns:
      'training_ratio',
      'train_MAE','train_MSE','train_RMSE','train_R2','train_MPD','train_MND',
      'test_MAE', 'test_MSE', 'test_RMSE', 'test_R2', 'test_MPD', 'test_MND'

    Args:
        csv_filepath (str): Path to the combined CSV file.
        EF (str or float): The EF value for labeling the plot.
        results_dir (str): Directory where the plot will be saved.
    """
    df = pd.read_csv(csv_filepath)
    df.sort_values('training_ratio', inplace=True)

    plt.figure(figsize=(10, 6))
    # Plot MAE
    plt.plot(df['training_ratio'], df['train_MAE'],  marker='o', linestyle='--', color = 'tab:blue', label='Train MAE')
    plt.plot(df['training_ratio'], df['test_MAE'],   marker='o', linestyle='-', color = 'tab:red', label='Test MAE')
    # Plot RMSE
    plt.plot(df['training_ratio'], df['train_RMSE'], marker='^', linestyle='--', color = 'tab:blue', label='Train RMSE')
    plt.plot(df['training_ratio'], df['test_RMSE'],  marker='^',linestyle='-',  color = 'tab:red', label='Test RMSE')
    # # Plot MPD
    # plt.plot(df['training_ratio'], df['train_MPD'],  marker='D', label='Train MPD')
    # plt.plot(df['training_ratio'], df['test_MPD'],   marker='d', label='Test MPD')
    # # Plot MND
    # plt.plot(df['training_ratio'], df['train_MND'],  marker='x', label='Train MND')
    # plt.plot(df['training_ratio'], df['test_MND'],   marker='*', label='Test MND')

    plt.xlabel('Training Ratio', fontsize=16)
    plt.ylabel('Metric Value', fontsize=16)
    plt.title(f'Active Learning Metrics vs. Training Ratio (EF = {EF})', fontsize=18)
    plt.legend(fontsize=12, framealpha=0.8)
    plt.grid(True)
    plt.tight_layout()

    # Ensure results_dir exists
    os.makedirs(results_dir, exist_ok=True)
    out_png = os.path.join(results_dir, f'active_learning_metrics_{EF}.png')
    plt.savefig(out_png, dpi=300)
    plt.close()
    print(f"✅ Plot saved to {out_png}")


    
def plot_active_predicting_from_csv(predict_csv_dir, ratio, EF):
    """
    Read training and testing prediction CSVs for a given iteration (ratio) and EF,
    plot true vs. predicted scatter for both sets, annotate with MAEs calculated
    directly from the data, and save the figure.

    Args:
        predict_csv_dir (str): Directory containing the CSVs.
        ratio (int or str): The iteration number (used in the filename).
        EF (float or str): The EF value (used in the filename and title).
    """
    # Paths to prediction CSVs
    train_fp = os.path.join(predict_csv_dir, f'train_preds_iter{ratio}_{EF}.csv')
    test_fp  = os.path.join(predict_csv_dir, f'test_preds_iter{ratio}_{EF}.csv')

    # Read data
    df_train = pd.read_csv(train_fp)
    df_test  = pd.read_csv(test_fp)

    # Compute MAEs directly
    mae_train = mean_absolute_error(df_train['Y_true'], df_train['Y_pred'])
    mae_test  = mean_absolute_error(df_test['Y_true'],  df_test['Y_pred'])

    # Create scatter plot
    plt.figure(figsize=(6,6))
    plt.scatter(df_train['Y_true'], df_train['Y_pred'],
                color='tab:blue', label=f'Train (MAE = {mae_train:.3f} eV)', alpha=0.6, linestyle='--')
    plt.scatter(df_test['Y_true'], df_test['Y_pred'],
                color='tab:red',  label=f'Test  (MAE = {mae_test:.3f} eV)',  alpha=0.6)

    # y = x reference line
    all_vals = pd.concat([df_train[['Y_true','Y_pred']],
                          df_test [['Y_true','Y_pred']]]).values.flatten()
    vmin, vmax = all_vals.min(), all_vals.max()
    plt.plot([vmin, vmax], [vmin, vmax], 'k--', lw=1)

    # # Annotate MAE values
    # text_x = vmin + 0.85*(vmax-vmin)
    # text_y = vmax - 0.85*(vmax-vmin)
    # plt.text(text_x, text_y,
    #          f"MAE_train: {mae_train:.3f} eV\nMAE_test:  {mae_test:.3f} eV",
    #          fontsize=12, bbox=dict(facecolor='white', alpha=0.8))

    # Labels and title
    plt.xlabel('True Eads',  fontsize=14)
    plt.ylabel('Predicted Eads', fontsize=14)
    plt.title(f'Iteration {ratio} Predictions (EF = {EF})', fontsize=16)
    plt.legend(framealpha=0.8, fontsize=12)
    plt.tight_layout()

    # Save figure
    out_png = os.path.join(predict_csv_dir, f'preds_iter{ratio}_{EF}.png')
    plt.savefig(out_png, dpi=300)
    plt.close()
    print(f"✅ Saved scatter plot with MAEs to {out_png}")
    
    # === Additional figure: ALL predictions (train + test) ===
    all_fp = os.path.join(predict_csv_dir, f'all_preds_iter{ratio}_{EF}.csv')
    if os.path.exists(all_fp):
        df_all = pd.read_csv(all_fp)
    
        # Compute MAE for ALL directly from file
        mae_all = mean_absolute_error(df_all['Y_true_all'], df_all['Y_pred_all'])
    
        # Create scatter: ALL
        plt.figure(figsize=(6,6))
        plt.scatter(
            df_all['Y_true_all'], df_all['Y_pred_all'],
            alpha=0.6, label=f'All (MAE = {mae_all:.3f} eV)'
        )
    
        # y = x reference line
        vmin_all = min(df_all['Y_true_all'].min(), df_all['Y_pred_all'].min())
        vmax_all = max(df_all['Y_true_all'].max(), df_all['Y_pred_all'].max())
        plt.plot([vmin_all, vmax_all], [vmin_all, vmax_all], 'k--', lw=1)
    
        # Labels and title
        plt.xlabel('True Eads', fontsize=14)
        plt.ylabel('Predicted Eads', fontsize=14)
        plt.title(f'Iteration {ratio} ALL Predictions (EF = {EF})', fontsize=16)
        plt.legend(framealpha=0.8, fontsize=12)
        plt.tight_layout()
    
        # Save figure
        out_png_all = os.path.join(predict_csv_dir, f'all_preds_iter{ratio}_{EF}.png')
        plt.savefig(out_png_all, dpi=300)
        plt.close()
        print(f"✅ Saved ALL scatter plot with MAE to {out_png_all}")
    else:
        print(f"⚠️ ALL predictions file not found: {all_fp}")
    
def run_or_plot_active_learning(models, model_name, EF, X, Y, sites,
                                initial_ratio=0.4, increment=0.05, max_ratio=0.95,
                                results_dir='results'):
    """
    Check for an existing overall CSV file with active learning results.
    If found, read and plot the results. Otherwise, run the active learning process,
    save the merged results to one CSV file, and then plot the results.
    
    Args:
        models (dict): Dictionary of models.
        model_name (str): Name of the model to use.
        EF: An identifier for saving results.
        X (np.array): Features dataset.
        Y (np.array): Target values.
        sites (np.array or list): Labels corresponding to each sample (e.g., the 'site' column).
        initial_ratio (float): Initial fraction of data for training.
        increment (float): Fraction of total data to add each iteration.
        max_ratio (float): Maximum fraction of data for training.
        results_dir (str): Directory where results will be saved.
    """
    overall_csv_file = os.path.join(results_dir, f'active_learning_metrics_{EF}.csv')
    
    if os.path.exists(overall_csv_file):
        print(f"CSV file found at {overall_csv_file}. Reading and plotting results...")
        plot_active_learning_from_csv(overall_csv_file, EF)
    else:
        print("CSV file not found, running active learning process...")
        active_learning_one_model(models, model_name, EF, X, Y, sites,
                                  initial_ratio=initial_ratio, increment=increment, 
                                  max_ratio=max_ratio, results_dir=results_dir)
        print("Active learning process completed. Plotting results...")
        plot_active_learning_from_csv(overall_csv_file, EF)

    # plot_active_predicting_from_csv('results', EF)
    iterations = [1, 2, 3, 4, 5, 6, 7]
    for l in iterations:
        plot_active_predicting_from_csv('results', l, EF)
### GA Applilcations
    
def find_all_rhombuses(atoms, connections, surface_indices, bond_length_threshold):
    # atoms, connections, cn_of_connected_atoms, exposed_top_sites = get_connection(path, metal='Ru', mult=0.9)
    """Find all possible rhombuses based on atom connections"""
    rhombuses = []
    for A_index in surface_indices:
        A_neighbors = connections[A_index]
        for B_index in A_neighbors:
            if B_index not in surface_indices or A_index == B_index:
                continue
            if np.linalg.norm(atoms[A_index].position - atoms[B_index].position) > bond_length_threshold:
                continue

            B_neighbors = connections[B_index]
            for C_index in B_neighbors:
                if C_index not in surface_indices or C_index in [A_index, B_index]:
                    continue
                if np.linalg.norm(atoms[B_index].position - atoms[C_index].position) > bond_length_threshold:
                    continue

                C_neighbors = connections[C_index]
                for D_index in C_neighbors:
                    if D_index not in surface_indices or D_index in [A_index, B_index, C_index]:
                        continue
                    if np.linalg.norm(atoms[C_index].position - atoms[D_index].position) > bond_length_threshold:
                        continue
                    if np.linalg.norm(atoms[D_index].position - atoms[A_index].position) > bond_length_threshold:
                        continue

                    # Save rhombus indices while maintaining the original order
                    rhombus_indices = [A_index, B_index, C_index, D_index]
                    rhombuses.append(rhombus_indices)

    return rhombuses


# ------------------ Site Conversion ------------------
def convert_sites(sites_list):
    """
    Given a list of four site numbers corresponding to atoms A, B, C, and D (in clockwise order),
    return a dictionary mapping standard site labels to their numeric string representations.
    
    For example, if sites_list = [20, 23, 24, 42]:
      - "top_A": "20"
      - "top_B": "23"
      - "top_C": "24"
      - "top_D": "42"
      - "bridge_A-B": "20-23"
      - "bridge_B-C": "23-24"
      - "bridge_C-D": "24-42"
      - "bridge_D-A": "20-42"  (always smaller-first)
      - "hollow_ABD": "20-23-42"  (sorted ascending among A, B, D)
      - "hollow_BCD": "23-24-42"  (sorted ascending among B, C, D)
    """
    if len(sites_list) != 4:
        raise ValueError("sites_list must contain exactly four elements corresponding to A, B, C, and D.")
    sites_list = [i + 1 for i in sites_list]
    A, B, C, D = sites_list
    site_dict = {}
    site_dict["top_A"] = str(A)
    site_dict["top_B"] = str(B)
    site_dict["top_C"] = str(C)
    site_dict["top_D"] = str(D)
    site_dict["bridge_A-B"] = f"{min(A, B)}_{max(A, B)}"
    site_dict["bridge_B-C"] = f"{min(B, C)}_{max(B, C)}"
    site_dict["bridge_C-D"] = f"{min(C, D)}_{max(C, D)}"
    site_dict["bridge_D-A"] = f"{min(D, A)}_{max(D, A)}"
    site_dict["hollow_ABD"] = "_".join(str(x) for x in sorted([A, B, D]))
    site_dict["hollow_BCD"] = "_".join(str(x) for x in sorted([B, C, D]))
    # print(site_dict)
    return site_dict

def convert_sites_triangle_site(sites_list):
    """
    Given a list of three site numbers corresponding to atoms A, B, C(in clockwise order),
    return a dictionary mapping standard site labels to their numeric string representations.
    
    For example, if sites_list = [20, 23, 24]:
      - "top_A": "20"
      - "top_B": "23"
      - "top_C": "24"
      - "bridge_A-B": "20-23"
      - "bridge_B-C": "23-24"
      - "bridge_A-C": "20-24"
      - "hollow_ABC": "20-23-24"  (sorted ascending among B, C, D)
    """
    if len(sites_list) != 3:
        raise ValueError("sites_list must contain exactly three elements corresponding to A, B, C.")
    sites_list = [i + 1 for i in sites_list]
    A, B, C = sites_list
    site_dict = {}
    site_dict["top_A"] = str(A)
    site_dict["top_B"] = str(B)
    site_dict["top_C"] = str(C)
    site_dict["bridge_A-B"] = f"{min(A, B)}_{max(A, B)}"
    site_dict["bridge_B-C"] = f"{min(B, C)}_{max(B, C)}"
    site_dict["bridge_A-C"] = f"{min(A, C)}_{max(A, C)}"
    site_dict["hollow_ABC"] = "_".join(str(x) for x in sorted([A, B, C]))
    return site_dict

# ------------------ Modified Helper Function ------------------
def select_best_site(cluster_path, species, sites, Prop, site_mapping=None):
    """
    Evaluate the predicted adsorption energy for a given species at each candidate site.
    If site_mapping is provided, convert the candidate label to its numeric string.
    
    Args:
        cluster_path (str): Path to the cluster data.
        species (str): The adsorbate species (e.g., "NH3", "NH2", "NH", "N", "H", "N2").
        sites (list): List of candidate site labels (e.g., "top_A", "bridge_A-B", etc.).
        Prop: Additional properties required by predict_Eads_site.
        site_mapping (dict, optional): Mapping of candidate labels to numeric strings.
    
    Returns:
        tuple: (best_site, energy_dict) where energy_dict is keyed by the candidate (symbolic label).
    """
    energy_dict = {}
    for site in sites:
        candidate = str(site_mapping[site]) if (site_mapping is not None and site in site_mapping) else site
        Eads = predict_Eads_site(cluster_path, species, candidate, Prop)
        energy_dict[site] = Eads
        # print(f"{species} at site {site} ({candidate}): Eads = {Eads:.3f} eV")
    best_site = min(energy_dict, key=energy_dict.get)
    # print(f"Best {species} site: {best_site} with Eads = {energy_dict[best_site]:.3f} eV\n")
    return best_site, energy_dict

# ------------------ Candidate Mappings ------------------

# (1) For NH3 adsorption:
nh3_candidates = ["top_A", "top_B", "top_C", "top_D"]

# (2) For NH3 → NH2 + H:
def get_nh2_candidates(nh3_site):
    mapping = {
        "top_A": ["top_A", "bridge_A-B", "bridge_D-A", "hollow_ABD"],
        "top_B": ["top_B", "bridge_A-B", "bridge_B-C", "hollow_ABD", "hollow_BCD"],
        "top_C": ["top_C", "bridge_B-C", "bridge_C-D", "hollow_BCD"],
        "top_D": ["top_D", "bridge_C-D", "bridge_D-A", "hollow_ABD", "hollow_BCD"]
    }
    return mapping.get(nh3_site, [])

# For H in Step 2 (H determined by NH2):
nh2_to_h_candidates_step2 = {
    "top_A":     ["top_B", "top_D", "hollow_ABD", "hollow_BCD"],
    "top_B":     ["top_A", "top_C", "hollow_ABD", "hollow_BCD"],
    "top_C":     ["top_B", "top_D", "hollow_ABD", "hollow_BCD"],
    "top_D":     ["top_A", "top_C", "hollow_ABD", "hollow_BCD"],
    "bridge_A-B": ["top_A", "top_B", "bridge_A-B", "hollow_ABD"],
    "bridge_B-C": ["top_B", "top_C", "bridge_B-C", "hollow_BCD"],
    "bridge_C-D": ["top_C", "top_D", "bridge_C-D", "hollow_BCD"],
    "bridge_D-A": ["top_D", "top_A", "bridge_D-A", "hollow_ABD"],
    "hollow_ABD": ["hollow_ABD", "top_A", "top_B", "top_D", "bridge_A-B", "bridge_D-A"],
    "hollow_BCD": ["hollow_BCD", "top_B", "top_C", "top_D", "bridge_B-C", "bridge_C-D"]
}

# (3) For NH2 → NH + H:
nh2_to_nh_candidates = {
    "top_A":     ["top_A", "bridge_A-B", "bridge_D-A", "hollow_ABD"],
    "top_B":     ["top_B", "bridge_A-B", "bridge_B-C", "hollow_ABD", "hollow_BCD"],
    "top_C":     ["top_C", "bridge_B-C", "bridge_C-D", "hollow_BCD"],
    "top_D":     ["top_D", "bridge_C-D", "bridge_D-A", "hollow_ABD", "hollow_BCD"],
    "bridge_A-B": ["top_A", "top_B", "bridge_A-B", "hollow_ABD"],
    "bridge_B-C": ["top_B", "top_C", "bridge_B-C", "bridge_A-B", "hollow_BCD"],
    "bridge_C-D": ["top_C", "top_D", "bridge_C-D", "hollow_BCD"],
    "bridge_D-A": ["top_D", "top_A", "bridge_D-A", "hollow_ABD"],
    "hollow_ABD": ["hollow_ABD", "top_A", "top_B", "top_D", "bridge_A-B", "bridge_D-A"],
    "hollow_BCD": ["hollow_BCD", "top_B", "top_C", "top_D", "bridge_B-C", "bridge_C-D"]
}

# For H in Step 3 (H determined by NH):
nh_to_h_candidates = {
    "top_A":     ["top_B", "top_D", "top_C", "hollow_ABD", "hollow_BCD"],
    "top_B":     ["top_A", "top_C", "top_D", "bridge_A-B", "bridge_B-C", "bridge_D-A", "bridge_C-D", "hollow_ABD", "hollow_BCD"],
    "top_C":     ["top_B", "top_D", "top_A", "bridge_B-C", "bridge_C-D", "bridge_A-B", "bridge_D-A", "hollow_BCD", "hollow_ABD"],
    "top_D":     ["top_A", "top_B", "top_C", "bridge_C-D", "bridge_D-A", "bridge_A-B", "bridge_B-C", "hollow_ABD", "hollow_BCD"],
    "bridge_A-B": ["top_C", "top_D", "bridge_B-C", "hollow_ABD", "hollow_BCD"],
    "bridge_B-C": ["top_A", "top_D", "bridge_D-A", "bridge_A-B", "bridge_C-D", "hollow_ABD", "hollow_BCD"],
    "bridge_C-D": ["top_A", "top_B", "bridge_A-B", "bridge_B-C", "bridge_D-A", "hollow_ABD", "hollow_BCD"],
    "bridge_D-A": ["top_B", "top_C", "bridge_B-C", "bridge_A-B", "bridge_C-D", "hollow_ABD", "hollow_BCD"],
    "hollow_ABD": ["hollow_BCD", "top_A", "top_B", "top_C", "top_D"],
    "hollow_BCD": ["hollow_ABD", "top_A", "top_B", "top_C", "top_D"]
}

# (4) For NH → N + H:
nh_to_n_candidates = {
    "top_A":     ["top_A", "bridge_A-B", "bridge_D-A", "hollow_ABD"],
    "top_B":     ["top_B", "bridge_A-B", "bridge_B-C", "hollow_ABD", "hollow_BCD"],
    "top_C":     ["top_C", "bridge_B-C", "bridge_C-D", "hollow_BCD"],
    "top_D":     ["top_D", "bridge_C-D", "bridge_D-A", "hollow_ABD", "hollow_BCD"],
    "bridge_A-B": ["top_A", "top_B", "bridge_A-B", "hollow_ABD"],
    "bridge_B-C": ["top_B", "top_C", "bridge_B-C", "bridge_A-B", "hollow_BCD"],
    "bridge_C-D": ["top_C", "top_D", "bridge_C-D", "hollow_BCD"],
    "bridge_D-A": ["top_D", "top_A", "bridge_D-A", "hollow_ABD"],
    "hollow_ABD": ["hollow_ABD", "top_A", "top_B", "top_D", "bridge_A-B", "bridge_D-A"],
    "hollow_BCD": ["hollow_BCD", "top_B", "top_C", "top_D", "bridge_B-C", "bridge_C-D"]
}
# For H in Step 4, reuse nh_to_h_candidates.

# ------------------ Reaction Chain Functions ------------------

# Step 1: NH₃ Adsorption
def determine_NH3_configuration(cluster_path, Prop, site_mapping=None):
    best_NH3_site, energies = select_best_site(cluster_path, "NH3", nh3_candidates, Prop, site_mapping)
    return {"site": best_NH3_site, "energy": energies[best_NH3_site]}

# Step 2: NH₃ → NH₂ + H (H determined by NH2)
def determine_NH2_configuration(cluster_path, Prop, best_NH3_site, site_mapping=None):
    nh2_candidates = get_nh2_candidates(best_NH3_site)
    # print('Good1',nh2_candidates )
    best_NH2_site, energies_NH2 = select_best_site(cluster_path, "NH2", nh2_candidates, Prop, site_mapping)
    candidate_H_sites = nh2_to_h_candidates_step2.get(best_NH2_site, [])
    if candidate_H_sites:
        best_H_site, energies_H = select_best_site(cluster_path, "H", candidate_H_sites, Prop, site_mapping)
    else:
        best_H_site, energies_H = None, {}
    return {
        "NH2": {"site": best_NH2_site, "energy": energies_NH2[best_NH2_site]},
        "H1": {"site": best_H_site, "energy": energies_H.get(best_H_site, None)}
    }

# Step 3: NH₂ → NH + H (H determined by NH)
def determine_NH_configuration(cluster_path, Prop, best_NH2_site, site_mapping=None):
    candidate_NH_sites = nh2_to_nh_candidates.get(best_NH2_site, [])
    best_NH_site, energies_NH = select_best_site(cluster_path, "NH", candidate_NH_sites, Prop, site_mapping)
    candidate_H_sites = nh_to_h_candidates.get(best_NH_site, [])
    if candidate_H_sites:
        best_H_site, energies_H = select_best_site(cluster_path, "H", candidate_H_sites, Prop, site_mapping)
    else:
        best_H_site, energies_H = None, {}
    return {
        "NH": {"site": best_NH_site, "energy": energies_NH[best_NH_site]},
        "H2": {"site": best_H_site, "energy": energies_H.get(best_H_site, None)}
    }

# Step 4: NH → N + H (H determined by N)
def determine_N_configuration(cluster_path, Prop, best_NH_site, site_mapping=None):
    candidate_N_sites = nh_to_n_candidates.get(best_NH_site, [])
    if candidate_N_sites:
        best_N_site, energies_N = select_best_site(cluster_path, "N", candidate_N_sites, Prop, site_mapping)
    else:
        best_N_site, energies_N = None, {}
    candidate_H_sites = nh_to_h_candidates.get(best_N_site, [])
    if candidate_H_sites:
        best_H_site, energies_H = select_best_site(cluster_path, "H", candidate_H_sites, Prop, site_mapping)
    else:
        best_H_site, energies_H = None, {}
    return {
        "N": {"site": best_N_site, "energy": energies_N.get(best_N_site, None)},
        "H3": {"site": best_H_site, "energy": energies_H.get(best_H_site, None)}
    }

# Step 5: N₂ Adsorption
def determine_N2_configuration(cluster_path, Prop, site_mapping=None):
    candidate_sites = [
        "top_A", "top_B", "top_C", "top_D",
        "bridge_A-B", "bridge_B-C", "bridge_C-D", "bridge_D-A",
        "hollow_ABD", "hollow_BCD"
    ]
    best_N2_site, energies_N2 = select_best_site(cluster_path, "N2", candidate_sites, Prop, site_mapping)
    return {"site": best_N2_site, "energy": energies_N2[best_N2_site]}

# ------------------ Top-Level Full Configuration ------------------
def determine_full_configuration(cluster_path, Prop, sites_list=None):
    """
    Determine the full stable configuration for the reaction chain and return a dictionary
    where each species is mapped to a tuple (site, energy):
      1. NH3 → NH2 + H      (H determined by NH2)
      2. NH2 → NH + H       (H determined by NH)
      3. NH  → N + H        (H determined by N)
      4. Also determine the N2 adsorption site.
    
    If sites_list is provided (a list of four numbers for atoms A, B, C, D in clockwise order),
    it is converted to a mapping for energy predictions.
    
    Returns a dictionary with keys:
      "NH3", "NH2", "NH", "N", "N2", "H1", "H2", and "H3".
    """
    site_mapping = convert_sites(sites_list) if sites_list is not None else None
    # print('site_mapping', site_mapping)
    config_NH3 = determine_NH3_configuration(cluster_path, Prop, site_mapping)
    config_NH2 = determine_NH2_configuration(cluster_path, Prop, config_NH3["site"], site_mapping)
    config_NH = determine_NH_configuration(cluster_path, Prop, config_NH2["NH2"]["site"], site_mapping)
    config_N = determine_N_configuration(cluster_path, Prop, config_NH["NH"]["site"], site_mapping)
    config_N2 = determine_N2_configuration(cluster_path, Prop, site_mapping)

    final_config = {
        "NH3": (config_NH3["site"], config_NH3["energy"]),
        "NH2": (config_NH2["NH2"]["site"], config_NH2["NH2"]["energy"]),
        "NH":  (config_NH["NH"]["site"], config_NH["NH"]["energy"]),
        "N":   (config_N["N"]["site"], config_N["N"]["energy"]),
        "N2":  (config_N2["site"], config_N2["energy"]),
        "H1":  (config_NH2["H1"]["site"], config_NH2["H1"]["energy"]),
        "H2":  (config_NH["H2"]["site"], config_NH["H2"]["energy"]),
        "H3":  (config_N["H3"]["site"], config_N["H3"]["energy"])
    }
    
    return final_config


### For triangle sites 
# Step 1: NH₃ Adsorption
nh3_candidates_tri = ["top_A", "top_B", "top_C"]
nh2_candidates_tri = ["top_A", "top_B", "top_C", "bridge_A-B", "bridge_B-C", "bridge_A-C"]
nh_candidates_tri = ["top_A", "top_B", "top_C", "bridge_A-B", "bridge_B-C", "bridge_A-C", "hollow_ABC" ]
n_candidates_tri =  nh_candidates_tri 
h_candidates_tri =  nh_candidates_tri 
n2_candidates_tri = nh_candidates_tri 
def determine_NH3_configuration_tri(cluster_path, Prop, site_mapping=None):
    best_NH3_site, energies = select_best_site(cluster_path, "NH3", nh3_candidates_tri, Prop, site_mapping)
    return {"site": best_NH3_site, "energy": energies[best_NH3_site]}
def determine_NH2_configuration_tri(cluster_path, Prop, site_mapping=None):
    best_NH2_site, energies_NH2 = select_best_site(cluster_path, "NH2", nh2_candidates_tri, Prop, site_mapping)
    return {"NH2": {"site": best_NH2_site, "energy": energies_NH2[best_NH2_site]}}
def determine_H_configuration_tri(cluster_path, Prop, site_mapping=None):    
    best_H_site, energies_H = select_best_site(cluster_path, "H", h_candidates_tri, Prop, site_mapping)
    return {"H": {"site": best_H_site, "energy": energies_H.get(best_H_site, None)}}
def determine_NH_configuration_tri(cluster_path, Prop, site_mapping=None):
    best_NH_site, energies_NH = select_best_site(cluster_path, "NH", nh_candidates_tri, Prop, site_mapping)
    return {"NH": {"site": best_NH_site, "energy": energies_NH[best_NH_site]}}
def determine_N_configuration_tri(cluster_path, Prop, site_mapping=None):
    best_N_site, energies_N = select_best_site(cluster_path, "N", n_candidates_tri, Prop, site_mapping)
    return {"N": {"site": best_N_site, "energy": energies_N.get(best_N_site, None)}}
def determine_N2_configuration_tri(cluster_path, Prop, site_mapping=None):
    best_N2_site, energies_N2 = select_best_site(cluster_path, "N2", n2_candidates_tri, Prop, site_mapping)
    return {"site": best_N2_site, "energy": energies_N2[best_N2_site]}

def determine_full_configuration_triangle_site(cluster_path, Prop, sites_list=None):
    """
    Determine the full stable configuration for the reaction chain and return a dictionary
    where each species is mapped to a tuple (site, energy):
      1. NH3 → NH2 + H      (H determined by NH2)
      2. NH2 → NH + H       (H determined by NH)
      3. NH  → N + H        (H determined by N)
      4. Also determine the N2 adsorption site.
    
    If sites_list is provided (a list of four numbers for atoms A, B, C, D in clockwise order),
    it is converted to a mapping for energy predictions.
    
    Returns a dictionary with keys:
      "NH3", "NH2", "NH", "N", "N2", "H1", "H2", and "H3".
    """
    site_mapping = convert_sites_triangle_site(sites_list) 
    config_NH3 = determine_NH3_configuration_tri(cluster_path, Prop, site_mapping)
    config_NH2 = determine_NH2_configuration_tri(cluster_path, Prop, site_mapping)
    config_NH = determine_NH_configuration_tri(cluster_path, Prop, site_mapping)
    config_N = determine_N_configuration_tri(cluster_path, Prop,  site_mapping)
    config_H = determine_H_configuration_tri(cluster_path, Prop,  site_mapping)
    config_N2 = determine_N2_configuration_tri(cluster_path, Prop, site_mapping)

    final_config = {
        "NH3": (config_NH3["site"], config_NH3["energy"]),
        "NH2": (config_NH2["NH2"]["site"], config_NH2["NH2"]["energy"]),
        "NH":  (config_NH["NH"]["site"], config_NH["NH"]["energy"]),
        "N":   (config_N["N"]["site"], config_N["N"]["energy"]),
        "N2":  (config_N2["site"], config_N2["energy"]),
        "H":   (config_H["H"]["site"], config_H["H"]["energy"]),
    }
    
    return final_config


def compute_EDFT(ef_value, final_config, gas_dict):
    """
    Compute the DFT simulated energy (EDFT) for each species using the following formulas:
      - EDFT(NH3) = Eads(NH3) + E(slab) + E(NH3_gas)
      - EDFT(NH2) = Eads(NH2) + E(slab) + E(NH2_gas)
      - EDFT(NH)  = Eads(NH)  + E(slab) + E(NH_gas)
      - For N, H1, H2, and H3, use:
            EDFT = Eads + E(slab) + 0.5 * E(H2_gas)
      - EDFT(N2)  = Eads(N2)  + E(slab) + E(N2_gas)

    Returns a dictionary:
      {
         "NH3": (site, Eads, EDFT),
         ...
      }
    """
    ef_value = float(ef_value)
    E_slab  = -6.2414 * ef_value**2 - 0.01405 * ef_value - 354.4197 
    
    edft = {}

    # Common pattern
    edft["NH3"] = (
        final_config["NH3"][0],
        final_config["NH3"][1],
        final_config["NH3"][1] + E_slab + gas_dict["NH3"]
    )

    edft["NH2"] = (
        final_config["NH2"][0],
        final_config["NH2"][1],
        final_config["NH2"][1] + E_slab + gas_dict["NH2"]
    )

    edft["NH"] = (
        final_config["NH"][0],
        final_config["NH"][1],
        final_config["NH"][1] + E_slab + gas_dict["NH"]
    )

    edft["N"] = (
        final_config["N"][0],
        final_config["N"][1],
        final_config["N"][1] + E_slab + 0.5 * gas_dict["N2"]
    )


    edft["H"] = (
        final_config["H"][0],
        final_config["H"][1],
        final_config["H"][1] + E_slab + 0.5 * gas_dict["H2"]
        )

    edft["N2"] = (
        final_config["N2"][0],
        final_config["N2"][1],
        final_config["N2"][1] + E_slab + gas_dict["N2"]
    )

    return edft, E_slab


def find_rhombus(atoms, top_list, B, D, bond_threshold=2.6):
    """
    Given a list of four atom indices (top_list) that form a quadrilateral in clockwise order,
    this function identifies a valid rhombus if:
      1) The pair of atoms with the longest distance is assigned as A and C (with the smaller index as A).
      2) The remaining two atoms are B and D (with the smaller index as B).
      3) The bonds A–B, B–C, C–D, and D–A all have distances less than or equal to bond_threshold.
         (In this design, B and D are assumed to be bonded, representing the shorter bridge.)
    
    Parameters:
      atoms: An object that supports atoms.get_distance(i, j) for distance between atoms.
      top_list: List of four atom indices.
      B, D: The two atoms chosen (from top_list) that form the short bridge (B–D).
      bond_threshold: Maximum allowed distance for a bond.
    
    Returns:
      A list [A, B, C, D] representing the valid rhombus in clockwise order,
      or an empty list if no valid rhombus is found.
    """
    # Compute the distance between B and D (the short bridge)
    BD_distance = atoms.get_distance(B, D)
    top_list_sorted = sorted(top_list)

    for A in top_list_sorted:
        if A == B or A == D:
            continue

        # Check whether A forms bonds with both B and D
        if atoms.get_distance(A, B) <= bond_threshold and atoms.get_distance(A, D) <= bond_threshold:
            for C in top_list_sorted:
                if C in (A, B, D):
                    continue

                # Check whether C forms bonds with both B and D
                if atoms.get_distance(C, B) <= bond_threshold and atoms.get_distance(C, D) <= bond_threshold:
                    AC_distance = atoms.get_distance(A, C)
                    # The longest distance should be between A and C.
                    if AC_distance > BD_distance:
                        return [A, B, C, D]
    return []


def dihedral_angle(points):
    """
    Calculate the dihedral angle (in degrees) between the plane defined by A, B, D
    and the plane defined by B, C, D.
    
    Parameters:
        points (array-like): A 4x3 array containing the positions of points A, B, C, and D.
    
    Returns:
        float: The dihedral angle in degrees.
    """
    A, B, C, D = points
    # Compute normals for the two planes
    n1 = np.cross(B - A, D - A)
    n1 /= np.linalg.norm(n1)
    n2 = np.cross(C - B, D - B)
    n2 /= np.linalg.norm(n2)
    # Compute the angle between the normals
    dot_val = np.clip(np.dot(n1, n2), -1.0, 1.0)
    angle = np.degrees(np.arccos(dot_val))
    return angle

def filter_rhombus_by_dihedral(rhombuses, atoms, dihedral_threshold=120):
    """
    Separate candidate rhombus configurations into planar and non-planar groups
    based on the dihedral angle between planes ABD and BCD.
    
    Parameters:
        rhombuses (list): List of candidate rhombus configurations (each is a list [A, B, C, D]).
        atoms: An ASE Atoms object (with .position and get_distance() method).
        dihedral_threshold (float): Angle (in degrees) above which a configuration is considered planar.
    
    Returns:
        tuple: (planar, non_planar)
            - planar: list of candidates with dihedral angle >= dihedral_threshold.
            - non_planar: list of candidates with dihedral angle < dihedral_threshold.
    """
    planar = []
    non_planar = []
    for indices in rhombuses:
        points = np.array([atoms[idx].position for idx in indices])
        angle = dihedral_angle(points)
        if angle >= dihedral_threshold:
            planar.append(indices)
        else:
            non_planar.append(indices)
    return planar, non_planar

def filter_rhombus_by_AC_BD(planar, non_planar, atoms, AC_cutoff=3.0):
    """
    From the candidate rhombus configurations (planar and non-planar), check if any candidate in the non-planar group
    has a very short A–C distance (less than AC_cutoff). If so, use the A–C pair (sorted) to find candidates
    that have this pair as their B–D bond. Those candidates are then removed from both the planar and non-planar lists
    and collected in an invalid_rhombus list.
    
    Parameters:
        planar (list): List of candidate rhombus configurations classified as planar.
        non_planar (list): List of candidate configurations classified as non-planar.
        atoms: An ASE Atoms object.
        AC_cutoff (float): Distance cutoff (in Å) for the A–C bond.
    
    Returns:
        tuple: (new_planar, new_non_planar, invalid_rhombus)
            - new_planar: Planar candidates after removal.
            - new_non_planar: Non-planar candidates after removal.
            - invalid_rhombus: Candidates removed because their A–C pair (from non-planar with short A–C)
                                is used as a B–D pair.
    """
    invalid_rhombus = []
    # Combine candidates from both groups
    all_candidates = planar + non_planar

    # Build a dictionary mapping each candidate's B–D pair (sorted) to candidate(s)
    bd_dict = {}
    for candidate in all_candidates:
        BD = tuple(sorted((candidate[1], candidate[3])))
        bd_dict.setdefault(BD, []).append(candidate)

    # For each candidate in the non-planar list, check its A–C distance.
    for candidate in non_planar:
        AC = tuple(sorted((candidate[0], candidate[2])))
        AC_distance = atoms.get_distance(candidate[0], candidate[2])
        if AC_distance < AC_cutoff:
            if AC in bd_dict:
                for cand in bd_dict[AC]:
                    if cand not in invalid_rhombus:
                        invalid_rhombus.append(cand)

    new_planar = [cand for cand in planar if cand not in invalid_rhombus]
    new_non_planar = [cand for cand in non_planar if cand not in invalid_rhombus]

    return new_planar, new_non_planar, invalid_rhombus


def point_in_prism(point, base_vertices, normal, height):
    """
    Check if a point is inside the prism defined by the base triangle and its top translated version.

    Parameters:
    - point (ndarray): The position of the atom being checked.
    - base_vertices (list): The three vertices of the base triangle.
    - normal (ndarray): The normal vector of the triangle.
    - height (float): The prism height limit.

    Returns:
    - bool: True if the point is inside the prism, False otherwise.
    """
    def is_point_in_triangle(p, v1, v2, v3):
        """Check if a point p lies inside a 2D projection of a triangle"""
        u = v2 - v1
        v = v3 - v1
        w = p - v1

        uu = np.dot(u, u)
        uv = np.dot(u, v)
        vv = np.dot(v, v)
        wu = np.dot(w, u)
        wv = np.dot(w, v)

        denom = uv * uv - uu * vv
        if abs(denom) < 1e-6:
            return False  # Degenerate triangle

        s = (uv * wv - vv * wu) / denom
        t = (uv * wu - uu * wv) / denom

        return (s >= 0) and (t >= 0) and (s + t <= 1)

    # Project point onto the base plane
    v0 = base_vertices[0]
    distance_to_base = np.dot(point - v0, normal)

    # **Check if point is between top and bottom planes**
    if not (-height <= distance_to_base <= height):
        return False

    # Project the point onto the base triangle's plane
    projected_point = point - distance_to_base * normal

    # **Check if projected point is inside the base triangle**
    return is_point_in_triangle(projected_point, *base_vertices)

def count_atoms_in_prism(atoms, indices, height):
    """
    Count the number of atoms within the prism formed by translating the base triangle
    along its normal vector by a given height.
    """
    pos = atoms.get_positions()

    # Get positions of the three triangle vertices
    p1, p2, p3 = pos[indices[0]], pos[indices[1]], pos[indices[2]]

    # Compute the normal vector
    normal = calculate_normal(p1, p2, p3)

    # **Translate base vertices to form the top and bottom triangles**
    top_triangle = [p1 + height * normal, p2 + height * normal, p3 + height * normal]
    bottom_triangle = [p1 - height * normal, p2 - height * normal, p3 - height * normal]

    # Count atoms in the top and bottom regions
    atom_count_top = 0
    atom_count_bottom = 0

    for i in range(len(pos)):
        if i in indices:  # Skip the three triangle atoms
            continue

        p = pos[i]

        # **Check if the atom is inside the top prism region**
        if point_in_prism(p, top_triangle, normal, height):
            print(f"Top Prism: {p}")
            atom_count_top += 1

        # **Check if the atom is inside the bottom prism region (now correctly defined)**
        if point_in_prism(p, bottom_triangle, normal, height):  
            atom_count_bottom += 1
            print(f"Bottom Prism: {p}")

    return atom_count_top, atom_count_bottom

def is_exposed_triangle(atoms, triangle_sites, height=2.2):
    """
    Check if a triangle site is exposed based on atom count in the prism region.

    Returns:
        True  -> Triangle is exposed
        False -> Triangle is not exposed (blocked)
    """
    atom_count_top, atom_count_bottom = count_atoms_in_prism(atoms, triangle_sites, height)

    if atom_count_top > 0 and atom_count_bottom > 0:
        print(atom_count_top, atom_count_bottom)
        return False 
    return True 

def is_exposed_rhombus(atoms, rhombus_indices, height=2.5):
    """
    Check if a rhombus formed by four atom indices is exposed.
    
    The function divides the rhombus (defined by indices [A, B, C, D]) into two triangles:
      - Triangle 1: [A, B, D]
      - Triangle 2: [B, C, D]
    
    For each triangle, it calls count_atoms_in_prism(atoms, triangle, height) to determine the number
    of atoms above (top) and below (bottom) the plane of the triangle. If, for both triangles, at least
    one triangle does NOT have atoms on both sides of its plane, the rhombus is considered exposed.
    
    Parameters:
        atoms: An ASE Atoms object (each atom has a .position attribute).
        rhombus_indices (list): A list of four atom indices [A, B, C, D] (in clockwise order).
        height (float): Height threshold (in Å) for the prism check.
    
    Returns:
        bool: True if the rhombus is exposed, False otherwise.
    """
    triangles = [
        [rhombus_indices[0], rhombus_indices[1], rhombus_indices[3]],  # Triangle ABD
        [rhombus_indices[1], rhombus_indices[2], rhombus_indices[3]]   # Triangle BCD
    ]
    
    results = []
    for triangle in triangles:
        atom_count_top, atom_count_bottom = count_atoms_in_prism(atoms, triangle, height)
        # If both top and bottom have atoms, then the triangle is not exposed.
        if atom_count_top > 0 and atom_count_bottom > 0:
            results.append(False)
        else:
            results.append(True)
    return all(results)


def filter_exposed_rhombus(candidates, atoms, height=2.5):
    """
    Filter candidate rhombus configurations based on their exposure.
    
    For each candidate (a list of four atom indices), this function uses is_exposed_rhombus()
    to determine whether the configuration is exposed.
    
    Parameters:
        candidates (list): List of candidate rhombus configurations (each is a list [A, B, C, D]).
        atoms: An ASE Atoms object.
        height (float): Height threshold (in Å) for exposure.
    
    Returns:
        tuple: (exposed_candidates, not_exposed_candidates)
            - exposed_candidates: List of candidates that are exposed.
            - not_exposed_candidates: List of candidates that are not exposed.
    """
    exposed = []
    not_exposed = []
    for candidate in candidates:
        if is_exposed_rhombus(atoms, candidate, height):
            exposed.append(candidate)
        else:
            not_exposed.append(candidate)
    return exposed, not_exposed


def filter_rhombus_by_exposure(planar, non_planar, atoms, height=2.5):
    """
    Given planar and non-planar candidate rhombus configurations, further filter them based on exposure.
    If a candidate is not exposed (i.e. is_exposed_rhombus() returns False), remove it from both groups
    and add it to the invalid_rhombus list.

    Parameters:
        planar (list): List of candidate rhombus configurations classified as planar.
        non_planar (list): List of candidate configurations classified as non-planar.
        atoms: An ASE Atoms object.
        height (float): Height threshold (in Å) for exposure.
    
    Returns:
        tuple: (planar_exposed, non_planar_exposed, invalid_rhombus)
            - planar_exposed: List of planar candidates that are exposed.
            - non_planar_exposed: List of non-planar candidates that are exposed.
            - invalid_rhombus: List of candidates that are not exposed.
    """
    exposed_planar, not_exposed_planar = filter_exposed_rhombus(planar, atoms, height)
    exposed_non_planar, not_exposed_non_planar = filter_exposed_rhombus(non_planar, atoms, height)
    invalid_rhombus = not_exposed_planar + not_exposed_non_planar
    return exposed_planar, exposed_non_planar, invalid_rhombus


def write_indices_to_file(file_path, indices_list):
    with open(file_path, 'w') as file:
        for indices in indices_list:
            text = ' '.join(map(str, indices))
            file.write(text + '\n')

def calculate_normal_vector_TOP(anchor_atom_index, cluster, cutoff):
    # Extract the position of the anchor atom
    anchor_atom = cluster[anchor_atom_index]
    anchor_position = anchor_atom.position
    
    # Find neighbors within the cutoff distance
    neighbors = []
    for neighbor in cluster:
        if not np.array_equal(neighbor.position, anchor_position):  # Ensure the neighbor is not the anchor atom
            vector = neighbor.position - anchor_position
            distance = np.linalg.norm(vector)
            if distance < cutoff:
                neighbors.append(neighbor.position)
    
    # Ensure we have enough neighbors to define a plane
    if len(neighbors) < 3:
        raise ValueError("Not enough neighbors to define a plane.")
    
    # Perform PCA on the neighbors to find the plane
    neighbors = np.array(neighbors)
    pca = PCA(n_components=3)
    pca.fit(neighbors)
    
    # The normal vector to the plane is the third principal component
    normal_vector = pca.components_[2]
    
    # Normalize the normal vector
    normal_vector /= np.linalg.norm(normal_vector)
    
    return normal_vector

def calculate_normal_vector_bri(atom_index, cluster, cutoff):
    vector_sum = np.zeros(3)
    atom = cluster[atom_index]

    for neighbor in cluster:
        if neighbor != atom:
            vector = atom.position - neighbor.position
            distance = np.linalg.norm(vector)
            if distance < cutoff and distance > 0:  # Avoid zero division
                vector_sum += vector / distance

    if np.linalg.norm(vector_sum) > 0:
        return vector_sum / np.linalg.norm(vector_sum)
    else:
        return np.array([0, 0, 1])  # Default to z-direction

def count_short_metal_nonmetal_bonds(cluster, metal='Ru', max_distance=1.6):
    """
    Count the number of bonds shorter than a specific distance between metal atoms and non-metal atoms in the cluster.

    Parameters:
    - cluster (ASE Atoms object): The cluster of atoms.
    - metal_symbol (str): Symbol of the metal atoms in the cluster.
    - max_distance (float): The maximum distance threshold for bond counting.

    Returns:
    - count (int): Number of bonds shorter than max_distance between metal atoms and non-metal atoms.
    """
    # Find indices of metal and non-metal atoms
    metal_indices = [i for i, atom in enumerate(cluster) if atom.symbol == metal]
    nonmetal_indices = [i for i, atom in enumerate(cluster) if atom.symbol != metal]

    count = 0

    # Calculate distances between metal and non-metal atoms
    for i in metal_indices:
        for j in nonmetal_indices:
            distance = cluster.get_distance(i, j, mic=True)
            if distance < max_distance:
                count += 1
    return count

def find_shortest_bond(atoms, cutoff=2.5):
    # read POSCAR
    # atoms = read(path)
    
    # get the index for N
    n_index = None
    ru_indices = []
    
    for i, atom in enumerate(atoms):
        if atom.symbol == 'N':
            n_index = i
        elif atom.symbol == 'Ru':
            ru_indices.append(i)

    if n_index is None:
        raise ValueError("No nitrogen (N) atom found in the POSCAR file.")  

    #Calculate the distance between N and each Ru atom, and filter out the distances that are less than the cutoff.
    distances = []

    for ru_index in ru_indices:
        distance = atoms.get_distance(n_index, ru_index, mic=True)
        if distance < cutoff:
            distances.append((ru_index, distance))

    
    ##"Sort by distance and find the shortest distance."
    if distances:
        distances.sort(key=lambda x: x[1])
        shortest_distance = distances[0]
        num_short_bonds = len(distances)
    else:
        shortest_distance = None
        num_short_bonds = 0

    return shortest_distance, num_short_bonds


##################=========================

def add_anchor_to_single_site_bri(poscar_path,exposed_indexs, metal='Ru', anchor_atom='N', output_prefix='POSCAR'):
    cluster = read(poscar_path)
    cutoff = 3.0  # Adjust based on your system
    modified_cluster = copy.deepcopy(cluster)
    def get_geo(modified_cluster, exposed_indexs):

        A = modified_cluster[exposed_indexs[0] ]  # Adjust index for 0-based indexing
        B = modified_cluster[exposed_indexs[1] ]  
        
        translation_vector = B.position - A.position
        # Calculate the normal vector for positioning the N atom
        normal_vector = calculate_normal_vector_bri(exposed_indexs[0] , modified_cluster, cutoff)
    
        # Position for N atom
        anchor_position = A.position + normal_vector * 1.6 +  0.5 * translation_vector  
        
        # Add N atom to the structure
        modified_cluster += Atoms(anchor_atom, positions=[anchor_position])
        return modified_cluster

    modified_cluster = copy.deepcopy(cluster)
    cluster_1 = get_geo(modified_cluster, exposed_indexs)
    num_bonds_1 = count_short_metal_nonmetal_bonds(cluster_1, metal, max_distance=1.6) 

    modified_cluster = copy.deepcopy(cluster)
    cluster_2 = get_geo(modified_cluster, exposed_indexs[::-1])
    
    num_bonds_2 = count_short_metal_nonmetal_bonds(cluster_2, metal, max_distance=1.6) 

    final_cluster = cluster_1
    if num_bonds_1 > num_bonds_2: 
        final_cluster = cluster_2 

    shortest_distance, num_short_bonds = find_shortest_bond(final_cluster, 2.5)
    # print(shortest_distance, num_short_bonds)
    if shortest_distance[1] > 1.5 and num_short_bonds < 3:     
        destination_path = os.path.join('bri', '_'.join([str(i) for i in exposed_indexs])) 
        os.makedirs(destination_path, exist_ok=True)
        output_file = destination_path + '/POSCAR'
    
        write(output_file, final_cluster, format='vasp', vasp5=True)

def add_one_bri(path, exposed_indices):
    poscar_path = path +'/POSCAR_bare'
    cluster = read(poscar_path)
    os.makedirs(path + '/bri', exist_ok=True)
    for combination in itertools.combinations(exposed_indices, 2):
        ## get the atoms that form the hollow site 
        sorted_combination = sorted(combination)
        bri_atom_indices = [i  for i in sorted_combination] # indice count from 0
        bri_atoms_positions = [cluster[i].position for i in bri_atom_indices]

        if not is_valid_bri(bri_atoms_positions):
            continue
        
        add_anchor_to_single_site_bri(poscar_path,sorted_combination,output_prefix='POSCAR')
 
def get_new_coordiates(cluster, A1, A2, B1, list_H):
    '''A1, A2, B1 are the coordinates, list_H is the list of coordinates '''
    # Calculate B2 as the center (midpoint) of C and D
    B2 = calculate_triangle_center(list_H) 
    translation_vector = A2 - B1
    B2_translated = B2 + translation_vector 
    
    # Step 2: Calculate vectors a and b
    a = A2 - A1
    a_normalized = a / np.linalg.norm(a)
    b = B2_translated - A2  # New b after moving B1 to A2
    b_normalized = b / np.linalg.norm(a)
    
    # Calculate rotation axis and angle
    axis = np.cross(b_normalized, a_normalized)
    
    axis_normalized = axis / np.linalg.norm(axis) if np.linalg.norm(axis) != 0 else b_normalized
    
    angle = np.arccos(np.clip(np.dot(b_normalized, a_normalized), -1.0, 1.0))     # Angle of rotation
    
    # Apply rotation around A2
    rotation = R.from_rotvec(axis_normalized * angle)
    
    new_list_H = [rotation.apply(i - B1) + A2 for i  in cluster.get_positions()[1:]]
    
    return new_list_H# C_final, D_final, E_final

def add_more_atoms(site):
    ''' Add species with more than three atoms to the anchor atoms. This function works for top, bri, and hollows sites '''
    sites = [int(i)  for i in site.split('_')]
    atoms = read(site + '/POSCAR')
    coordinates = [atoms[i].position for i in sites]
    A1 = calculate_triangle_center(coordinates)  # fcc site 
    A2 = atoms[-1].position     # N site, add N first and then put the NH,NH2,NH3 et al using the N site as reference.

    atoms_temp = read('../POSCAR_temp')
    B1 = atoms_temp[0].position

    # Initialize list_H to store atoms forming bonds with atoms_temp[0]
    list_H = []
    # Define bond distance threshold
    bond_distance = 1.9  # in Ångstroms
    positions = atoms_temp.get_positions()

    # Iterate over all atoms except the first one (atoms_temp[0])
    for i in range(1, len(positions)):
        atom_position = positions[i]
        distance = euclidean(B1, atom_position)
        
        if distance <= bond_distance:
            list_H.append(atom_position)
    
    # Convert list_H to numpy array if needed
    list_H = np.array(list_H)
    
    new_list_H  = get_new_coordiates(atoms_temp, A1, A2, B1, list_H)
    list_ele = atoms_temp.get_chemical_symbols()[1:]
    out_geo = copy.deepcopy(atoms)
    
    for num, coord in enumerate(new_list_H):
        out_geo += Atoms(list_ele[num], positions=[coord])
        
    write(site+'/POSCAR_bri', out_geo, format='vasp', vasp5=True)

def add_more_bri(path):
    """Add larger moelcules to all the bridge sites. The sites are the folders generated by the add_one_bri function"""
    os.chdir(path + '/bri')
    all_items = os.listdir('.')
    folders = [item for item in all_items if os.path.isdir(item)]
    folders = [folder for folder in folders if folder.count('_') == 1]
        
    for site in folders:     
        add_more_atoms(site)
        
def get_active_sites(cluster_path,metal = 'Ru'):   # Path: the cluster model directory
    poscar = os.path.join(cluster_path, 'POSCAR')
    atoms = read(poscar)
    bond_length_threshold = 2.7
    

    """ Step1: obtain the top sites""" 
    list_file = os.path.join(cluster_path, 'list')   ### list all the top sites in one line
    if not os.path.exists(list_file):
        # print("Warning: No list file found. Examine the exposed sites using Coordination Numbers.")
        top_indices = get_top_sites(atoms, metal, mult=0.9)  ## Seems that this method is not ideal
    else:
        with open(list_file, 'r') as f_in:
            top_indices = f_in.readline().rstrip().split()
            top_indices = [int(i) for i in top_indices]
            surface_indices = [i-1 for i in top_indices]
 
    """ Step2: obtain the bridge sites """     
    bridge_list = []
    for combination in itertools.combinations(top_indices, 2):
        ## get the atoms that form the hollow site 
        sorted_combination = sorted(combination)
        distance = atoms.get_distance(sorted_combination[0], sorted_combination[1])
        if distance < 2.6:
            bridge_list.append(sorted_combination)

    l_rhombus  = []
    for bridge_site in bridge_list: 
        B, D  =  bridge_site[:]
        indices = find_rhombus(atoms, top_indices, B, D, bond_threshold=3.0)
        l_rhombus.append(indices)
       
    l_rhombus = [item for item in l_rhombus if item] ## remove the empty 

    ############## Step3: categorize  rhombus sites: planar or edge
    planar, non_planar = filter_rhombus_by_dihedral(l_rhombus, atoms, dihedral_threshold=120)
    new_planar, new_non_planar, invalid_rhombus = filter_rhombus_by_AC_BD(planar, non_planar, atoms, AC_cutoff=3.0)
    # coplanar_threshold = 0.5  # Adjust threshold as needed, unit is Å
    # coplanar_rhombus = filter_coplanar_rhombuses(l_rhombus, atoms, coplanar_threshold)
    # edge_rhombus = filter_rhombuses_by_dihedral(l_rhombus, atoms, 40, 120)

    write_indices_to_file(os.path.join(cluster_path, 'all_sites.txt'), l_rhombus)
    write_indices_to_file(os.path.join(cluster_path, 'planar_sites.txt'), new_planar)
    write_indices_to_file(os.path.join(cluster_path, 'edge_sites.txt'), new_non_planar)
    write_indices_to_file(os.path.join(cluster_path, 'invalid_sites.txt'), invalid_rhombus)

    return l_rhombus, new_planar, new_non_planar

################## Section MKM

def get_k(Ea, T=673):
    kb = 8.617333262145E-5  # Boltzmann's constant in eV/K
    kbT = kb * T  # Boltzmann's constant times Temperature (in eV)
    A = kbT / (h / eV)  # Pre-exponential factor in s^-1, converting h to eV·s
    kf = A * math.exp(-Ea / kbT)  # Activation energy in eV, so no conversion needed
    return kf

def get_rate(kf0, kr0):
    # Given constants and conversion factors
    J2eV = 6.24150974E18            # eV/J
    Na = 6.0221415E23               # mol-1
    T = 823                         # Temperature in Kelvin
    h = 6.626068E-34 * J2eV         # Planck's constant in eV*s
    kb = 1.3806503E-23 * J2eV       # Boltzmann's constant in eV/K
    kbT = kb * T                    # Boltzmann's constant times Temperature
    kJ_mol2eV = 0.01036427
    wave2eV = 0.00012398
    FNH30 = 1                       # Inlet molar flow rate (assumed in mol/s)
    p0 = 1                          # Total pressure in atm
    # Function to calculate the partial pressures
    def get_pressures(FNH30, X, p0):
        # Calculating flow rates of NH3, N2, and H2
        FNH3 = FNH30 * (1 - X)
        FN2 = FNH30 * 0.5 * X
        FH2 = FNH30 * 1.5 * X
        
        # Total flow rate
        FT = FNH3 + FN2 + FH2
        
        # Calculating partial pressures
        pNH3 = (FNH3 / FT) * p0
        pN2 = (FN2 / FT) * p0
        pH2 = (FH2 / FT) * p0
        
        return pNH3, pN2, pH2

    # kf0 = [5631186144.388563, 50.01374140854966, 167.3703043694318, 35375.898901179375, 1.4277669985218864e-08, 10257.81417760215]
    # kr0 = [133466475.11618523, 78037.92020523513, 51662039.04911003, 526671805.90369725, 1.6185464638531345e-08, 66391314202.84105]
    
    # Function to get the rates
    def get_rates(theta, kf, kr, X, FNH30, p0):
        pNH3, pN2, pH2 = get_pressures(FNH30, X, p0)
        
        tNH3, tNH2, tNH, tN, tH = theta
        tstar = 1.0 - tNH3 - tNH2 - tNH - tN - tH  # Site balance for tstar
        
        rate = np.zeros(6)
        rate[0] = kf[0] * pNH3 * tstar - kr[0] * tNH3
        rate[1] = kf[1] * tNH3 * tstar - kr[1] * tNH2 * tH
        rate[2] = kf[2] * tNH2 * tstar - kr[2] * tNH * tH
        rate[3] = kf[3] * tNH * tstar - kr[3] * tN * tH
        rate[4] = kf[4] * tN ** 2 - kr[4] * pN2 * tstar
        rate[5] = kf[5] * tH ** 2 - kr[5] * pH2 * tstar
        return rate
    
    # Function to get the ODEs
    def get_odes(rate):
        dt = [0] * 5
        dt[0] = rate[0] - rate[1]                          # d(tNH3)/dt
        dt[1] = rate[1] - rate[2]                          # d(tNH2)/dt
        dt[2] = rate[2] - rate[3]                          # d(tNH)/dt
        dt[3] = rate[3] - 2 * rate[4]                      # d(tN)/dt
        dt[4] = rate[1] + rate[2] + rate[3] - 2 * rate[5]  # d(tH)/dt
        return dt
    
    def combined_odes(initial_state, t, kf, kr, FNH30, p0):
        X = initial_state[-1]
        theta = initial_state[:-1]
        
        rates = get_rates(theta, kf, kr, X, FNH30, p0)
        
        ds_dt = get_odes(rates)
        
        # Example improvement: Calculate dx_dt based on a specific rate
        # Assuming rate[5] corresponds to a key product formation step
        dx_dt = rates[4] / FNH30  # Normalize by inlet flow to represent conversion rate
        
        comb_odes = np.append(ds_dt, dx_dt)
        
        return comb_odes
    
    def solve_combined_odes(kf, kr, X0, FNH30, p0):
        thetas0 = np.zeros(5)
        initial_state = np.append(thetas0, X0)
        start_time = 0
        end_time = 1e6
        
        def combined_odes_wrapper(t, y):
            return combined_odes(y, t, kf, kr, FNH30, p0)
            
        # Solving the ODEs
        solution = solve_ivp(combined_odes_wrapper, [start_time, end_time], initial_state, method='LSODA', dense_output=True)
    
        t = np.linspace(start_time, end_time, 10000) 
        sol = solution.sol(t)
        thetas_solution = sol[:-1, :]
        X_solution = sol[-1, :]
            
        return thetas_solution, X_solution, t 
    
    # Solving for theta0 and rates
    X0 = 0.0
    # print('Running')
    thetas_t, X_t, t = solve_combined_odes(kf0, kr0, X0, FNH30, p0)
    # print('Done')
    rates = get_rates(thetas_t[:,-1], kf0, kr0, X_t[-1],  FNH30 , p0)
    tof = p0 *  X_t[-1]  / 10000 
    print('tof:', tof, 'lg(tof):', math.log10(tof), 'Conversion:', X_t[-1], 'rate_rds:',  min(rates)  )
    
    # print(thetas_t[:,-1],  X_t[-1])
    # print(math.log(p0 /100000 * X_t[-1]  / 1000000 ))
    # print("XNH3:", X_t[-1], )

    return rates



#### Construct adsorption configurations 


# Define bonding criteria (in Angstrom)
R_N_BOND_MAX = 2.4  # Max Ru-N bond distance
R_H_BOND_MAX = 2.2  # Max Ru-H bond distance
N_N_BOND_MAX = 1.4  # Max N-N bond distance for N2 detection

def get_bonded_atoms(atoms, atom_index, element, cutoff):
    """Find all atoms of a given element bonded to a specified atom."""
    bonded = []
    for i, atom in enumerate(atoms):
        if atom.symbol == element and i != atom_index:
            dist = atoms.get_distance(atom_index, i)
            if dist < cutoff:
                bonded.append(i)
    return bonded



def classify_single_atom_adsorption(atoms, element, bond_cutoff):
    """Classify adsorption of single atoms (N or H) based on Ru coordination."""
    atom_indices = [i for i, atom in enumerate(atoms) if atom.symbol == element]
    results = []

    for idx in atom_indices:
        ru_bonded = get_bonded_atoms(atoms, idx, "Ru", bond_cutoff)

        # Determine adsorption type
        if len(ru_bonded) == 1:
            adsorption_type = "top"
        elif len(ru_bonded) == 2:
            adsorption_type = "bridge"
        elif len(ru_bonded) == 3:
            adsorption_type = "hollow"
        else:
            adsorption_type = "unknown"

        ru_site_str = format_ru_sites(ru_bonded)
        results.append((idx, adsorption_type, ru_site_str))

    # return results
    return ru_site_str


def format_ru_sites(ru_indices):
    """Convert a list of Ru indices into a sorted string format."""
    ru_indices = [i + 1 for i in ru_indices]  # Convert to 1-based index
    ru_indices.sort()  # Sort in ascending order
    return "_".join(map(str, ru_indices)) if ru_indices else "N/A"

def is_valid_bridge3(atoms, n1, n2, ru_bonded_n1, ru_bonded_n2):
    """
    Check if N2 forms a Bridge-3 adsorption:
    - Both N atoms bind to the same two Ru atoms.
    - N2 bond is at an angle between 75°-90° relative to Ru-Ru bond.

    Returns:
        True  -> Valid Bridge-3
        False -> Not a Bridge-3
    """
    if set(ru_bonded_n1) != set(ru_bonded_n2):
        return False  # Ensure both N atoms are bonded to the same two Ru atoms

    ru1, ru2 = ru_bonded_n1  # Get the two Ru atoms

    # Get atomic positions
    n1_pos = atoms[n1].position
    n2_pos = atoms[n2].position
    ru1_pos = atoms[ru1].position
    ru2_pos = atoms[ru2].position

    # Compute N2 and Ru-Ru vectors
    n2_vector = n2_pos - n1_pos
    n2_vector /= np.linalg.norm(n2_vector)

    ru_vector = ru2_pos - ru1_pos
    ru_vector /= np.linalg.norm(ru_vector)

    # Compute angle in degrees
    angle_rad = np.arccos(np.clip(np.dot(n2_vector, ru_vector), -1.0, 1.0))
    angle_deg = np.degrees(angle_rad)

    return 75 <= angle_deg <= 115  # Allow angle range between 75°-90°



# def is_valid_fcc0(atoms, n1, n2, ru_bonded_n1, ru_bonded_n2):
#     """
#     Check if N2 forms an FCC-0 adsorption:
#     - N1 binds to two Ru atoms (A, B).
#     - N2 binds to a third, separate Ru atom (C), instead of one of (A, B).

#     Returns:
#         True  -> Valid FCC-0
#         False -> Not an FCC-0 site (likely Bridge-2)
#     """
#     if len(ru_bonded_n1) != 2 or len(ru_bonded_n2) != 2:
#         return False  # Ensure N1 and N2 each bind to exactly two Ru atoms

#     # Get the two Ru atoms bound to N1 (A, B)
#     ru_n1_set = set(ru_bonded_n1)

#     # Get the two Ru atoms bound to N2 (should include C)
#     ru_n2_set = set(ru_bonded_n2)

#     # **For Bridge-2: N2 must bind to both A and B**
#     if ru_n2_set.issubset(ru_n1_set):
#         return False

#     # **For FCC-0: N2 must bind to a third Ru (C) instead of both A and B**
#     return True

def is_valid_fcc0(atoms, n1, n2, ru_bonded_n1, ru_bonded_n2):
    """
    Check if N2 forms an FCC-0 adsorption:
    - One N (N1 or N2) binds to two Ru atoms (A, B).
    - The other N (N2 or N1) binds to a third, separate Ru atom (C), instead of both A and B.

    Returns:
        True  -> Valid FCC-0
        False -> Not an FCC-0 site (likely Bridge-2)
    """
    if len(ru_bonded_n1) != 2 or len(ru_bonded_n2) != 1:
        if len(ru_bonded_n1) != 1 or len(ru_bonded_n2) != 2:
            return False  # Ensure N1 binds to two Ru, and N2 binds to one OR vice versa

    # Get the two Ru atoms bound to N1 (A, B) and the single Ru bound to N2 (C)
    ru_n1_set = set(ru_bonded_n1)
    ru_n2_set = set(ru_bonded_n2)

    # **For FCC-0: N2 must bind to a separate Ru (C) not in (A, B)**
    return len(ru_n1_set.intersection(ru_n2_set)) == 0

def is_valid_rhombus_site(atoms, n1, n2, ru_bonded_n1, ru_bonded_n2):
    """
    Check if N2 forms a Rhombus adsorption:
    - N1 binds to three Ru atoms (A, B, C).
    - N2 binds to one of (A, B, C) and a new Ru atom (D).
    - Distinguishes from FCC-2 where N2 binds to two of (A, B, C).

    Returns:
        True  -> Valid Rhombus site
        False -> Not a Rhombus site (likely FCC-2)
    """
    if len(ru_bonded_n1) != 3 or len(ru_bonded_n2) != 2:
        return False  # Ensure correct Ru coordination

    # Get the three Ru atoms bound to N1 (A, B, C)
    ru_n1_set = set(ru_bonded_n1)

    # Get the two Ru atoms bound to N2
    ru_n2_set = set(ru_bonded_n2)

    # **Check if N2 binds to only one of (A, B, C) and one new Ru (D)**
    common_rus = ru_n1_set.intersection(ru_n2_set)
    unique_rus = ru_n2_set.difference(ru_n1_set)

    return len(common_rus) == 1 and len(unique_rus) == 1



def classify_N2_adsorption(atoms):
    """Classify N2 adsorption types and format Ru sites as strings."""
    nitrogen_indices = [i for i, atom in enumerate(atoms) if atom.symbol == "N"]
    # print('N2', nitrogen_indices)
    # Pair N atoms into N2 molecules based on N-N bonding distance
    n2_molecules = []
    visited = set()
    for i in nitrogen_indices:
        if i in visited:
            continue
        for j in nitrogen_indices:
            if i != j and atoms.get_distance(i, j) < N_N_BOND_MAX:
                n2_molecules.append((i, j))
                visited.add(i)
                visited.add(j)
                break

    results = []
    
    for n1, n2 in n2_molecules:
        
        ru_bonded_n1 = get_bonded_atoms(atoms, n1, "Ru", R_N_BOND_MAX)
        ru_bonded_n2 = get_bonded_atoms(atoms, n2, "Ru", R_N_BOND_MAX)
        
        # **Normalize Bond Counts** (Handle N1/N2 ordering issue)
        bond_counts = tuple(sorted([len(ru_bonded_n1), len(ru_bonded_n2)]))
        
        
        # **Check for Gas-Phase N2**
        if atoms.get_distance(n1, n2) < 1.3 and len(ru_bonded_n1) == 0 and len(ru_bonded_n2) == 0:
            adsorption_type = "gas"
            ru_site_str = "N/A"
        
        # **Classify Adsorption Type**
        elif bond_counts == (0, 1):
            adsorption_type = "top"
            ru_site_str = format_ru_sites(ru_bonded_n1 if len(ru_bonded_n1) == 1 else ru_bonded_n2)
        elif bond_counts == (1, 1) and ru_bonded_n1 == ru_bonded_n2:
            adsorption_type = "top"
            ru_site_str = format_ru_sites(ru_bonded_n1)  # One Ru index
        elif bond_counts == (1, 1):
            adsorption_type = "bridge-1"
            ru_site_str = format_ru_sites(list(set(ru_bonded_n1 + ru_bonded_n2)))  # Two Ru atoms
        
        elif bond_counts == (0, 2):
            # **Bridge-0: One N binds to two Ru atoms, the other is in the gas phase**
            adsorption_type = "bridge-0"
            ru_site_str = format_ru_sites(set(ru_bonded_n1 + ru_bonded_n2))
    
        elif bond_counts == (1, 2):
            # **Distinguish between FCC-0 and Bridge-2**
            if is_valid_fcc0(atoms, n1, n2, ru_bonded_n1, ru_bonded_n2):
                adsorption_type = "fcc-0"
            else:
                adsorption_type = "bridge-2"
            ru_site_str = format_ru_sites(set(ru_bonded_n1 + ru_bonded_n2))
        elif bond_counts == (2, 2):
            # **Bridge-3 detection (Both N bind to the same two Ru atoms & angle between 75°-90°)**
            if is_valid_bridge3(atoms, n1, n2, ru_bonded_n1, ru_bonded_n2):
                adsorption_type = "bridge-3"
                ru_site_str = format_ru_sites(set(ru_bonded_n1 + ru_bonded_n2))
            else:
                if len(set(ru_bonded_n1 + ru_bonded_n2)) == 3:
                    adsorption_type = "fcc-1"
                    ru_site_str = format_ru_sites(list(set(ru_bonded_n1 + ru_bonded_n2)))  # Three Ru atoms
                elif len(set(ru_bonded_n1 + ru_bonded_n2)) == 4:
                    adsorption_type = "trapezoid-1"
                    ru_site_str = format_ru_sites(list(set(ru_bonded_n1 + ru_bonded_n2)))  # Three Ru atoms
                    
                else:
                    adsorption_type = "unknown"
                    ru_site_str = "N/A"
        elif bond_counts == (2, 3):
            # **Distinguish between FCC-2 and Rhombus**
            if is_valid_rhombus_site(atoms, n1, n2, ru_bonded_n1, ru_bonded_n2):
                adsorption_type = "rhombus"
            else:
                adsorption_type = "fcc-2"
            ru_site_str = format_ru_sites(set(ru_bonded_n1 + ru_bonded_n2))  # Three Ru atoms
        elif bond_counts == (1, 3):
            adsorption_type = "fcc-3"
            ru_site_str = format_ru_sites(list(set(ru_bonded_n1 + ru_bonded_n2)))  # Three Ru atoms
        else:
            adsorption_type = "unknown"
            ru_site_str = "N/A"

        #results.append((n1, n2, adsorption_type, ru_site_str))
        results.append((adsorption_type, ru_site_str))

    return results   


def get_shortest_ru_n_distance(atoms):
    """Get the shortest Ru-N distance in the given structure."""
    ru_indices = [i for i, atom in enumerate(atoms) if atom.symbol == 'Ru']
    n_indices = [i for i, atom in enumerate(atoms) if atom.symbol == 'N']

    min_distance = float('inf')
    for ru in ru_indices:
        for n in n_indices:
            distance = atoms.get_distance(ru, n)
            if distance < min_distance:
                min_distance = distance

    return min_distance

def add_N2_top_sites(atoms, n_ru_distance=2.0, n_n_distance=1.19, metal='Ru'):
    """
    Add N2 molecules at top adsorption sites (Top1 and Top2).
    """

    connections, cn_of_connected_atoms, top_sites, bridge_sites, hollow_sites = get_connection(atoms, metal=metal, mult=0.9)
    mass_center = np.mean([atom.position for atom in atoms if atom.symbol == 'Ru'], axis=0)

    for i, ru_index in enumerate(top_sites):
        pos_top = atoms[ru_index].position  # Ru atom serving as the top site
        direction = pos_top - mass_center
        direction /= np.linalg.norm(direction)

        if np.dot(direction, pos_top - mass_center) < 0:
            direction = -direction  # Flip direction if needed

        ### **Top1 Placement**
        n1_position_top1 = pos_top + direction * n_ru_distance
        n2_position_top1 = n1_position_top1 + direction * n_n_distance
        n2_top1 = Atoms('N2', positions=[n1_position_top1, n2_position_top1])

        new_atoms_top1 = atoms.copy()
        new_atoms_top1.extend(n2_top1)

        if get_shortest_ru_n_distance(new_atoms_top1) < 1.5:
            write(f"POSCAR_top1_{i+1}_check1", new_atoms_top1, format='vasp')
            direction = -direction
            n1_position_top1 = pos_top + direction * n_ru_distance
            n2_position_top1 = n1_position_top1 + direction * n_n_distance
            n2_top1 = Atoms('N2', positions=[n1_position_top1, n2_position_top1])
            new_atoms_top1 = atoms.copy()
            new_atoms_top1.extend(n2_top1)

            if get_shortest_ru_n_distance(new_atoms_top1) < 1.5:
                write(f"POSCAR_top1_{i+1}_check2", new_atoms_top1, format='vasp')
            else:
                write(f"POSCAR_top1_{i+1}.vasp", new_atoms_top1, format='vasp')
        else:
            write(f"POSCAR_top1_{i+1}.vasp", new_atoms_top1, format='vasp')

        ### **Top2 Placement**
        nn_center = pos_top + direction * n_ru_distance
        perp_vector = np.cross(direction, [0, 0, 1])
        if np.linalg.norm(perp_vector) < 1e-6:
            perp_vector = np.cross(direction, [1, 0, 0])
        perp_vector /= np.linalg.norm(perp_vector)

        n1_position_top2 = nn_center - perp_vector * (n_n_distance / 2)
        n2_position_top2 = nn_center + perp_vector * (n_n_distance / 2)
        n2_top2 = Atoms('N2', positions=[n1_position_top2, n2_position_top2])

        new_atoms_top2 = atoms.copy()
        new_atoms_top2.extend(n2_top2)

        write(f"POSCAR_top2_{i+1}.vasp", new_atoms_top2, format='vasp')

def add_N2_bridge_sites(atoms, n_ru_distance=2.15, n_n_distance=1.19):
    """
    Add N2 molecules at bridge adsorption sites (Bridge-1, Bridge-2, Bridge-3, FCC-0).
    """

    connections, cn_of_connected_atoms, top_sites, bridge_sites, hollow_sites = get_connection(atoms, metal=metal, mult=0.9)
    mass_center = np.mean([atom.position for atom in atoms if atom.symbol == 'Ru'], axis=0)

    for i, (ru1, ru2) in enumerate(bridge_sites):
        pos1, pos2 = atoms[ru1].position, atoms[ru2].position
        bridge_mid = (pos1 + pos2) / 2
        direction = bridge_mid - mass_center
        direction /= np.linalg.norm(direction)

        ### **Bridge-1 Placement (Parallel to Ru-Ru)**
        nn_center = bridge_mid + direction * n_ru_distance
        ru_vector = pos2 - pos1
        ru_vector /= np.linalg.norm(ru_vector)

        n1_position_bridge1 = nn_center - ru_vector * (n_n_distance / 2)
        n2_position_bridge1 = nn_center + ru_vector * (n_n_distance / 2)
        n2_bridge1 = Atoms('N2', positions=[n1_position_bridge1, n2_position_bridge1])

        new_atoms_bridge1 = atoms.copy()
        new_atoms_bridge1.extend(n2_bridge1)

        if get_shortest_ru_n_distance(new_atoms_bridge1) < 1.5:
            write(f"POSCAR_bridge1_{i+1}_check1", new_atoms_bridge1, format='vasp')
            direction = -direction
            n1_position_bridge1 = nn_center - ru_vector * (n_n_distance / 2)
            n2_position_bridge1 = nn_center + ru_vector * (n_n_distance / 2)
            n2_bridge1 = Atoms('N2', positions=[n1_position_bridge1, n2_position_bridge1])
            new_atoms_bridge1 = atoms.copy()
            new_atoms_bridge1.extend(n2_bridge1)

            if get_shortest_ru_n_distance(new_atoms_bridge1) < 1.5:
                write(f"POSCAR_bridge1_{i+1}_check2", new_atoms_bridge1, format='vasp')
            else:
                write(f"POSCAR_bridge1_{i+1}.vasp", new_atoms_bridge1, format='vasp')
        else:
            write(f"POSCAR_bridge1_{i+1}.vasp", new_atoms_bridge1, format='vasp')

        ### **Bridge-2 & FCC-0 Placement**
        possible_c_rus = [ru for ru in connections[ru1] if ru != ru2] + [ru for ru in connections[ru2] if ru != ru1]

        if possible_c_rus:  # Ensure a third Ru atom exists
            ru3 = possible_c_rus[0]  # Select one possible C atom
            pos3 = atoms[ru3].position

            if is_valid_fcc0(atoms, ru1, ru2, [ru1, ru2], [ru3]):  
                adsorption_type = "fcc-0"
                filename = f"POSCAR_fcc0_{i+1}.vasp"
            else:
                adsorption_type = "bridge-2"
                filename = f"POSCAR_bridge2_{i+1}.vasp"

            n1_position = (pos1 + pos2) / 2 + direction * n_ru_distance
            n2_position = pos3 + direction * n_ru_distance  # N2 binds to Ru3 in FCC-0 or Ru1/Ru2 in Bridge-2

            n2_new = Atoms('N2', positions=[n1_position, n2_position])

            new_atoms = atoms.copy()
            new_atoms.extend(n2_new)

            if get_shortest_ru_n_distance(new_atoms) < 1.5:
                write(f"{filename.replace('.vasp', '_check1.vasp')}", new_atoms, format='vasp')
                direction = -direction
                n1_position = (pos1 + pos2) / 2 + direction * n_ru_distance
                n2_position = pos3 + direction * n_ru_distance
                n2_new = Atoms('N2', positions=[n1_position, n2_position])

                new_atoms = atoms.copy()
                new_atoms.extend(n2_new)

                if get_shortest_ru_n_distance(new_atoms) < 1.5:
                    write(f"{filename.replace('.vasp', '_check2.vasp')}", new_atoms, format='vasp')
                else:
                    write(filename, new_atoms, format='vasp')
            else:
                write(filename, new_atoms, format='vasp')

        ### **Bridge-3 Placement (N₂ perpendicular to Ru-Ru, both N have 2 Ru-N bonds)**
        perp_vector = np.cross(ru_vector, direction)
        perp_vector /= np.linalg.norm(perp_vector)

        nn_center_bridge3 = bridge_mid + direction * n_ru_distance
        n1_position_bridge3 = nn_center_bridge3 + perp_vector * (n_n_distance / 2)
        n2_position_bridge3 = nn_center_bridge3 - perp_vector * (n_n_distance / 2)
        n2_bridge3 = Atoms('N2', positions=[n1_position_bridge3, n2_position_bridge3])

        new_atoms_bridge3 = atoms.copy()
        new_atoms_bridge3.extend(n2_bridge3)

        if get_shortest_ru_n_distance(new_atoms_bridge3) < 1.5:
            write(f"POSCAR_bridge3_{i+1}_check1", new_atoms_bridge3, format='vasp')
            direction = -direction
            n1_position_bridge3 = nn_center_bridge3 + perp_vector * (n_n_distance / 2)
            n2_position_bridge3 = nn_center_bridge3 - perp_vector * (n_n_distance / 2)
            n2_bridge3 = Atoms('N2', positions=[n1_position_bridge3, n2_position_bridge3])
            new_atoms_bridge3 = atoms.copy()
            new_atoms_bridge3.extend(n2_bridge3)

            if get_shortest_ru_n_distance(new_atoms_bridge3) < 1.5:
                write(f"POSCAR_bridge3_{i+1}_check2", new_atoms_bridge3, format='vasp')
            else:
                write(f"POSCAR_bridge3_{i+1}.vasp", new_atoms_bridge3, format='vasp')
        else:
            write(f"POSCAR_bridge3_{i+1}.vasp", new_atoms_bridge3, format='vasp')

        print(f"Saved POSCAR_bridge3_{i+1}")
  

def add_hollow_sites(atoms, n_ru_distance=1.95, n_n_distance=1.19, n1_height=1.5,):
    """
    Add N2 molecules at hollow adsorption sites (FCC-1, FCC-2, FCC-3).
    """

    connections, cn_of_connected_atoms, top_sites, bridge_sites, hollow_sites = get_connection(atoms, metal=metal, mult=0.9)
    ru_positions = np.array([atom.position for atom in atoms if atom.symbol == 'Ru'])
    mass_center = np.mean(ru_positions, axis=0)

    for i, (ru1, ru2, ru3) in enumerate(hollow_sites):
        triangle_sites = [ru1, ru2, ru3]

        if not is_exposed_triangle(atoms, triangle_sites, height=2.5):
            print(f"Skipping hollow site {ru1+1}, {ru2+1}, {ru3+1} (not exposed)")
            continue

        pos1, pos2, pos3 = atoms[ru1].position, atoms[ru2].position, atoms[ru3].position
        hollow_center = np.mean([pos1, pos2, pos3], axis=0)

        # **Compute outward direction from mass center to hollow site**
        direction = hollow_center - mass_center
        direction /= np.linalg.norm(direction)

        if np.dot(direction, hollow_center - mass_center) < 0:
            direction = -direction  # Flip direction vector if needed

        normal = np.cross(pos2 - pos1, pos3 - pos1)
        normal /= np.linalg.norm(normal)

        ### **FCC-1 Placement**
        n1_position_fcc1 = (pos1 + pos2) / 2 + direction * n_ru_distance
        n2_position_fcc1 = (pos2 + pos3) / 2 + direction * n_ru_distance
        n2_fcc1 = Atoms('N2', positions=[n1_position_fcc1, n2_position_fcc1])

        new_atoms_fcc1 = atoms.copy()
        new_atoms_fcc1.extend(n2_fcc1)

        if get_shortest_ru_n_distance(new_atoms_fcc1) < 1.5:
            write(f"POSCAR_fcc1_{i+1}_check1", new_atoms_fcc1, format='vasp')
            direction = -direction
            n1_position_fcc1 = (pos1 + pos2) / 2 + direction * n_ru_distance
            n2_position_fcc1 = (pos2 + pos3) / 2 + direction * n_ru_distance
            n2_fcc1 = Atoms('N2', positions=[n1_position_fcc1, n2_position_fcc1])
            new_atoms_fcc1 = atoms.copy()
            new_atoms_fcc1.extend(n2_fcc1)

            if get_shortest_ru_n_distance(new_atoms_fcc1) < 1.5:
                write(f"POSCAR_fcc1_{i+1}_check2", new_atoms_fcc1, format='vasp')
            else:
                write(f"POSCAR_fcc1_{i+1}.vasp", new_atoms_fcc1, format='vasp')
        else:
            write(f"POSCAR_fcc1_{i+1}.vasp", new_atoms_fcc1, format='vasp')

        ### **FCC-2 Placement**
        n1_position_fcc2 = hollow_center + normal * n1_height
        perp_vector = np.cross(normal, direction)
        perp_vector /= np.linalg.norm(perp_vector)
        n2_position_fcc2 = n1_position_fcc2 + perp_vector * n_n_distance
        n2_fcc2 = Atoms('N2', positions=[n1_position_fcc2, n2_position_fcc2])

        new_atoms_fcc2 = atoms.copy()
        new_atoms_fcc2.extend(n2_fcc2)

        if get_shortest_ru_n_distance(new_atoms_fcc2) < 1.5:
            write(f"POSCAR_fcc2_{i+1}_check1", new_atoms_fcc2, format='vasp')
            direction = -direction
            n2_position_fcc2 = n1_position_fcc2 + perp_vector * n_n_distance
            n2_fcc2 = Atoms('N2', positions=[n1_position_fcc2, n2_position_fcc2])
            new_atoms_fcc2 = atoms.copy()
            new_atoms_fcc2.extend(n2_fcc2)

            if get_shortest_ru_n_distance(new_atoms_fcc2) < 1.5:
                write(f"POSCAR_fcc2_{i+1}_check2", new_atoms_fcc2, format='vasp')
            else:
                write(f"POSCAR_fcc2_{i+1}.vasp", new_atoms_fcc2, format='vasp')
        else:
            write(f"POSCAR_fcc2_{i+1}.vasp", new_atoms_fcc2, format='vasp')

        ### **FCC-3 Placement**
        n1_position_fcc3 = hollow_center + normal * n1_height
        closest_ru = min([pos1, pos2, pos3], key=lambda p: np.linalg.norm(p - n1_position_fcc3))
        n2_position_fcc3 = n1_position_fcc3 + perp_vector * n_n_distance
        n2_fcc3 = Atoms('N2', positions=[n1_position_fcc3, n2_position_fcc3])

        new_atoms_fcc3 = atoms.copy()
        new_atoms_fcc3.extend(n2_fcc3)

        if get_shortest_ru_n_distance(new_atoms_fcc3) < 1.5:
            write(f"POSCAR_fcc3_{i+1}_check1", new_atoms_fcc3, format='vasp')
            direction = -direction
            n2_position_fcc3 = n1_position_fcc3 + perp_vector * n_n_distance
            n2_fcc3 = Atoms('N2', positions=[n1_position_fcc3, n2_position_fcc3])
            new_atoms_fcc3 = atoms.copy()
            new_atoms_fcc3.extend(n2_fcc3)

            if get_shortest_ru_n_distance(new_atoms_fcc3) < 1.5:
                write(f"POSCAR_fcc3_{i+1}_check2", new_atoms_fcc3, format='vasp')
            else:
                write(f"POSCAR_fcc3_{i+1}.vasp", new_atoms_fcc3, format='vasp')
        else:
            write(f"POSCAR_fcc3_{i+1}.vasp", new_atoms_fcc3, format='vasp')


### MKM 

def generate_replacement_dict(ef_value, adsorption_data: dict) -> dict:
    """
    根据吸附态字典和 Ru 的能量，生成 replacement_dict，用于替换 Excel Sheet1 中的数据。

    参数：
    - adsorption_data: 吸附物字典，格式为 {ads: (site, adsorption_energy, total_energy)}

    返回：
    - replacement_dict: key 为 Excel 中的 A 列项（如 "NH3(T)"），value 为替换后的能量值（float）
    """
    # 固定的 slab 能量
    # E_slab = -354.40596642
    E_slab  = -6.2414 * ef_value**2 - 0.01405 * ef_value - 354.4197 

    # 初始化 replacement_dict
    replacement_dict = {
        "RU(T)": E_slab
    }

    # 常规吸附物种
    for species in ['NH3', 'NH2', 'NH', 'N', 'N2']:
        if species in adsorption_data:
            replacement_dict[f"{species}(T)"] = adsorption_data[species][2]

    # H1, H2, H3 -> Hv1(T), Hv2(T), Hv3(T)
    for i in range(1, 4):
        h_key = f"H{i}"
        hv_key = f"Hv{i}(T)"
        if h_key in adsorption_data:
            replacement_dict[hv_key] = adsorption_data[h_key][2]

    
    E_NH3 = adsorption_data['NH3'][2]
    E_NH2 = adsorption_data['NH2'][2]
    E_NH = adsorption_data['NH'][2]
    E_N  = adsorption_data['N'][2]
    E_N_N = 2 * adsorption_data['N'][2] - E_slab + 0.5
    E_N2 = adsorption_data['N2'][2] 
    try:
        E_H1  = adsorption_data['H1'][2]  
        E_H2  = adsorption_data['H2'][2] 
        E_H3  = adsorption_data['H3'][2] 
        E_H   = min([E_H1,E_H2,E_H3])
    except:
        E_H   =  adsorption_data['H'][2]
        
    replacement_dict['H(T)'] = E_H
    
    
    if float(adsorption_data['NH3'][1]) <= -1.9:
        print('E_NH3', adsorption_data['NH3'])
    
    # print(E_NH3, E_NH2, E_H1, E_slab)
    
    replacement_dict['N_N(T)'] = E_N_N
    
    print('E_N_ads:', adsorption_data['N'][1])
    print('E_H_ads:', adsorption_data['H'][1])
    print('E_N2_ads:', adsorption_data['N2'][1])
    print('E_NH3_ads:', adsorption_data['NH3'][1])

    reaction_energy_1 = E_NH2 + E_H - E_NH3 - E_slab
    reaction_energy_2 = E_NH  + E_H - E_NH2 - E_slab
    reaction_energy_3 = E_N   + E_H - E_NH  - E_slab
    reaction_energy_4 = E_N2  - E_N_N

    # Ea = a * ΔE + b
    Ea_params = [
        (0.42, 1.36),  # TS1_NH3(T)
        (0.88, 0.82),  # TS1_NH2(T)
        (0.47, 0.81),  # TS1_NH(T)
        (0.87, 2.10)   # TS4_N2(T)
    ]
    
    Ea1 = Ea_params[0][0] * reaction_energy_1 + Ea_params[0][1]
    Ea2 = Ea_params[1][0] * reaction_energy_2 + Ea_params[1][1]
    Ea3 = Ea_params[2][0] * reaction_energy_3 + Ea_params[2][1]
    Ea4 = Ea_params[3][0] * reaction_energy_4 + Ea_params[3][1]

    Ea_list = [Ea1, Ea2, Ea3, Ea4]
    # print('DE1', reaction_energy_1)
    # print('Ea_list', Ea_list)
    # 
    Ea_list = [max(0.01, ea) for ea in Ea_list]
    
    # print('Ea', Ea_list)
    Ea1, Ea2, Ea3, Ea4 = Ea_list

    # 过渡态能量
    replacement_dict['TS1_NH3(T)'] = E_NH3 + Ea1
    replacement_dict['TS2_NH2(T)'] = E_NH2 + Ea2
    replacement_dict['TS3_NH(T)']  = E_NH  + Ea3
    replacement_dict['TS4_N2(T)']  = E_N_N + Ea4

    ref_dict = {k: v - E_slab for k, v in replacement_dict.items()}

    return ref_dict
def build_refdict_and_plot_rc_double_NH3(ef_value: float, adsorption_data: dict, outputs_folder: str = "outputs") -> dict:
    os.makedirs(outputs_folder, exist_ok=True)

    # Gas-phase reference energies
    E_gas = {
        "NH3": -19.53586573,
        "NH2": -13.53307777,
        "NH": -8.10061060,
        "N2": -16.62922486,
        "H2": -6.76668776
    }

    # Field-dependent slab energy
    E_slab = -6.2414 * ef_value**2 - 0.01405 * ef_value - 354.4197

    # Adsorbed species energies
    E_NH3_ads = adsorption_data['NH3'][2]
    E_NH2_ads = adsorption_data['NH2'][2]
    E_NH_ads = adsorption_data['NH'][2]
    E_N_ads = adsorption_data['N'][2]
    try:
        E_H_ads = min(adsorption_data[k][2] for k in ('H1', 'H2', 'H3'))
    except:
        E_H_ads = adsorption_data['H'][2]
    E_N2_ads = adsorption_data['N2'][2]

    # Reaction energies (1 NH3 unit)
    ΔE1 = E_NH3_ads - E_slab - E_gas['NH3']
    ΔE2 = (E_NH2_ads + E_H_ads - E_slab) - E_NH3_ads
    ΔE3 = (E_NH_ads + E_H_ads - E_slab) - E_NH2_ads
    ΔE4 = (E_N_ads + E_H_ads - E_slab) - E_NH_ads
    ΔE5 = (E_N2_ads + E_slab) - 2 * E_N_ads
    ΔE6 = -adsorption_data['N2'][1]          # full N2 desorption
    try:
        ΔE7 = -adsorption_data['H1'][1] * 3      # 6H* → 3 H2(g)
    except:
        ΔE7 = -adsorption_data['H'][1] * 3

    # Barriers
    Ea_param = [(0.42, 1.36), (0.88, 0.82), (0.47, 0.81), (0.87, 2.10)]
    Ea1 = max(0.01, Ea_param[0][0] * ΔE2 + Ea_param[0][1])
    Ea2 = max(0.01, Ea_param[1][0] * ΔE3 + Ea_param[1][1])
    Ea3 = max(0.01, Ea_param[2][0] * ΔE4 + Ea_param[2][1])
    Ea4 = max(0.01, Ea_param[3][0] * ΔE5 + Ea_param[3][1])

    # State labels with TS included
    state_labels = [
        '2NH3(g)', '2NH3*', 'TS1', '2NH2*+2H*', 'TS2',
        '2NH*+4H*', 'TS3', '2N*+6H*', 'TS4',
        'N2*+6H*', 'N2(g)+6H*', 'N2(g)+3H2(g)'
    ]

    # Compute relE
    relE = [0]
    relE.append(relE[-1] + 2 * ΔE1)           # 2NH3*
    relE.append(relE[-1] + 2 * Ea1)           # TS1
    relE.append(relE[-2] + 2 * ΔE2)           # 2NH2*+2H*
    relE.append(relE[-1] + 2 * Ea2)           # TS2
    relE.append(relE[-2] + 2 * ΔE3)           # 2NH*+4H*
    relE.append(relE[-1] + 2 * Ea3)           # TS3
    relE.append(relE[-2] + 2 * ΔE4)           # 2N*+6H*
    relE.append(relE[-1] + Ea4)               # TS4
    relE.append(relE[-2] + ΔE5)               # N2*+6H*
    relE.append(relE[-1] + ΔE6)               # N2(g)+6H*
    relE.append(relE[-1] + ΔE7)               # N2(g)+3H2(g)

    # Sanity check: match to total gas-phase reaction energy
    ΔE_total_gas = E_gas['N2'] + 3 * E_gas['H2'] - 2 * E_gas['NH3']
    relE[-1] = ΔE_total_gas  # enforce exact gas-phase energy consistency

    # Build x/y for platform plot
    x_vals, y_vals = [], []
    for i, e in enumerate(relE):
        x_vals.extend([i, i + 1])
        y_vals.extend([e, e])

    labels_expanded = []
    for label in state_labels:
        labels_expanded.extend([label, label])

    df_rc = pd.DataFrame({'x': x_vals, 'y': y_vals, 'label': labels_expanded})
    df_rc.to_csv(os.path.join(outputs_folder, 'rc.csv'), index=False)

    # Plot
    plt.figure(figsize=(10, 5))
    plt.plot(x_vals, y_vals, '-k', lw=2)
    plt.scatter([i + 0.5 for i in range(len(relE))], relE, color='k', s=30, zorder=3)
    plt.axhline(0, ls='--', color='gray')
    plt.xticks([i + 0.5 for i in range(len(state_labels))], state_labels, rotation=45, ha='right')
    plt.ylabel('Relative energy (eV)')
    plt.title(f'2 NH₃ → N₂ + 3 H₂   |   EF = {ef_value:.1f} V/Å')
    plt.tight_layout()
    plt.savefig(os.path.join(outputs_folder, f'reaction_coordinate_2NH3_EF_{ef_value}.png'), dpi=300)
    plt.close()

    # print("E_end:", relE[-1], "E_end_gas_ref:", ΔE_total_gas)
    # return df_rc, relE[-1], ΔE_total_gas  # return for optional use


def build_refdict_and_plot_rc(ef_value: float, adsorption_data: dict, outputs_folder: str = "outputs") -> dict:
    os.makedirs(outputs_folder, exist_ok=True)

    # Gas-phase reference energies
    E_gas = {"NH3": -19.53586573, "NH2": -13.53307777, "NH": -8.10061060,
             "N2": -16.62922486, "H2": -6.76668776}

    # Field-dependent slab energy
    E_slab = -6.2414 * ef_value**2 - 0.01405 * ef_value - 354.4197

    # Adsorbed species energies
    E_NH3_ads = adsorption_data['NH3'][2]
    E_NH2_ads = adsorption_data['NH2'][2]
    E_NH_ads = adsorption_data['NH'][2]
    E_N_ads = adsorption_data['N'][2]
    E_H_ads = min(adsorption_data[k][2] for k in ('H1', 'H2', 'H3'))
    E_N2_ads = adsorption_data['N2'][2]

    # Reaction energies
    ΔE1 = E_NH3_ads - E_slab - E_gas['NH3']
    ΔE2 = (E_NH2_ads + E_H_ads - E_slab) - E_NH3_ads
    ΔE3 = (E_NH_ads + E_H_ads - E_slab) - E_NH2_ads
    ΔE4 = (E_N_ads + E_H_ads - E_slab) - E_NH_ads
    ΔE5 = (E_N2_ads + E_slab) - 2 * E_N_ads
    ΔE6 = -adsorption_data['N2'][1] * 0.5
    ΔE7 = -adsorption_data['H1'][1] * 1.5

    # Barrier estimation: Ea = a * ΔE + b
    Ea_param = [(0.42, 1.36), (0.88, 0.82), (0.47, 0.81), (0.87, 2.10)]
    Ea1 = max(0.01, Ea_param[0][0] * ΔE2 + Ea_param[0][1])
    Ea2 = max(0.01, Ea_param[1][0] * ΔE3 + Ea_param[1][1])
    Ea3 = max(0.01, Ea_param[2][0] * ΔE4 + Ea_param[2][1])
    Ea4 = max(0.01, Ea_param[3][0] * ΔE5 + Ea_param[3][1])

    # Extended state sequence including TS
    state_labels = [
        'NH3(g)', 'NH3*', 'TS1_NH3', 'NH2*+H*', 'TS2_NH2',
        'NH*+2H*', 'TS3_NH', 'N*+3H*', 'TS4_N2',
        'N2*+3H*', 'N2(g)+3H*', 'N2(g)+1.5H2(g)'
    ]

    # Compute state energies
    relE = [0]
    relE.append(relE[-1] + ΔE1)              # NH3*
    relE.append(relE[-1] + Ea1)              # TS1
    relE.append(relE[-2] + ΔE2)              # NH2* + H*
    relE.append(relE[-1] + Ea2)              # TS2
    relE.append(relE[-2] + ΔE3)              # NH* + 2H*
    relE.append(relE[-1] + Ea3)              # TS3
    relE.append(relE[-2] + ΔE4)              # N* + 3H*
    relE.append(relE[-1] + Ea4)              # TS4
    relE.append(relE[-2] + ΔE5)              # N2* + 3H*
    relE.append(relE[-1] + ΔE6)              # N2(g) + 3H*
    relE.append(relE[-1] + ΔE7)              # N2(g) + 1.5H2(g)

    # Build platform data for plotting
    x_vals = []
    y_vals = []
    for i in range(len(relE)):
        x_vals.extend([i, i + 1])
        y_vals.extend([relE[i]] * 2)
    # Expand state labels to match platform points
    labels_expanded = []
    for label in state_labels:
        labels_expanded.extend([label, label])
    
    # Save rc.csv with labels
    df_rc = pd.DataFrame({'x': x_vals, 'y': y_vals, 'label': labels_expanded})
    df_rc.to_csv(os.path.join(outputs_folder, 'rc.csv'), index=False)
    # Save rc.csv

    # Plot
    plt.figure(figsize=(10, 5))
    plt.plot(x_vals, y_vals, '-k', lw=2)
    plt.scatter([i + 0.5 for i in range(len(relE))], relE, color='k', s=30, zorder=3)
    plt.axhline(0, ls='--', color='gray')

    # Label x-axis
    plt.xticks([i + 0.5 for i in range(len(state_labels))], state_labels, rotation=45, ha='right')
    plt.ylabel('Relative energy (eV)')
    plt.title(f'EF = {ef_value:.1f} V/Å')
    plt.tight_layout()
    plt.savefig(os.path.join(outputs_folder, f'reaction_coordinate_EF_{ef_value}.png'), dpi=300)
    plt.close()

    # Replacement dict
    repl = {
        "RU(T)": E_slab,
        "NH3(T)": E_NH3_ads,
        "NH2(T)": E_NH2_ads,
        "NH(T)": E_NH_ads,
        "N(T)": E_N_ads,
        "N2(T)": E_N2_ads,
        "Hv1(T)": adsorption_data['H1'][2],
        "Hv2(T)": adsorption_data['H2'][2],
        "Hv3(T)": adsorption_data['H3'][2],
        "H(T)": E_H_ads,
        "N_N(T)": 2 * E_N_ads - E_slab + 0.5,
        "TS1_NH3(T)": E_NH3_ads + Ea1,
        "TS2_NH2(T)": E_NH2_ads + Ea2,
        "TS3_NH(T)": E_NH_ads + Ea3,
        "TS4_N2(T)": 2 * E_N_ads - E_slab + 0.5 + Ea4
    }

    ref_dict = {k: v - E_slab for k, v in repl.items()}
    return ref_dict

def plot_rc(EF, edft, e_slab, outputs_folder):
    E_gas = {                 # gas
        "H": -1.11671722,
        "H2": -6.76668776,
        "N": -3.12298738,
        "N2": -16.62922486,
        "NH": -8.10061060,
        "NH2": -13.53307777,
        "NH3": -19.53586573,
    }

    # ---------- 3. 取常用能量 ----------
    E_NH3_ads  = edft['NH3'][2]          # slab+NH3
    E_NH2_ads  = edft['NH2'][2]
    E_NH_ads   = edft['NH'][2]
    E_N_ads    = edft['N'][2]
    E_H_ads    = min(edft[k][2] for k in ('H1','H2','H3'))  # 最稳定 H
    E_N2_ads   = edft['N2'][2]
    
    # ---------- 4. 逐步反应热 ----------
    steps   = []
    labels  = []
    
    # 0) 初始气相
    cumE = 0.0
    steps.append(cumE); labels.append('NH$_3$(g)')
    
    # 1) NH3 吸附
    dE1  = E_NH3_ads - e_slab - E_gas['NH3']
    cumE += dE1
    steps.append(cumE); labels.append('NH$_3$*')
    
    # 2) NH3 → NH2* + H*
    dE2  = (E_NH2_ads + E_H_ads - e_slab) - E_NH3_ads
    cumE += dE2
    steps.append(cumE); labels.append('NH$_2$* + H*')
    
    # 3) NH2* → NH* + H*
    dE3  = (E_NH_ads + E_H_ads - e_slab) - E_NH2_ads
    cumE += dE3
    steps.append(cumE); labels.append('NH* + 2H*')
    
    # 4) NH* → N* + H*
    dE4  = (E_N_ads + E_H_ads - e_slab) - E_NH_ads
    cumE += dE4
    steps.append(cumE); labels.append('N* + 3H*')
    
    # 5) N* + N* → N2*
    dE5  = (E_N2_ads + e_slab) - 2 * E_N_ads
    cumE += dE5/2
    steps.append(cumE); labels.append('N$_2$* + 3H*')
    
    # 6) N2 脱附
    dE6  = - edft['N2'][1] * 0.5          # 按给定规则：吸附能的 -½
    cumE += dE6
    steps.append(cumE); labels.append('N$_2$(g) + 3H*')
    
    # 7) 3H* → 1.5 H2(g)
    dE7  = - edft['H1'][1] * 1.5          # 最稳定 H 的吸附能 × -1.5
    cumE += dE7
    steps.append(cumE); labels.append('N$_2$(g) + 1.5H$_2$(g)')
    
    # ---------- 5. 画 Reaction Coordinate ----------
    # x = range(len(steps))
    # plt.figure(figsize=(8,5))
    # plt.plot(x, steps, '-o', lw=2, ms=6, color='k')
    # plt.axhline(0, ls='--', color='grey', lw=1)
    
    # plt.xticks(x, labels, rotation=45, ha='right')
    # plt.ylabel('Relative energy (eV)')
    # plt.tight_layout()
    # plt.savefig('reaction_coordinate.png', dpi=300)
    # plt.show()
    
    # --- 构造平台坐标 ---
    x_vals = []
    y_vals = []
    for i in range(len(steps)):
        x_vals.extend([i, i+1])
        y_vals.extend([steps[i]] * 2)
    df_rc = pd.DataFrame({'x': x_vals, 'y': y_vals})
    df_rc.to_csv(os.path.join(outputs_folder, 'rc.csv'), index=False)
    
    # --- 设置平台中点 label 的位置 ---
    label_pos = [(i + i + 1)/2 for i in range(len(steps))]
    
    # --- 绘图 ---
    plt.figure(figsize=(8,5))
    plt.plot(x_vals, y_vals, '-', lw=2, color='k', label=f'EF = {EF:.1f} V/Å')
    plt.scatter(label_pos, steps, s=30, color='k', zorder=3)  # 小圆点标示每个状态
    plt.axhline(0, ls='--', color='gray', lw=1)
    
    # x-axis
    plt.xticks(label_pos, labels, rotation=45, ha='right')
    plt.xlim(0, len(steps))
    plt.ylabel('Relative energy (eV)')
    plt.legend()

    plt.tight_layout()
    filename = os.path.join(outputs_folder, f'reaction_coordinate_EF_{EF}.png')
    plt.savefig(filename, dpi=300)
    plt.close()



def update_excel_with_replacement(index_list, replacement_dict, ef_value,
                                  template_file="NH3_temp.xlsx", base_dir="mkm_inputs", verbose=True):
    """
    Create a structured directory based on index and EF, update the "species" sheet 
    in the Excel template with new data, execute additional workflow commands, and save the result.
    
    Parameters:
        index_list: List of indices used for naming subfolders, e.g. [12, 43, 40, 44]
        replacement_dict: Dictionary with replacement items, e.g. {"NH3(T)": -375.2, ...}
        ef_value: External EF value (float, can be negative) for naming subfolders (e.g. -0.3)
        template_file: Excel template file name; now expected to be located inside base_dir.
        base_dir: Base directory (e.g. "mkm_inputs")
        verbose: Whether to print verbose output.
    """
    # 1. Build directory paths
    index_folder = "_".join(map(str, index_list))
    ef_folder = f"EF_{ef_value:.1f}"  # e.g. EF_-0.3
    inputs_folder = os.path.join(base_dir, index_folder, ef_folder, 'inputs')
    os.makedirs(inputs_folder, exist_ok=True)
    
    # Create outputs folder at the same level as inputs
    outputs_folder = os.path.join(base_dir, index_folder, ef_folder, 'outputs')
    os.makedirs(outputs_folder, exist_ok=True)
    
    # 2. Load the Excel template
    template_path = os.path.join(base_dir, template_file)
    if not os.path.exists(template_path):
        sys.exit(f"Template file not found: {template_path}")
        
        
    from openpyxl import load_workbook    
    wb = load_workbook(template_path)
    dft_sheet = wb["species"]
    
    # 3. Replace data in the sheet: get keys from column A, replace potential energy in column L"
    # print(f"✅ Fill the Excel Template File")
    
    missing_keys = []
    for row in dft_sheet.iter_rows(min_row=2, min_col=1, max_col=12):
        key_cell = row[0]    # Column A (key)
        target_cell = row[11]  # Column L (the value to replace)
        key = key_cell.value
        if key in replacement_dict:
            target_cell.value = replacement_dict[key]
            # if verbose:
                # print(f"✅ Replace: {key} -> {replacement_dict[key]:.6f}")
        else:
            missing_keys.append(key)
            
    # for key in missing_keys:
        # print(f"  - {key}")
    
    # 4. Save the updated Excel file into the inputs folder
    output_path = os.path.join(inputs_folder, 'NH3_Input_Data.xlsx')
    wb.save(output_path)
    # print(f"📁 File saved to: {output_path}")
    
    # 5. Copy files to the outputs folder (from base_dir)
    for filename in ["NH3_MKM_Ru45.py"]:
        src = os.path.join(base_dir, filename)
        dst = os.path.join(outputs_folder, filename)
        if os.path.exists(src):
            shutil.copy2(src, dst)
            # if verbose:
                # print(f"✅ Copied {filename} to {outputs_folder}")
        # else:
            # print(f"⚠️ File {filename} not found!")
    
    # Copy OpenMKM_IO.py to the folder: mkm_inputs/index_folder/ef_folder
    src_IO = os.path.join(base_dir, "OpenMKM_IO.py")
    dst_IO_dir = os.path.join(base_dir, index_folder, ef_folder)
    dst_IO = os.path.join(dst_IO_dir, "OpenMKM_IO.py")
    if os.path.exists(src_IO):
        shutil.copy2(src_IO, dst_IO)
    #     if verbose:
    #         print(f"✅ Copied OpenMKM_IO.py to {dst_IO_dir}")
    # else:
    #     print("⚠️ File OpenMKM_IO.py not found!")
    
    # 6. Workflow:
    main_folder = os.path.join(base_dir, index_folder, ef_folder)    
    
    # 6.1 Change to main_folder and execute "python3 OpenMKM_IO.py"
    # print("✅ Generating the MKM inputs...")
    # subprocess.run(["python3", "OpenMKM_IO.py"], cwd=main_folder, check=True)
    prepare_yaml(main_folder)
    
    # # 6.2 Instead of running "bash modify.sh", update the YAML file.
    # yaml_file = os.path.join(main_folder, "outputs", "thermo.yaml")
    # print("✅ Updating YAML file in outputs folder...")
    # update_yaml(yaml_file)
    
    # 6.3 In outputs folder, run "python3 NH3_v2.py 600 > mkm.out"
    # print("✅ Running MKM...")
    #T = 400   # Celcius
    #run_mkm(main_folder, T)
    # subprocess.run("python3 NH3_v2.py 600 > mkm.out", cwd=outputs_folder, shell=True, check=True)
    
    # print("✅ Workflow completed!\n")

