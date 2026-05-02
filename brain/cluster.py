import os, sys
import csv
import json
import numpy as np
import pandas as pd
import itertools
import copy
from pathlib import Path
import math
import matplotlib.pyplot as plt
import joblib  # import the joblib library
import platform
import re
import shutil
import subprocess
from openpyxl import load_workbook
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from ase import Atoms
from ase.io import read, write
from matplotlib.ticker import LinearLocator, FormatStrFormatter


# Set default font properties
# plt.rc('font', size=18)  # Default text font size
# plt.rc('axes', titlesize=18)  # Title font size
# plt.rc('axes', labelsize=18)  # X and Y label font size
# plt.rc('xtick', labelsize=16)  # X tick label font size
# plt.rc('ytick', labelsize=16)  # Y tick label font size
# plt.rc('legend', fontsize=16)  # Legend font size
# plt.rc('figure', titlesize=20)  # Figure title font size
plt.rcParams['font.family'] = 'Times New Roman'
# plt.rcParams["font.family"] = "Times New Roman"
plt.rcParams["font.size"]   = 14

from matplotlib.patches import Patch
import matplotlib.patches as patches
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.collections import PatchCollection
from matplotlib.colors import ListedColormap
from matplotlib import cm
from matplotlib.colors import ListedColormap


from scipy.spatial import KDTree, ConvexHull
from scipy.spatial.transform import Rotation as R
from scipy.spatial.distance import euclidean
from scipy.integrate import odeint
from scipy.integrate import solve_ivp
from scipy.constants import h,  eV

from ase import Atoms
from ase.io import read, write
from ase.neighborlist import NeighborList, natural_cutoffs, neighbor_list
from ase.io.vasp import read_vasp, write_vasp
from ase.visualize import view

from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import cross_val_score, KFold
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, BayesianRidge, LogisticRegression
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.metrics import make_scorer, mean_absolute_error, mean_squared_error, r2_score
from sklearn.neural_network import MLPRegressor
from sklearn.decomposition import PCA
from sklearn.cross_decomposition import PLSRegression
from sklearn.pipeline import make_pipeline
# from xgboost import XGBRegressor
from scipy.optimize import curve_fit
from scipy.stats import pearsonr
from sklearn.metrics import mean_absolute_error, mean_squared_error

from mkm import * 

def round_scientific(number, decimals):
    # Format the number in scientific notation with the desired number of decimals
    formatted_number = f"{number:.{decimals}e}"
    return formatted_number

covalent_radii = np.array([1,
                           0.32, 0.46, 1.20, 0.94, 0.77, 0.75, 0.71, 0.63, 0.64, 0.67,
                           1.40, 1.25, 1.13, 1.04, 1.10, 1.02, 0.99, 0.96, 1.76, 1.54,
                           1.33, 1.22, 1.21, 1.10, 1.07, 1.04, 1.00, 0.99, 1.01, 1.09,
                           1.12, 1.09, 1.15, 1.10, 1.14, 1.17, 1.89, 1.67, 1.47, 1.39,
                           1.32, 1.24, 1.15, 1.13, 1.13, 1.08, 1.15, 1.23, 1.28, 1.26,
                           1.26, 1.23, 1.32, 1.31, 2.09, 1.76, 1.62, 1.47, 1.58, 1.57,
                           1.56, 1.55, 1.51, 1.52, 1.51, 1.50, 1.49, 1.49, 1.48, 1.53,
                           1.46, 1.37, 1.31, 1.23, 1.18, 1.16, 1.11, 1.12, 1.13, 1.32,
                           1.30, 1.30, 1.36, 1.31, 1.38, 1.42, 2.01, 1.81, 1.67, 1.58,
                           1.52, 1.53, 1.54, 1.55])

cnmax_values = {
    'top': 12,
    'bridge': 18,
    'three_fold': 22,
    'four_fold': 26
}

gas_dict = {
    "CH": -6.05202894,
    "CH2": -11.78901352,
    "CH3": -18.19317746,
    "CO": -14.78130054,
    "H": -1.11671722,
    "H2": -6.76668776,
    "N": -3.12298738,
    "N2": -16.62922486,
    "NH": -8.10061060,
    "NH2": -13.53307777,
    "NH3": -19.53586573,
    "O": -1.90384114,
    "O2": -9.85600535,
    "OH": -7.72794651,
    "OH2": -14.21849996,
    "H2O": -14.21849996,
    "S": -1.08009627,
    "SH": -6.01226074,
    "SH2": -11.19981098,
    "H2S": -11.19981098,
    "thiol": -55.96808919,
    'thiolate_lateral':-50.67527044,
    "thiolate_upright": -50.67527044,
}

def calculate_triangle_center(list_coordinates):
    return np.mean(list_coordinates, axis=0)

def is_valid_bri(atoms_positions, max_length=3.0):
    if np.linalg.norm(atoms_positions[0] - atoms_positions[1]) > max_length:
        return False
    return True

def calculate_normal(v1, v2, v3):
    """Calculate the unit normal vector of a triangle defined by three points"""
    normal = np.cross(v2 - v1, v3 - v1)
    return normal / np.linalg.norm(normal)

def calculate_cn(structure, metal, cutoff=2.7):
    """Calculate the coordination number for each atom in the structure, considering only metal atoms."""
    distances = structure.get_all_distances(mic=True)
    cn = np.zeros(len(structure))
    for i, atom in enumerate(structure):
        if atom.symbol == metal:
            neighbors = np.where((distances[i] < cutoff) & (distances[i] > 0))[0]
            cn[i] = sum(1 for n in neighbors if structure[n].symbol == metal)
    return cn

def get_CN(atoms, mult=1.0):
    radii = natural_cutoffs(atoms, mult=mult)
    nl = NeighborList(radii, bothways=True, self_interaction=False)
    nl.update(atoms)
    CN_matrix = nl.get_connectivity_matrix(sparse=False)
    CN = CN_matrix.sum(axis=0)
    return CN, CN_matrix

def calculate_gcn(structure, cn, cnmax, metal='Ru'):
    """Calculate the GCN for each atom in the structure, considering only metal atoms."""
    distances = structure.get_all_distances(mic=True)
    gcn = np.zeros(len(structure))
    for i, atom in enumerate(structure):
        if atom.symbol == metal:
            neighbors = np.where((distances[i] < 3) & (distances[i] > 0))[0]
            cn_neighbors = [cn[n] for n in neighbors if structure[n].symbol == metal]
            gcn[i] = np.sum(cn_neighbors) / cnmax
    return np.round(gcn, 2)  # 保留两位小数

def get_GCN(CN, CN_matrix, cn_max: float=None):
    if cn_max is None:
        cn_max = CN_matrix.sum(axis=0).max()
    num_atoms = CN_matrix.shape[0]
    GCN = np.zeros(num_atoms)
    for A in range(num_atoms):
        for B in range(num_atoms):
            GCN[A] += CN_matrix[A, B] * CN[B]
        GCN[A] /=  cn_max
    return GCN

def calculate_gcn_for_sites(structure, site_indices, gcn):
    """Calculate GCN values for specified sites in a structure, considering only metal atoms."""
    gcn_values = []
    for sites in site_indices:
        total_gcn = 0
        for site in sites:
            total_gcn += gcn[site]
        gcn_values.append(round(total_gcn / len(sites), 2))  # 保留两位小数
    return gcn_values

def calculate_cluster_size(structure, metal='Ru'):
    """Calculate the cluster size based on the maximum distance between metal atoms."""
    metal_positions = [atom.position for atom in structure if atom.symbol == metal]
    max_distance = 0
    for i in range(len(metal_positions)):
        for j in range(i + 1, len(metal_positions)):
            distance = np.linalg.norm(metal_positions[i] - metal_positions[j])
            if distance > max_distance:
                max_distance = distance
    return max_distance


def get_top_sites(atoms_in, metal = 'Ru', mult=0.9):
    """Obtain the exposed  top sites"""   
    filtered_atoms = [atom for atom in atoms_in if atom.symbol  in [metal]]
    atoms = Atoms(filtered_atoms)
    radii = natural_cutoffs(atoms, mult=mult)
    nl = NeighborList(radii, bothways=True, self_interaction=False)
    nl.update(atoms)
    CN_matrix = nl.get_connectivity_matrix(sparse=False)
    CN = CN_matrix.sum(axis=0)
    exposed_top_sites = [i for i, cn in enumerate(CN) if cn <= 9]
    return exposed_top_sites

def get_connection(atoms_in, metal='Ru', mult=0.9):
    # atoms_in = read(path + '/POSCAR')

    filtered_atoms = [atom for atom in atoms_in if atom.symbol in [metal]]
    atoms = Atoms(filtered_atoms)
    radii = natural_cutoffs(atoms, mult=mult)
    nl = NeighborList(radii, bothways=True, self_interaction=False)
    nl.update(atoms)
    CN_matrix = nl.get_connectivity_matrix(sparse=False)
    CN = CN_matrix.sum(axis=0)
    exposed_top_sites = [i for i, cn in enumerate(CN) if cn <= 9]       # Exposed top sites (CN <= 9)

    connections = {}     # Get the connections of all atoms
    for i in range(len(atoms)):
        connections[i] = list(np.where(CN_matrix[i])[0])

    cn_of_connected_atoms = {}      # Request 2: Calculate the CN of the connected atoms
    for i in range(len(atoms)):
        connected_indices = connections[i]
        cn_of_connected_atoms[i] = [CN[j] for j in connected_indices]

    exposed_connections = {i: [j for j in connections[i] if j in exposed_top_sites] for i in exposed_top_sites}

    # Bridge sites (pairs of connected exposed top sites)
    bridge_sites = []
    for i in exposed_top_sites:
        for j in exposed_connections[i]:
            if i < j:  # Ensure each pair is considered only once
                # The CN < 9 condition is not strictly necessary for bridge detection,
                # but it is often used to focus on undercoordinated (surface) atoms.
                bridge_sites.append([i, j])

    # Hollow sites (triplets of closely connected exposed top sites)
    hollow_sites = []
    for i in exposed_top_sites:
        for j in exposed_connections[i]:
            for k in exposed_connections[j]:
                if i < j and j < k and i in exposed_connections[k]:  # Ensure each triplet is considered only once and forms a triangle
                    # The CN < 9 condition is not strictly necessary for hollow detection,
                    # but it is often used to focus on undercoordinated (surface) atoms.
                    hollow_sites.append([i, j, k])

    # Square sites (quartets of connected exposed top sites forming a closed loop)
    square_sites = []
    for i in exposed_top_sites:
        for j in exposed_connections[i]:
            if j <= i:
                continue
            for k in exposed_connections[j]:
                if k in (i, j) or k <= j:
                    continue
                for l in exposed_connections[k]:
                    if l in (i, j, k) or l <= k:
                        continue
                    # Check if l connects back to i to form a closed loop
                    if i in exposed_connections[l]:
                        # Ensure all connections exist: i-j, j-k, k-l, l-i
                        square = [i, j, k, l]
                        # The CN < 9 condition is not strictly necessary for square detection,
                        # but it is often used to focus on undercoordinated (surface) atoms.
                        square_sites.append(square)

    return connections, cn_of_connected_atoms, exposed_top_sites, bridge_sites, hollow_sites, square_sites

def number_to_letter(num):
    """Conver the CN to letters"""
    if num == 0:
        return '0'
    return chr(ord('a') + num - 1)


def get_CN_GA(path, mult=0.9, metal='Ru'):
    
    '''
    Important: it is the path for the clean cluster. 
    1) the exposed surface of the cluster will be scanned and all the top, bridge, and hollow sites will be stored.
    2) the sites will be stored as a dictionray 
    3) all the groups will be stored too to  fix the order of the  groups in the matrix
    '''    
    file_in = os.path.join(path,'POSCAR')
    atoms = read(file_in)
    print(metal)
    connections, cn_of_connected_atoms, exposed_top_sites, bridge_sites, hollow_sites, square_sites  = get_connection(atoms, metal=metal, mult=0.9)
    print('top', exposed_top_sites)
    def get_one_site_string(site):
        """Obtain the text string representation for single site """
        cn_values = sorted(cn_of_connected_atoms[site])
        # print(cn_values)
        if len(cn_values) < 13:
            cn_values += [0] * (13 - len(cn_values))  # Pad with zeros if less than 12 elements
        else: 
            cn_values += [0] * (13 - 12)  # Pad with zeros if less than 12 elements
        
        cn_string = ''.join(number_to_letter(num) for num in cn_values)
         
        return cn_string
    
    def get_groups(site_indices, site_type):
        """
        Generic function to get group strings for a site (top, bridge, hollow, square).
        site_indices: int for top, list of ints for others.
        site_type: 'top', 'bridge', 'hollow', or 'square'
        """
        if site_type == 'top':
            site = site_indices
            groups = [get_one_site_string(site)]
            groups += [get_one_site_string(s) for s in connections[site]]
            return groups
        else:
            sites = site_indices
            strings = [get_one_site_string(i) for i in sites]
            groups = ['-'.join(sorted(strings))]
            # Collect all unique surrounding atoms except the site atoms themselves
            all_neighbors = set()
            for i in sites:
                all_neighbors.update(connections[i])
            surrounding_atoms = list(all_neighbors - set(sites))
            groups += [get_one_site_string(s) for s in surrounding_atoms]
            return groups

    GA_dict = {}
    for site in exposed_top_sites:
        key = site + 1
        GA_dict[key] = get_groups(site, 'top')
        
        if site == 30:
            print(site, get_groups(site, 'top'))

    for sites in bridge_sites:
        key = '_'.join(str(i + 1) for i in sites)
        GA_dict[key] = get_groups(sites, 'bridge')

    for sites in hollow_sites:
        key = '_'.join(str(i + 1) for i in sites)
        GA_dict[key] = get_groups(sites, 'hollow')

    for sites in square_sites:
        key = '_'.join(str(i + 1) for i in sites)
        GA_dict[key] = get_groups(sites, 'square')

    GA_file = os.path.join(path, 'GA_dict.txt')

    # Writing JSON data
    with open(GA_file, 'w') as f:
        json.dump(GA_dict, f)
        
    # Summarize and extract all unique groups from GA_dict
    groups = {group for sublist in GA_dict.values() for group in sublist}
    
    # Save groups to a text file, one per line
    groups_file = os.path.join(path, 'groups.txt')
    with open(groups_file, 'w') as f:
        for group in groups:
            f.write(f"{group}\n")


def func(x, a, b, c):
    return -0.5 * a * x**2 - b * x + c

def surface_site_center(positions: np.ndarray) -> np.ndarray:
    """
    Given positions of 1, 2, or 3 atoms (shape: (N, 3) with N in {1,2,3}),
    return the site coordinate:
      - N=1: the atom position
      - N=2: midpoint
      - N=3: triangle centroid (simple average of vertices)
    """
    if positions.shape[0] not in (1, 2, 3):
        raise ValueError("Provide 1, 2, or 3 surface atom indices.")
    return positions.mean(axis=0)

def angle_com_to_surface_xy(atoms: Atoms, surf_indices) -> float:
    """
    Parameters
    ----------
    atoms : ase.Atoms
        Your Ru45 cluster (or any cluster); masses must be set for COM (ASE sets by symbol).
    surf_indices : int | list[int] | tuple[int, ...]
        1, 2, or 3 indices of surface atoms defining the surface site.

    Returns
    -------
    angle_deg : float
        Angle in degrees between the vector (COM -> surface_site) and the xy-plane
        that passes through the COM (i.e., 0° = parallel to plane, 90° = straight up/down).
    """
    # 1) Center of mass of the cluster
    com = atoms.get_center_of_mass()  # (3,)

    # 2) Surface-site position from indices
    if isinstance(surf_indices, (int, np.integer)):
        idx_list = [int(surf_indices) -1]
    else:
        idx_list = [int(i) -1 for i in surf_indices]


    site = atoms.get_positions()[idx_list].mean(axis=0)
    v = site - com
    r_xy = np.linalg.norm(v[:2])          # 与平面的投影长度
    angle_rad = np.arctan2(v[2], r_xy)    # 保留符号：vz>0 正，vz<0 负
    return float(np.degrees(angle_rad))


def get_dipole_polarization(GA_matrix, list_EF, fill_nans_with_fit=True, verbose=True):

    # x_values = np.array([-0.6, -0.4, -0.2, 0.0, 0.2, 0.4, 0.6])
    # eads_cols = ['E_ads_' + i for i in ["-0.6", "-0.4", "-0.2", "0.0", "0.2", "0.4", "0.6"] ]
    # x_values = np.array([-0.6, -0.4, -0.2, -0.1, 0.0, 0.1, 0.2, 0.4, 0.6])
    eads_cols = [f"E_ads_{val}" for val in list_EF]
    site_col = 'Site' if 'Site' in GA_matrix.columns else 'site'
    angle_col = 'angle'

    results_list = []
    GA_matrix = GA_matrix.copy()  # 避免修改原始 DataFrame

    x_values = np.array(list_EF, dtype=float)
    for idx, row in GA_matrix.iterrows():
        site = row[site_col]
        angle = row[angle_col]
        y_values_orig = row[eads_cols].values.astype(float)
        valid_mask = ~np.isnan(y_values_orig)
        x_fit = x_values[valid_mask] * math.sin(math.radians(angle))
        y_fit = y_values_orig[valid_mask]

        if len(x_fit) < 3:
            if verbose:
                print(f"Skipping site '{site}' (index {idx}) due to insufficient data points.")
            continue

        try:
            # 第一次拟合
            params, _ = curve_fit(func, x_fit, y_fit)
            y_pred = func(x_fit, *params)
            deviation = np.abs(y_pred - y_fit)
            outlier_found = False

            if np.any(deviation > 0.2):
                outlier_found = True
                if verbose:
                    print(f"Site '{site}' (index {idx}) has deviations > 0.2 eV:")
                for xi, yi, ypi, dev in zip(x_fit, y_fit, y_pred, deviation):
                    if dev > 0.2:
                        if verbose:
                            print(f"  E={xi:.1f}: actual={yi:.3f}, predicted={ypi:.3f}, Δ={dev:.3f}")
                        GA_matrix.at[idx, f"{xi:.1f}"] = np.nan

                # 重新提取数据
                y_values_clean = GA_matrix.loc[idx, eads_cols].values.astype(float)
                valid_mask = ~np.isnan(y_values_clean)
                x_fit = x_values[valid_mask]
                y_fit = y_values_clean[valid_mask]

                if len(x_fit) < 3:
                    if verbose:
                        print(f"After removing outliers, not enough data for site '{site}' (index {idx})")
                    continue

                # 第二次拟合
                params, _ = curve_fit(func, x_fit, y_fit)
                y_pred = func(x_fit, *params)

            # 填补初始 NaN（如果没有出现 outlier）
            elif fill_nans_with_fit and np.any(np.isnan(y_values_orig)):
                full_pred = func(x_values, *params)
                for j, val in enumerate(y_values_orig):
                    if np.isnan(val):
                        if verbose:
                            print(f"  Filling NaN for site '{site}' at E={x_values[j]:.1f} with predicted value {full_pred[j]:.3f}")
                        GA_matrix.at[idx, f"{x_values[j]:.1f}"] = full_pred[j]

            a, b, c = params
            R2 = r2_score(y_fit, y_pred)
            mae = mean_absolute_error(y_fit, y_pred)
            rmse = np.sqrt(mean_squared_error(y_fit, y_pred))

            results_list.append({
                'site': site,
                'polarizability': a,
                'dipole': b,
                'c': c,
                # 'R2': R2,
                # 'MAE': mae,
                # 'RMSE': rmse
            })

        except Exception as e:
            if verbose:
                print(f"Error fitting site '{site}' (index {idx}): {e}")
            continue

    results_df = pd.DataFrame(results_list)
    return results_df#, GA_matrix


def load_ga_data(cluster_path):
    """
    Loads the GA dictionary and the groups list from files.

    Parameters:
        load_path (str): The directory where the files are stored.
                         Files are expected to be named 'GA_dict.txt' and 'groups.txt'.
    
    Returns:
        tuple: A tuple containing the GA dictionary and the groups list.
    """
    # Load GA_dict from JSON file
    ga_file = os.path.join(cluster_path, 'GA_dict.txt')
    with open(ga_file, 'r') as f:
        GA_dict = json.load(f)
    
    # Load groups from the text file
    groups_file = os.path.join(cluster_path, 'groups.txt')
    with open(groups_file, 'r') as f:
        groups = [line.strip() for line in f if line.strip()]
    
    return GA_dict, groups


def get_full_GA_matrix(cluster_path):
    """
    Generate the GA (Group Additivity) matrix for all the surface sites
    Alos, we add the angle information of surface sites to the GA matrix.
    """
    GA_dict, groups = load_ga_data(cluster_path)
    atoms  = read(cluster_path + '/POSCAR')
    # Prepare header for the GA matrix DataFrame
    header = ['site'] + groups + ['angle']
    GA_matrix = []

    for key, values in GA_dict.items():  # key is the index of the atom in the POSCAR, 1-based
        groups_geo = values
        row = [key]
        # Create the GA matrix row: first column is the site label, then counts for each group
        for group in groups:
            count = groups_geo.count(group)
            row.append(count)
            
        ## Obtain The angle (A) between the mass center-surface atom vector and the xy-plane    
        indices = [int(i) for i in key.split('_')]
        angle = angle_com_to_surface_xy(atoms, indices)
        row.append(angle)
        
        GA_matrix.append(row)

    # Create a DataFrame from the GA matrix
    df_GA = pd.DataFrame(GA_matrix, columns=header)

    
    # Save the GA matrix to CSV
    output_csv = cluster_path + "/GA_matrix_full.csv"
    
    df_GA.to_csv(output_csv, index=False)
    print("GA matrix saved as", output_csv)
    
    return df_GA


def generate_GA_matrix_species(cluster_path, list_EF):
    """  
    This function will update the adsorption energies to the GA_matrix file.
    """
    
    ga_matrix_file = os.path.join(cluster_path, 'GA_matrix_full.csv')
    print("Reading from:", ga_matrix_file)

    with open(ga_matrix_file, 'r') as f:
        first_line = f.readline()
        print("Actual header line in file:")
        print(first_line.strip())
        
    # GA_matrix_full = get_full_GA_matrix(cluster_path)
    GA_matrix_full = pd.read_csv(ga_matrix_file)
    print("Original GA_matrix_full columns from slab folder:")
    print(GA_matrix_full.columns.tolist())
    
    sites = GA_matrix_full['site'].tolist()
    GA_matrix_species = GA_matrix_full.copy()
    # Ensure results directory exists
    if not os.path.exists('results'):
        os.makedirs('results')
    
    data_file = './sorted_data.csv'
    df_data = pd.read_csv(data_file)

    species = get_species_from_cwd()  

    for ef_val in list_EF:
        eads_ef = []
        for site in sites:         
        # Retrieve the slab energy from df where Species equals 'slab'
            try:
                slab_energy = df_data.loc[df_data['Species'] == 'slab', str(ef_val)].values[0]
                # print(slab_energy)
            except KeyError:
                raise KeyError(f"EF value '{ef_val}' not found in the data.")
    
            # Compute the adsorption energy: subtract slab energy and a gas-phase correction
    
            if species in ['H', 'N', 'O']:
                gas_species = species + '2'
                E_gas = float(gas_dict[gas_species]) / 2
            else:
                gas_species = species
                E_gas = float(gas_dict[gas_species])
                
            # print('Eslab', slab_energy)
            try:
                ef_value = df_data.loc[df_data['site'] == site, str(ef_val)].values[0]
                eads_value = ef_value - slab_energy - E_gas
            except :
                eads_value = float('nan') 
            eads_ef.append(eads_value)
            
            # Add the computed energy column to the GA matrix DataFrame.
            # (Assumes the row order of df and df_GA is the same.)
        GA_matrix_species[f'E_ads_{ef_val}'] = eads_ef

    # Save the GA matrix to CSV
    output_csv = "./GA_matrix.csv"
    
    GA_matrix_species.to_csv(output_csv, index=False)
    print("GA matrix of the species is saved as", output_csv)
    
    return GA_matrix_species

def update_GA_matrix_with_dipole_polarization(csv_filepath, list_EF):
    
    """
    Update the GA_matrix by adding 'polarizability' and 'dipole' columns based on curve fitting.
    
    This function reads the GA_matrix CSV file, computes the dipole polarization fits for each row 
    (using get_dipole_polarization, which returns columns 'polarizability' and 'dipole'), merges these 
    values into the original GA_matrix based on the 'site' column, and then overwrites the CSV file.
    
    Args:
        csv_filepath (str): Path to the GA_matrix CSV file (default: './GA_matrix.csv').
    
    Returns:
        updated_df (pandas.DataFrame): The updated GA_matrix DataFrame with the new columns.
        
            
    """
    # Read the current GA_matrix.
    df_GA = pd.read_csv(csv_filepath)
    
    # Get the dipole polarization results.
    # This function must be defined and should return a DataFrame with columns: 
    # 'site', 'polarizability', 'dipole', 'c', 'R2', 'MAE', 'RMSE'
    dipole_results = get_dipole_polarization(df_GA, list_EF)
    
    # Merge the dipole polarization results into the GA_matrix based on the 'site' column.
    updated_df = df_GA.merge(dipole_results[['site', 'polarizability', 'dipole', 'c']], on='site', how='left')
    
    # Overwrite the original GA_matrix CSV file with the updated DataFrame.
    updated_df.to_csv(csv_filepath, index=False)
    print(f"Updated GA_matrix saved to {csv_filepath}.")
    
    return updated_df

    
def get_species_from_cwd():
    # cwd = os.getcwd()
    cwd = os.path.realpath(os.getcwd())
    # print(cwd, 'cwd')
    species = None
    try:
        # path_parts = cwd.replace(' ', '').split(os.path.sep)
        part = os.path.basename(cwd).strip()
        # print('part', part)
        if '_ads' in part:
            species = part.split('_')[0]
        
    except Exception as e:
        print(f"An error occurred: {e}")
    
    if species:
        return species
    else:
        print("No species found in the current working directory.")


def get_GA_matrix(cluster_path, EF, list_EF, Taylor=False):
    """
    Retrieve the GA matrix from the CSV file if it exists; otherwise, generate it.
    Then extract the feature matrix X and target vector Y.

    Args:
        EF (str): The EF value (as a string) used to label the adsorption energy column.
                  For example, if EF='0.6', the energy column is assumed to be 'E_ads_0.6'.

    Returns:
        tuple: (df_matrix, X, Y, n_group)
            - df_matrix: The full GA matrix DataFrame (including 'site' and adsorption energy columns).
            - X: The features as a NumPy array (with 'site' and all adsorption energy columns removed).
            - Y: The target adsorption energies (from the corresponding column).
            - n_group: The number of feature columns in X.
    """
    csv_filepath = './GA_matrix.csv'
    df_matrix = pd.read_csv(csv_filepath)
    list_EF = [float(x) for x in list_EF]   # sanitize once

    if Taylor:
        eads_cols = [f"E_ads_{ef:.1f}" for ef in list_EF]
        prop_list = eads_cols + ["polarizability", "dipole", "c"]
    else:
        eads_cols = [f"E_ads_{ef:.1f}" for ef in list_EF]
        prop_list = eads_cols + ["polarizability", "dipole", "c"]
        
    # Extract Y values (adsorption energies)
    if EF not in  ["polarizability", "dipole", "c"]:
        prop = 'E_ads_' + str(EF) 
    else: 
        prop = EF
    try:
        Y = df_matrix[prop].values
    except KeyError:
        raise KeyError(f"The expected adsorption energy column '{EF}' is not found in the GA matrix.")
    
    # Drop the 'site' column and all adsorption energy columns (those starting with the prefix) to get X.
    cols_to_drop = ['site'] + [col for col in prop_list]
    df_features = df_matrix.drop(columns=cols_to_drop)
    X = df_features.values
    n_group = df_features.shape[1]

    return df_matrix, X, Y, n_group


def GA_prediction(X_training, Y_training):
    '''Run conventional GA approach '''
 
    model = PseudoInverseLinearRegression() 
    model.fit(X_training,Y_training)

    Y_pred = model.predict(X_training)
    
    # Calculate the MAE, RMSE, and R²
    mae = mean_absolute_error(Y_training, Y_pred)
    rmse = np.sqrt(mean_squared_error(Y_training, Y_pred))
    r2 = r2_score(Y_training, Y_pred)
    
    # Print the errors and R²
    print(f'Mean Absolute Error (MAE): {mae}')
    print(f'Root Mean Squared Error (RMSE): {rmse}')
    print(f'R-squared (R²): {r2}')
         
    # # Plot the data and regression line
    plt.figure(figsize=(10, 6))
    plt.scatter(Y_training, Y_pred, color='blue', label='Model')
    
    # Add text box for metrics
    textstr = f'MAE: {mae:.2f}\nRMSE: {rmse:.2f}\nR2: {r2:.2f}'
    plt.gca().text(0.65, 0.25, textstr, transform=plt.gca().transAxes,
                    fontsize=16, verticalalignment='top', bbox=dict(facecolor='white', alpha=0))
    
    # Customize the plot
    plt.xlabel('X / Unit', fontsize=16)
    plt.ylabel('Y / Unit', fontsize = 16)
    plt.xticks(fontsize=16)
    plt.yticks(fontsize=16)
    plt.legend(loc='upper left', frameon=False, fontsize=18)
    #plt.title('Linear Regression of De vs Ea')
    plt.tight_layout()
    # Save the figure
    plt.savefig('XY.png')
    return Y_training, Y_pred

    
def save_one_model_split(species, df, EF,  X, Y, models, model_name, ratio=0.95):
    model = models[model_name]
    n_samples = int(len(X) * ratio)  # 80% of the data for training
    indices = np.random.choice(len(X), size=len(X), replace=True)
    X_train, Y_train = X[indices[:n_samples]], Y[indices[:n_samples]]   
    model.fit(X_train,Y_train)
    
    joblib.dump(model, f'{species}_{model_name}_{EF}.pkl')


def test_one_model_split(df, EF,  X, Y, models, model_name, ratio):
    '''Run conventional GA approach '''
    model = models[model_name]
    
    n_samples = int(len(X) * ratio)  # 80% of the data for training
    
    indices = np.random.choice(len(X), size=len(X), replace=True)
    X_train, Y_train = X[indices[:n_samples]], Y[indices[:n_samples]]
    X_test, Y_test = X[indices[n_samples:]], Y[indices[n_samples:]]
    
    model.fit(X_train,Y_train)
    Y_pred = model.predict(X_test)
    
    # Calculate the MAE, RMSE, and R²
    mae = mean_absolute_error(Y_test, Y_pred)
    rmse = np.sqrt(mean_squared_error(Y_test, Y_pred))
    r2 = r2_score(Y_test, Y_pred)

    fig, ax  = plt.subplots(figsize=(10,8))  
    
    species = get_species_from_cwd()
    text = '%s adsorption' %(species)
    # ax.set_title(text)
    ax.scatter(Y_test, Y_pred, edgecolor='k', alpha=0.7, s=200, label = text, color='#1f77b4')
    ax.plot([min(Y), max(Y)], [min(Y), max(Y)], 'r--', linewidth=3)
    ax.set_xlabel(f'DFT Eads (eV) at EF = {EF} eV/A', fontsize=36)
    ax.set_ylabel('Pred. Eads(eV)', fontsize=36)

    textstr = f'MAE: {mae:.2f}  \nRMSE: {rmse:.2f}  \nR2: {r2:.2f}'
    plt.gca().text(0.65, 0.25, textstr, transform=plt.gca().transAxes,
                    fontsize=20, verticalalignment='top', bbox=dict(facecolor='white', alpha=0))


    for i in range(len(Y_test)):
        if abs(Y_test[i] - Y_pred[i]) > 0.15: # 0.15 eV difference
            ax.annotate(df['site'][i], (Y_test[i], Y_pred[i]), textcoords="offset points", xytext=(0,10), ha='center')
            ax.scatter(Y_test[i], Y_pred[i], color='red', s=100, edgecolor='k')
    
    
    plt.legend(loc='upper left', frameon=False, fontsize=40)

    plt.grid(False)
    ax = plt.gca()
    for spine in ax.spines.values():
        spine.set_linewidth(3)
        
    #plt.title('Linear Regression of De vs Ea')
    ax.tick_params(axis='both', which='major', width=2, length=8, labelsize=28)
    plt.tight_layout()
    plt.savefig(f'results/XY_%s_testing_%s_{EF}.png' %(model_name, str(1-ratio)), dpi=600)
    
    return mae, rmse, r2 
        
 
class PseudoInverseLinearRegression:
    def __init__(self, fit_intercept=True):
        self.beta_ = None
        self.fit_intercept = fit_intercept

    def fit(self, X, Y):
        if self.fit_intercept:
            # Add a column of ones to X to account for the intercept
            X = np.hstack([np.ones((X.shape[0], 1)), X])
        
        # Compute the beta coefficients using the Moore-Penrose pseudoinverse
        self.beta_ = np.linalg.pinv(X).dot(Y)

    def predict(self, X):
        if self.fit_intercept:
            # Add a column of ones to X for predictions
            X = np.hstack([np.ones((X.shape[0], 1)), X])
        
        # Make predictions using the computed beta coefficients
        return X.dot(self.beta_)
    

# # Now include this custom model in the models dictionary
def refine_models(n_group):
    models = {
        "Ridge": Ridge(alpha=1.0),
        "Lasso": Lasso(alpha=1/n_group),
        "EN": ElasticNet(alpha=0.1, l1_ratio=0.5),
        "DT": DecisionTreeRegressor(max_depth=n_group),
        "RF": RandomForestRegressor(n_estimators=n_group*5, max_depth=n_group*5),
        "GB": GradientBoostingRegressor(n_estimators=n_group*5, max_depth=n_group),
        "SVR": SVR(kernel='linear', C=1.0),
        "GA": PseudoInverseLinearRegression(),
        # "XGB": XGBRegressor(n_estimators=11*5, max_depth=11, learning_rate=0.1),
        }
    return models


def bootstrap_cross_validation_one_model(models, model_name, EF, X, Y, ratio, n_bootstrap=100):
    """
    Perform bootstrapped cross-validation for a single model and save results.
    
    Args:
        models (dict): Dictionary of models with their names as keys.
        model_name (str): The name of the model to use for cross-validation.
        X (np.array): Features dataset.
        Y (np.array): Target values.
        ratio (float): Proportion of data to use for training (e.g., 0.8 for 80%).
        n_bootstrap (int): Number of bootstrap iterations.
    
    Returns:
        dict: Dictionary containing metric results for the model.
    """
    model = models[model_name]
    results = {metric: [] for metric in ['MAE', 'MSE', 'RMSE', 'R2', 'MPD', 'MND']}
    n_samples = int(len(X) * ratio)  # Number of training samples

    for i in range(n_bootstrap):
        print(f'Iteration: {i + 1}/{n_bootstrap}')
        
        # Generate bootstrap indices
        indices = np.random.choice(len(X), size=len(X), replace=True)
        X_train, Y_train = X[indices[:n_samples]], Y[indices[:n_samples]]
        X_test, Y_test = X[indices[n_samples:]], Y[indices[n_samples:]]
        
        # Train the model and predict
        model.fit(X_train, Y_train)
        predictions = model.predict(X_test)
        
        # Calculate metrics
        results['MAE'].append(mean_absolute_error(Y_test, predictions))
        results['MSE'].append(mean_squared_error(Y_test, predictions))
        results['RMSE'].append(np.sqrt(mean_squared_error(Y_test, predictions)))
        results['R2'].append(r2_score(Y_test, predictions))
        results['MPD'].append(np.max(predictions - Y_test))
        results['MND'].append(np.min(predictions - Y_test))
    
    # Save results to CSV files
    results_dir = 'results'
    os.makedirs(results_dir, exist_ok=True)  # Ensure the results directory exists
    
    for metric, values in results.items():
        metric_file = f'{results_dir}/{metric.lower()}_{ratio}_{EF}.csv'
        averages_file = f'{results_dir}/{metric.lower()}_averages_{ratio}_{EF}.csv'
        
        # Save all values
        pd.DataFrame({model_name: values}).to_csv(metric_file, index=False)
        print(f"Saved all {metric} values to {metric_file}_{EF}")
        
        # Save averages
        pd.DataFrame({model_name: [np.mean(values)]}).to_csv(averages_file, index=False)
        print(f"Saved average {metric} value to {averages_file}_{EF}")

    return results


def clean_ga_matrix(df):
    """
    Cleans a GA matrix DataFrame by:
    1. Removing the first row (assumed to be an extra header),
    2. Dropping the 'site' column,
    3. Converting all remaining values to numeric.
    
    Returns a cleaned DataFrame.
    """
    return df.iloc[1:].drop(columns=['site']).reset_index(drop=True).apply(pd.to_numeric, errors='coerce')


def plot_active_learning_from_csv(EF, results_dir='results'):
    """
    Read the combined active learning results CSV and plot both training and testing
    performance metrics vs. training ratio.

    The CSV is expected to have columns:
      'training_ratio',
      'train_MAE','train_MSE','train_RMSE','train_R2','train_MPD','train_MND',
      'test_MAE', 'test_MSE', 'test_RMSE', 'test_R2', 'test_MPD', 'test_MND'

    Args:
        csv_filepath (str): Path to the combined CSV file.
        EF (str or float): The EF value for labeling the plot.
        results_dir (str): Directory where the plot will be saved.
    """
    overall_csv_file = os.path.join(results_dir, f'active_learning_metrics_{EF}.csv')
    df = pd.read_csv(overall_csv_file)
    df.sort_values('training_ratio', inplace=True)

    plt.figure(figsize=(10, 6))
    # Plot MAE
    plt.plot(df['training_ratio'], df['train_MAE'],  marker='o', linestyle='--', color = 'tab:blue', label='Train MAE')
    plt.plot(df['training_ratio'], df['test_MAE'],   marker='o', linestyle='-', color = 'tab:red', label='Test MAE')
    # Plot RMSE
    plt.plot(df['training_ratio'], df['train_RMSE'], marker='^', linestyle='--', color = 'tab:blue', label='Train RMSE')
    plt.plot(df['training_ratio'], df['test_RMSE'],  marker='^',linestyle='-',  color = 'tab:red', label='Test RMSE')
    # # Plot MPD
    # plt.plot(df['training_ratio'], df['train_MPD'],  marker='D', label='Train MPD')
    # plt.plot(df['training_ratio'], df['test_MPD'],   marker='d', label='Test MPD')
    # # Plot MND
    # plt.plot(df['training_ratio'], df['train_MND'],  marker='x', label='Train MND')
    # plt.plot(df['training_ratio'], df['test_MND'],   marker='*', label='Test MND')

    plt.xlabel('Training Ratio', fontsize=16)
    plt.ylabel('Metric Value', fontsize=16)
    plt.title(f'Active Learning Metrics vs. Training Ratio (EF = {EF})', fontsize=18)
    plt.legend(fontsize=12, framealpha=0.8)
    plt.grid(True)
    plt.tight_layout()

    # Ensure results_dir exists
    os.makedirs(results_dir, exist_ok=True)
    out_png = os.path.join(results_dir, f'active_learning_metrics_{EF}.png')
    plt.savefig(out_png, dpi=300)
    # plt.show()
    plt.close()
    print(f"✅ Plot saved to {out_png}")

def plot_active_predicting_from_csv(ratio, EF, predict_csv_dir='results'):
    """
    Read training and testing prediction CSVs for a given iteration (ratio) and EF,
    plot true vs. predicted scatter for both sets, annotate with MAEs calculated
    directly from the data, and save the figure.

    Args:
        predict_csv_dir (str): Directory containing the CSVs.
        ratio (int or str): The iteration number (used in the filename).
        EF (float or str): The EF value (used in the filename and title).
    """
    # Paths to prediction CSVs
    train_fp = os.path.join(predict_csv_dir, f'train_preds_iter{ratio}_{EF}.csv')
    test_fp  = os.path.join(predict_csv_dir, f'test_preds_iter{ratio}_{EF}.csv')

    # Read data
    df_train = pd.read_csv(train_fp)
    df_test  = pd.read_csv(test_fp)

    # Compute MAEs directly
    mae_train = mean_absolute_error(df_train['Y_true'], df_train['Y_pred'])
    mae_test  = mean_absolute_error(df_test['Y_true'],  df_test['Y_pred'])

    # Create figure/axes
    fig, ax = plt.subplots(figsize=(5, 6))

    # Scatter (圆点)
    ax.scatter(df_train['Y_true'], df_train['Y_pred'],
               color='tab:blue', label=f'Train (MAE = {mae_train:.3f} eV)', alpha=0.6)
    ax.scatter(df_test['Y_true'], df_test['Y_pred'],
               color='tab:red',  label=f'Test  (MAE = {mae_test:.3f} eV)',  alpha=0.6)

    # y = x reference line & limits
    all_vals = np.concatenate([
        df_train[['Y_true','Y_pred']].values.ravel(),
        df_test [['Y_true','Y_pred']].values.ravel()
    ])
    vmin, vmax = np.min(all_vals), np.max(all_vals)
    if vmin == vmax:
        eps = 1e-6
        vmin -= eps; vmax += eps
        
    pad = 0.2
    tick_step = 0.1   # 你想要的每个 tick 间隔（最终会有 4 个间隔）
    
    # 1) nice vmin/vmax（例：2.04 -> 2.0）
    vmin0 = np.floor(vmin / tick_step) * tick_step
    vmax0 = np.ceil (vmax / tick_step) * tick_step
    
    # 2) 加 padding
    lo0 = vmin0 - pad
    hi0 = vmax0 + pad
    
    # 3) 把 lo/hi 再对齐到 tick_step
    lo0 = np.floor(lo0 / tick_step) * tick_step
    hi0 = np.ceil (hi0 / tick_step) * tick_step
    
    # 4) 让区间长度 = 4 * tick_step * integer
    span0 = hi0 - lo0
    n_steps = int(np.ceil(span0 / (4 * tick_step)))   # 至少需要多少个(4*tick_step)
    span = n_steps * 4 * tick_step
    hi = lo0 + span
    lo = lo0
    
    ax.plot([lo, hi], [lo, hi], 'k--', lw=1)
    ax.set_xlim(lo, hi)
    ax.set_ylim(lo, hi)
        
    # ax.plot([vmin-0.2, vmax+0.2], [vmin-0.2, vmax+0.2], 'k--', lw=1)
    # ax.set_xlim(vmin-0.2, vmax+0.2)
    # ax.set_ylim(vmin-0.2, vmax+0.2)
    
    # #For N_ads
    # ax.plot([-1.1, 0.7], [-1.1, 0.7], 'k--', lw=1)
    # ax.set_xlim(-2.0, 0.4)
    # ax.set_ylim(-1.1, 0.7)

    # === 关键：每个轴固定 5 个主刻度 ===
    ax.xaxis.set_major_locator(LinearLocator(5))  # exactly 5 ticks including ends
    ax.yaxis.set_major_locator(LinearLocator(5))
    ax.xaxis.set_major_formatter(FormatStrFormatter('%.2f'))
    ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

    # ax.tick_params(axis='both', which='major', labelsize=18)  # set x/y tick label font size

    # Labels and title
    ax.set_xlabel('Eads(DFT) /eV',  fontsize=18)
    ax.set_ylabel('Eads(GA) / eV', fontsize=18)
    # ax.set_title(f'Iteration {ratio} Predictions (EF = {EF})', fontsize=16)
    ax.legend(framealpha=0, fontsize=14, loc='upper left', handletextpad=0.2)
    ax.tick_params(axis='both', which='major', labelsize=16)  # major ticks

    ax.set_aspect('equal', adjustable='box')  # 可选：图形上 y=x 视觉上为 45°

    plt.tight_layout()

    # Save figure
    out_png = os.path.join(predict_csv_dir, f'preds_iter{ratio}_{EF}.png')
    plt.savefig(out_png, dpi=300)
    plt.close()
    print(f"✅ Saved scatter plot with MAEs to {out_png}")

    # === Additional figure: ALL predictions (train + test) ===
    all_fp = os.path.join(predict_csv_dir, f'all_preds_iter{ratio}_{EF}.csv')
    if os.path.exists(all_fp):
        df_all = pd.read_csv(all_fp)
    
        # Compute MAE for ALL directly from file
        mae_all = mean_absolute_error(df_all['Y_true_all'], df_all['Y_pred_all'])
    
        # Create scatter: ALL
        plt.figure(figsize=(6,6))
        plt.scatter(
            df_all['Y_true_all'], df_all['Y_pred_all'],
            alpha=0.6, label=f'All (MAE = {mae_all:.3f} eV)'
        )
    
        # y = x reference line
        vmin_all = min(df_all['Y_true_all'].min(), df_all['Y_pred_all'].min())
        vmax_all = max(df_all['Y_true_all'].max(), df_all['Y_pred_all'].max())
        plt.plot([vmin_all, vmax_all], [vmin_all, vmax_all], 'k--', lw=1)
    
        # Labels and title
        plt.xlabel('Eads (DFT) / eV', fontsize=22)
        plt.ylabel('Eads (GA) / eV', fontsize=22)
        plt.tick_params(axis='both', which='major', labelsize=16,)
        # plt.title(f'Iteration {ratio} ALL Predictions (EF = {EF})', fontsize=16)
        plt.legend(framealpha=0.8, fontsize=14, loc='upper left', handletextpad=0.2)
        plt.tight_layout()
    
        # Save figure
        out_png_all = os.path.join(predict_csv_dir, f'all_preds_iter{ratio}_{EF}.png')
        plt.savefig(out_png_all, dpi=300)
        plt.close()
        print(f"✅ Saved ALL scatter plot with MAE to {out_png_all}")
    else:
        print(f"⚠️ ALL predictions file not found: {all_fp}")
        
def active_learning_one_model(models, model_name, EF, X, Y, sites, 
                              initial_ratio=0.4, increment=0.05, max_ratio=0.95,
                              error_threshold=0.5, results_dir='results'):
    """
    Perform an active learning process for a single model, saving both training
    and testing predictions and metrics each iteration, and finally output a
    combined CSV of train/test metrics.

    Returns:
        list of dict: metrics per iteration.
    """


    # Ensure arrays
    X = np.asarray(X)
    print(X.shape)
    Y = np.asarray(Y)
    
    # If Y might be (N,1), flatten it
    Y = Y.ravel()
    
    # Build a row-wise validity mask
    valid_mask = np.all(np.isfinite(X), axis=1) & np.isfinite(Y)
    
    # Apply the mask consistently
    X = X[valid_mask]
    Y = Y[valid_mask]
    sites = np.asarray(sites)[valid_mask]


    # if EF == 0:
    #     X = X[:, :-1]   # remove last column

    total_samples = len(X)
    idx = np.arange(total_samples)
    np.random.shuffle(idx)

    # Initial train/test split
    n_train = int(total_samples * initial_ratio)
    train_idx = idx[:n_train].tolist()
    test_idx  = idx[n_train:].tolist()

    os.makedirs(results_dir, exist_ok=True)
    results = []
    iteration = 0

    while len(train_idx)/total_samples < max_ratio and test_idx:
        
        iteration += 1

        # Prepare data
        X_tr, Y_tr = X[train_idx], Y[train_idx]
        X_te, Y_te = X[test_idx], Y[test_idx]
        sites_tr = np.array(sites)[train_idx]
        sites_te = np.array(sites)[test_idx]
        
        # Fit model
        model = models[model_name]
        model.fit(X_tr, Y_tr)

        # Predict
        Y_tr_pred = model.predict(X_tr)
        Y_te_pred = model.predict(X_te)
        Y_all_pred = model.predict(X)

        # Save model
        species = get_species_from_cwd()
        pipe = Pipeline([('scaler', StandardScaler()),('model', model) ])
        pipe.fit(X_tr, Y_tr)
        joblib.dump(pipe, f'{species}_{model_name}_{EF}.pkl')

        # Compute train metrics
        train_mae  = mean_absolute_error(Y_tr, Y_tr_pred)
        train_mse  = mean_squared_error(Y_tr, Y_tr_pred)
        train_rmse = np.sqrt(train_mse)
        train_r2   = r2_score(Y_tr, Y_tr_pred)
        tr_dev     = Y_tr_pred - Y_tr
        train_mpd  = tr_dev.max()
        train_mnd  = tr_dev.min()

        # Compute test metrics
        mae  = mean_absolute_error(Y_te, Y_te_pred)
        mse  = mean_squared_error(Y_te, Y_te_pred)
        rmse = np.sqrt(mse)
        r2   = r2_score(Y_te, Y_te_pred)
        dev  = Y_te_pred - Y_te
        mpd  = dev.max()
        mnd  = dev.min()

        # Save predictions
        pd.DataFrame({
            'site': sites_tr,
            'Y_true': Y_tr,
            'Y_pred': Y_tr_pred
        }).to_csv(
            os.path.join(results_dir, f'train_preds_iter{iteration}_{EF}.csv'),
            index=False
        )

        pd.DataFrame({
            'site': sites_te,
            'Y_true': Y_te,
            'Y_pred': Y_te_pred
        }).to_csv(
            os.path.join(results_dir, f'test_preds_iter{iteration}_{EF}.csv'),
            index=False
        )
            
        #Save ALL predictions
        pd.DataFrame({
            'site': sites,
            'Y_true_all': Y,
            'Y_pred_all': Y_all_pred
        }).to_csv(
            os.path.join(results_dir, f'all_preds_iter{iteration}_{EF}.csv'),
            index=False
        )
        
        # Print all deviations larger than 0.3 eV
        deviations = np.abs(np.array(Y) - np.array(Y_all_pred))
        large_dev_mask = deviations > 0.3
        if large_dev_mask.any():
            large_dev_indices = np.where(large_dev_mask)[0]
            print(f"\n[Iteration {iteration}, EF={EF}] Deviations > 0.3 eV:")
            for idx in large_dev_indices[np.argsort(-deviations[large_dev_indices])]:
                print(f"  Site: {sites[idx]:>6} | Y_true: {Y[idx]:>8.4f} | Y_pred: {Y_all_pred[idx]:>8.4f} | Deviation: {deviations[idx]:>6.4f} eV")
        else:
            print(f"\n[Iteration {iteration}, EF={EF}] No deviations > 0.3 eV")


        # Log iteration
        print(f"Iteration {iteration}: train MAE={train_mae:.4f}, test MAE={mae:.4f}")

        # Outlier analysis
        errs = np.abs(dev)
        mask = errs > error_threshold
        if mask.any():
            outs = np.where(mask)[0]
            for i in outs[np.argsort(-errs[outs])]:
                print(f"  Outlier: site={sites_te[i]}, err={errs[i]:.3f}")
        else:
            print("  No outliers above threshold.")

        # Record metrics
        results.append({
            'iteration':        iteration,
            'training_ratio':   len(train_idx)/total_samples,
            'n_train':          len(train_idx),
            'n_test':           len(test_idx),

            'train_MAE':        train_mae,
            'train_MSE':        train_mse,
            'train_RMSE':       train_rmse,
            'train_R2':         train_r2,
            'train_MPD':        train_mpd,
            'train_MND':        train_mnd,

            'test_MAE':         mae,
            'test_MSE':         mse,
            'test_RMSE':        rmse,
            'test_R2':          r2,
            'test_MPD':         mpd,
            'test_MND':         mnd,
        })

        # Active learning: add top-error test samples to train
        n_add = max(1, int(total_samples * increment))
        n_add = min(n_add, len(test_idx))
        top_idxs = np.argsort(-errs)[:n_add]
        selected = [test_idx[i] for i in top_idxs]
        train_idx.extend(selected)
        test_idx = [i for i in test_idx if i not in selected]
        
        plot_active_predicting_from_csv(iteration, EF, predict_csv_dir='results')

    # Save all iteration metrics
    df = pd.DataFrame(results)
    base_cols = ['iteration','training_ratio','n_train','n_test']
    metric_cols = ['MAE','MSE','RMSE','R2','MPD','MND']
    cols = base_cols + [f'train_{m}' for m in metric_cols] + [f'test_{m}' for m in metric_cols]
    out_file = os.path.join(results_dir, f'active_learning_metrics_{EF}.csv')
    df[cols].to_csv(out_file, index=False)
    print(f"✅ Saved combined metrics to {out_file}")
    plot_active_learning_from_csv(EF, results_dir='results')

    return results


def get_matrix_to_be_predicted(cluster_path, site):
    
    site = str(site)
    
    ga_matrix_file = os.path.join(cluster_path, 'GA_matrix_full.csv')
    GA_matrix_full = pd.read_csv(ga_matrix_file)
    
    def get_feature_by_site(df, site):
        """
        根据 site 值获取对应的一行特征（去掉 'site' 列）。
    
        参数：
            df (pd.DataFrame): 原始DataFrame，包含'site'列和多个特征列
            site (int): 要查询的 site 编号
    
        返回：
            np.ndarray: shape 为 (1, n_features) 的 NumPy 数组
        """
        row = df[df['site'] == site]
        if row.empty:
            raise ValueError(f"Site {site} not found in the DataFrame.")
        return row.drop(columns=['site']).values
    # print(site)
    matrix_site = get_feature_by_site(GA_matrix_full, site)

    return matrix_site

def predict_Eads_site(cluster_path, species, site, Prop):
    '''Print the adsorption energy at the specific site, 
    the data in the data_path is applied to train the GA model to predict the 
    same species at the provided site.
    Prop values: -0.6, -0.4, -0.2, 0.0, 0.2, 0.4, 0.6, polarizability, dipole, all are strings.
    '''
    predict_matrix =  get_matrix_to_be_predicted(cluster_path, site)
    #print(predict_matrix.shape)
    pkl_file = os.path.join(cluster_path,f'{species}_GA_{Prop}.pkl')
    GA_model = joblib.load(pkl_file)  
    E_species_ef = GA_model.predict(predict_matrix)[0]

    return E_species_ef

def predict_Eads_site_Taylor(cluster_path, species, site, Prop):
    '''Print the adsorption energy at the specific site, 
    the data in the data_path is applied to train the GA model to predict the 
    same species at the provided site.
    Prop values: -0.6, -0.4, -0.2, 0.0, 0.2, 0.4, 0.6, polarizability, dipole, all are strings.
    '''
    preidct_matrix =  get_matrix_to_be_predicted(cluster_path, site)
    
    # Prediction via Taylor Expansion
    pkl_polarizability = os.path.join(cluster_path,f'{species}_GA_polarizability.pkl')
    pkl_dipole = os.path.join(cluster_path,f'{species}_GA_dipole.pkl')
    pkl_constant = os.path.join(cluster_path,f'{species}_GA_c.pkl')
    
    polarizability_model = joblib.load(pkl_polarizability)  
    dipole_model = joblib.load(pkl_dipole)  
    constant_model = joblib.load(pkl_constant)  
    
    polarizability  = polarizability_model.predict(preidct_matrix)[0]
    dipole  = dipole_model.predict(preidct_matrix)[0]
    constant = constant_model.predict(preidct_matrix)[0]
    
    E_species_ef = func(float(Prop), polarizability, dipole, constant)  
    return E_species_ef  

### GA Applilcations
    
def find_all_rhombuses(atoms, connections, surface_indices, bond_length_threshold):
    # atoms, connections, cn_of_connected_atoms, exposed_top_sites = get_connection(path, metal='Ru', mult=0.9)
    """Find all possible rhombuses based on atom connections"""
    rhombuses = []
    for A_index in surface_indices:
        A_neighbors = connections[A_index]
        for B_index in A_neighbors:
            if B_index not in surface_indices or A_index == B_index:
                continue
            if np.linalg.norm(atoms[A_index].position - atoms[B_index].position) > bond_length_threshold:
                continue

            B_neighbors = connections[B_index]
            for C_index in B_neighbors:
                if C_index not in surface_indices or C_index in [A_index, B_index]:
                    continue
                if np.linalg.norm(atoms[B_index].position - atoms[C_index].position) > bond_length_threshold:
                    continue

                C_neighbors = connections[C_index]
                for D_index in C_neighbors:
                    if D_index not in surface_indices or D_index in [A_index, B_index, C_index]:
                        continue
                    if np.linalg.norm(atoms[C_index].position - atoms[D_index].position) > bond_length_threshold:
                        continue
                    if np.linalg.norm(atoms[D_index].position - atoms[A_index].position) > bond_length_threshold:
                        continue

                    # Save rhombus indices while maintaining the original order
                    rhombus_indices = [A_index, B_index, C_index, D_index]
                    rhombuses.append(rhombus_indices)

    return rhombuses


def convert_sites_triangle_site(sites_list):
    """
    Given a list of three site numbers corresponding to atoms A, B, C(in clockwise order),
    return a dictionary mapping standard site labels to their numeric string representations.
    
    For example, if sites_list = [20, 23, 24]: # Note the index is 1-based
      - "top_A": "20"
      - "top_B": "23"
      - "top_C": "24"
      - "bridge_A-B": "20-23"
      - "bridge_B-C": "23-24"
      - "bridge_A-C": "20-24"
      - "hollow_ABC": "20-23-24"  (sorted ascending among B, C, D)
    """
    if len(sites_list) != 3:
        raise ValueError("sites_list must contain exactly three elements corresponding to A, B, C.")
    
    A, B, C = sites_list
    sites_ads = [str(i) for i in sites_list]
    bri_1 = f"{min(A, B)}_{max(A, B)}" 
    bri_2 = f"{min(B, C)}_{max(B, C)}"  
    bri_3 = f"{min(A, C)}_{max(A, C)}"

    sites_ads.append(bri_1)
    sites_ads.append(bri_2)
    sites_ads.append(bri_3)
    sites_ads.append("_".join(str(x) for x in sorted([A, B, C])))

    return sites_ads

def load_slab_energy_from_csv(cluster_path, ef_value):
    """
    Load slab energy from e_slab.csv file based on the given EF value.
    
    The e_slab.csv file should have the format:
    Species,site,-0.6,-0.4,-0.2,0.0,0.2,0.4,0.6,site_type
    slab,slab,-356.76171876,-355.45868095,...
    
    Args:
        cluster_path (str or Path): Path to the slab folder containing e_slab.csv
        ef_value (float): The EF value to look up (e.g., -0.6, 0.0, 0.6)
    
    Returns:
        float: The slab energy corresponding to the given EF value
    
    Raises:
        FileNotFoundError: If e_slab.csv is not found
        ValueError: If the EF value is not found in the CSV columns
    """
    import pandas as pd
    
    csv_path = Path(cluster_path) / 'e_slab.csv'
    
    if not csv_path.exists():
        raise FileNotFoundError(f"e_slab.csv not found at {csv_path}")
    
    # Read the CSV file
    df = pd.read_csv(csv_path)
    
    # The EF values are in the column headers (except the first two: 'Species', 'site', and last: 'site_type')
    # Find the column that corresponds to the ef_value
    ef_columns = [col for col in df.columns if isinstance(col, (int, float)) or (isinstance(col, str) and col.lstrip('-').replace('.', '').isdigit())]
    
    # Convert column names to floats for comparison
    ef_column_map = {}
    for col in df.columns:
        try:
            ef_col_float = float(col)
            ef_column_map[ef_col_float] = col
        except (ValueError, TypeError):
            pass
    
    # Find the closest EF value if exact match not found
    if ef_value in ef_column_map:
        ef_col = ef_column_map[ef_value]
    else:
        # Find closest EF value
        available_ef = sorted(ef_column_map.keys())
        if ef_value < min(available_ef) or ef_value > max(available_ef):
            raise ValueError(f"EF value {ef_value} is outside the range of available values: {available_ef}")
        
        # Linear interpolation between two closest values
        for i, ef in enumerate(available_ef):
            if ef >= ef_value:
                if i == 0:
                    ef_col = ef_column_map[ef]
                else:
                    # Interpolate between available_ef[i-1] and available_ef[i]
                    ef_lower = available_ef[i-1]
                    ef_upper = available_ef[i]
                    col_lower = ef_column_map[ef_lower]
                    col_upper = ef_column_map[ef_upper]
                    
                    energy_lower = float(df.loc[df['Species'] == 'slab', col_lower].values[0])
                    energy_upper = float(df.loc[df['Species'] == 'slab', col_upper].values[0])
                    
                    # Linear interpolation
                    weight = (ef_value - ef_lower) / (ef_upper - ef_lower)
                    e_slab = energy_lower + weight * (energy_upper - energy_lower)
                    return e_slab
                break
    
    # Get the slab energy for the selected column
    e_slab = float(df.loc[df['Species'] == 'slab', ef_col].values[0])
    return e_slab

def obtain_lowest_site(data_path, species, sites, EF):
    """
    Evaluate the predicted adsorption energy for a given species at each candidate site.
    If site_mapping is provided, convert the candidate label to its numeric string.
    
    Args:
        cluster_path (str): Path to the cluster data.
        species (str): The adsorbate species (e.g., "NH3", "NH2", "NH", "N", "H", "N2").
        sites (list): List of 1-based candidate site labels (e.g., "20", "21", '20_21', '18_19_20').
        EF: Additional properties required by predict_Eads_site. It can be either the EF values or polarizability 
    
    Returns:
        tuple: (best_site, energy_dict) where energy_dict is keyed by the candidate (symbolic label).
    """

    cluster_path = data_path / 'slab' 
    
    # Ensure EF is a float
    ef_value = float(EF)

    # Load slab energy from e_slab.csv
    E_slab = load_slab_energy_from_csv(cluster_path, ef_value)
        
    energy_dict = {}
    sites_ads = convert_sites_triangle_site(sites)
    

    # sites_ads:     ['5', '13', '29', '5_13', '13_29', '5_29', '5_13_29']
    
    def site_kind(site: str) -> str:
        s = str(site)
        n = s.count('_')
        if n == 0:
            return 'top'
        elif n == 1:
            return 'bridge'
        elif n == 2:
            return 'hollow'
        else:
            return 'other'

    def is_allowed_site(species: str, site: str) -> bool:
        k = site_kind(site)
        if species == 'NH3':
            return k == 'top'
        if species == 'NH2':
            return k in ('top', 'bridge')
        return True  # 其他物种不限制



    for site in sites_ads:
        if not is_allowed_site(species, site):
            continue
    
        filename = f"{species}_ads/results/all_preds_iter4_{EF}.csv"
        filepath = os.path.join(data_path, filename)
        df = pd.read_csv(filepath)
    
        result_dict = df.set_index("site")[["Y_true_all", "Y_pred_all"]].T.to_dict("list")
        result_dict = {str(k): v for k, v in result_dict.items()}
    
        site_str = str(site)
        if site_str in result_dict:
            Eads = float(result_dict[site_str][0])
        else:
            Eads = predict_Eads_site(cluster_path, species, site_str, EF)
    
        energy_dict[site_str] = Eads
        
    
    best_site = min(energy_dict, key=energy_dict.get)
    E_ads_best = energy_dict[best_site]
    
    # 
    EADS_FLOOR = {
        'H':   -1.0,
        'NH3': -1.8,
        'NH2': -4.0,
        'NH':  -5.3,
        'N':   -1.0,
        'N2': -1.5
    }
    
    floor = EADS_FLOOR.get(species)
    E_ads_best = max(float(E_ads_best), floor)  # 若 E_ads_best < floor，则用 floor
    
    if species == 'N':
        DFT_best = E_ads_best + E_slab + gas_dict['N2'] * 0.5
    elif species == 'H':
        DFT_best = E_ads_best + E_slab + gas_dict['H2'] * 0.5
    else: 
        DFT_best = E_ads_best + E_slab + gas_dict[species]
        
            
    print(f"Best {species} site: {best_site} under {EF} with Eads = {energy_dict[best_site]:.3f} eV\n")
    
    return {species: [best_site, E_ads_best, DFT_best]}


def compute_EDFT(data_path, site, EF):
    """
    Compute the DFT simulated energy (EDFT) for each species using the following formulas:
      - EDFT(NH3) = Eads(NH3) + E(slab) + E(NH3_gas)
      - EDFT(NH2) = Eads(NH2) + E(slab) + E(NH2_gas)
      - EDFT(NH)  = Eads(NH)  + E(slab) + E(NH_gas)
      - For N, H1, H2, and H3, use:
            EDFT = Eads + E(slab) + 0.5 * E(H2_gas)
      - EDFT(N2)  = Eads(N2)  + E(slab) + E(N2_gas)

    Returns a dictionary:
      {
         "NH3": (site, Eads, EDFT),
         ...
      }
    """

    ef_value = float(EF)
    # E_slab  = -6.2414 * EF**2 - 0.01405 * EF - 354.4197 

    # Load slab energy from e_slab.csv file
    E_slab = load_slab_energy_from_csv(data_path / 'slab', ef_value)

    edft = {}

    for species in ["NH3", "NH2", "NH", "N", "H", "N2"]:
        dict_species = obtain_lowest_site(data_path, species, site, EF)
        edft.update(dict_species)

    return edft, E_slab


### MKM 

def generate_replacement_dict(ef_value, adsorption_data: dict, data_path=None) -> dict:
    """
    根据吸附态字典和 Ru 的能量，生成 replacement_dict，用于替换 Excel Sheet1 中的数据。

    参数：
    - adsorption_data: 吸附物字典，格式为 {ads: (site, adsorption_energy, total_energy)}
    - ef_value: EF值
    - data_path: 数据路径，用于读取 e_slab.csv（可选，如果提供则使用CSV，否则使用默认字典）

    返回：
    - replacement_dict: key 为 Excel 中的 A 列项（如 "NH3(T)"），value 为替换后的能量值（float）
    """

    # E_slab energy dictionary for different EF values
    E_slab = load_slab_energy_from_csv(Path(data_path) / 'slab', ef_value)

    dict_H = {
        -0.7: -361.62253006,
        -0.6: -360.8042509,
        -0.5: -360.11240866,
        -0.4: -359.55221139,
        -0.3: -359.12110435,
        -0.2: -358.8141788,
        -0.1: -358.63071895,
        0.0: -358.57019498,
        0.1: -358.63223688,
        0.2: -358.81681762,
        0.3: -359.12893043,
        0.4: -359.57204797,
        0.5: -360.14213764,
        0.6: -360.83989921,
        0.7: -361.66972994
        }
    
    replacement_dict = {"RU(T)": E_slab}
    # replacement_dict['H(T)'] = E_H

    # Adsorption species 
    for species in ['NH3', 'NH2', 'NH', 'N', 'N2', 'H']:
        if species in adsorption_data:
            replacement_dict[f"{species}(T)"] = adsorption_data[species][2]
    E_N_N = 2 * adsorption_data['N'][1] + gas_dict['N2'] +  E_slab  + 0.5 
    
    replacement_dict['N_N(T)'] = E_N_N
    
    # Calculate energies for transition states and intermediates

    E_NH3 = adsorption_data['NH3'][2]
    E_NH2 = adsorption_data['NH2'][2]
    E_NH = adsorption_data['NH'][2]
    E_N  = adsorption_data['N'][2]
    E_N2 = adsorption_data['N2'][2]
    E_H = adsorption_data['H'][2]

    # if float(adsorption_data['NH3'][1]) <= -1.9:
    #     print('E_NH3', adsorption_data['NH3'])


    #V1
    # E1  = E_NH2 + E_H - E_NH3 - E_slab
    # E2  = E_NH + E_H - E_NH2 - E_slab
    # E3  = E_N + E_H - E_NH - E_slab
    # E4 = E_slab + gas_dict['N2'] - E_N_N     
    
    #V2
    # E1  = E_NH2 + E_H - E_NH3 - E_slab
    # E2  = E_NH + E_H - E_NH2 - E_slab
    # E3  = E_N + E_H - E_NH - E_slab
    # E4 = E_N2 - E_N_N    
    
    # v3
    E1 = E_NH2 + (-6.76668776)/2 - E_NH3 
    E2 = E_NH  + (-6.76668776)/2 - E_NH2  
    E3 = E_N   + (-6.76668776)/2 - E_NH  
    E4 = E_slab + gas_dict['N2'] - E_N_N     
 
    # #v4
    # E1 = E_NH2 + (-6.76668776)/2 - E_NH3 
    # E2 = E_NH  + (-6.76668776)/2 - E_NH2  
    # E3 = E_N   + (-6.76668776)/2 - E_NH  
    # E4 = E_N2 - E_N_N   
        
    # BEP scaling Ea = a * ΔE + b
    Ea_params = [
        (0.52, 0.90),  # TS1_NH3(T)
        (0.86, 0.73),  # TS2_NH2(T)
        (0.62, 0.75),  # TS3_NH(T)
        (0.62, 1.39)   # TS4_N2(T)
    ]

    Ea1 = Ea_params[0][0] * E1 + Ea_params[0][1]
    Ea2 = Ea_params[1][0] * E2 + Ea_params[1][1]
    Ea3 = Ea_params[2][0] * E3 + Ea_params[2][1]
    Ea4 = Ea_params[3][0] * E4 + Ea_params[3][1] 

    Ea_list = [Ea1, Ea2, Ea3, Ea4]
    Ea1, Ea2, Ea3, Ea4 = [max(0.01, ea) for ea in Ea_list]
    
    dict_energy = {
        "E_H":   adsorption_data["H"][1],
        "E_N":   adsorption_data["N"][1],
        "E_NH":  adsorption_data["NH"][1],
        "E_NH2": adsorption_data["NH2"][1],
        "E_NH3": adsorption_data["NH3"][1],
        "Ea1":   Ea_list[0],
        "Ea2":   Ea_list[1],
        "Ea3":   Ea_list[2],
        "Ea4":   Ea_list[3],
    }

    # print(adsorption_data["NH3"])
    # dft energy for transition states
    replacement_dict['TS1_NH3(T)'] = E_NH3 + Ea1
    replacement_dict['TS2_NH2(T)'] = E_NH2 + Ea2
    replacement_dict['TS3_NH(T)']  = E_NH  + Ea3
    replacement_dict['TS4_N2(T)']  = E_N_N + Ea4

    ref_dict = {k: v - E_slab for k, v in replacement_dict.items()}

    return ref_dict, dict_energy

def update_excel_with_replacement(index_list, replacement_dict, ef_value,
                                  template_file="NH3_temp.xlsx", base_dir="mkm_inputs", verbose=True):
    """
    Create a structured directory based on index and EF, update the "species" sheet 
    in the Excel template with new data, execute additional workflow commands, and save the result.
    
    Parameters:
        index_list: List of indices used for naming subfolders, e.g. [12, 43, 40, 44]
        replacement_dict: Dictionary with replacement items, e.g. {"NH3(T)": -375.2, ...}
        ef_value: External EF value (float, can be negative) for naming subfolders (e.g. -0.3)
        template_file: Excel template file name; now expected to be located inside base_dir.
        base_dir: Base directory (e.g. "mkm_inputs")
        verbose: Whether to print verbose output.
    """
    # 1. Build directory paths
    index_folder = "_".join(map(str, index_list))
    ef_folder = f"EF_{ef_value:.1f}"  # e.g. EF_-0.3
    inputs_folder = os.path.join(base_dir, index_folder, ef_folder, 'inputs')
    os.makedirs(inputs_folder, exist_ok=True)
    # Create outputs folder at the same level as inputs
    outputs_folder = os.path.join(base_dir, index_folder, ef_folder, 'outputs')
    os.makedirs(outputs_folder, exist_ok=True)
    
    # 2. Load the Excel template
    template_path = os.path.join(base_dir, template_file)
    if not os.path.exists(template_path):
        sys.exit(f"Template file not found: {template_path}")
        
    from openpyxl import load_workbook    
    wb = load_workbook(template_path)
    dft_sheet = wb["species"]
    
    # 3. Replace data in the sheet: get keys from column A, replace potential energy in column L"
    if verbose:
        print(f"✅ Fill the Excel Template File")
    
    missing_keys = []
    replaced_count = 0
    for row in dft_sheet.iter_rows(min_row=2, min_col=1, max_col=12):
        key_cell = row[0]    # Column A (key)
        target_cell = row[11]  # Column L (the value to replace)
        key = key_cell.value
        if key in replacement_dict:
            target_cell.value = replacement_dict[key]
            if verbose:
                print(f"   ✅ Replace: {key} -> {replacement_dict[key]:.6f}")
            replaced_count += 1
        else:
            missing_keys.append(key)
            
    if verbose:
        print(f"   Total replaced: {replaced_count}")
        if missing_keys:
            print(f"   Missing keys: {missing_keys}")
    
    # 4. Save the updated Excel file into the inputs folder
    output_path = os.path.join(inputs_folder, 'NH3_Input_Data.xlsx')
    wb.save(output_path)
    # print(f"📁 File saved to: {output_path}")
    
    # 5. Copy files to the outputs folder (from base_dir)
    for filename in ["drc_qli.py"]:
        src = os.path.join(base_dir, filename)
        dst = os.path.join(outputs_folder, filename)
        if os.path.exists(src):
            shutil.copy2(src, dst)
            # if verbose:
                # print(f"✅ Copied {filename} to {outputs_folder}")
        # else:
            # print(f"⚠️ File {filename} not found!")
    
    # Copy OpenMKM_IO.py to the folder: mkm_inputs/index_folder/ef_folder
    src_IO = os.path.join(base_dir, "OpenMKM_IO.py")
    dst_IO_dir = os.path.join(base_dir, index_folder, ef_folder)
    dst_IO = os.path.join(dst_IO_dir, "OpenMKM_IO.py")
    if os.path.exists(src_IO):
        shutil.copy2(src_IO, dst_IO)
    
    # 6. generate YAML file by executing OpenMKM_IO.py
    main_folder = os.path.join(base_dir, index_folder, ef_folder)    
    prepare_yaml(main_folder)

def prepare_yaml(main_folder):
    """
    Execute OpenMKM_IO.py to generate YAML thermodynamic data file.
    
    Parameters:
        main_folder: Path to the folder containing OpenMKM_IO.py and inputs folder
    
    Output:
        Generates thermo.yaml in the outputs folder
    """
    print("✅ Generating YAML file from Excel data...")
    
    yaml_path = os.path.join(main_folder, "outputs", "thermo.yaml")
    outputs_folder = os.path.join(main_folder, "outputs")
    
    # Create outputs folder if it doesn't exist
    os.makedirs(outputs_folder, exist_ok=True)
    
    # Run OpenMKM_IO.py to generate YAML from Excel input
    try:
        result = subprocess.run(
            ["python3", "OpenMKM_IO.py"], 
            cwd=main_folder, 
            capture_output=True,
            text=True,
            timeout=60
        )
        
        if result.returncode == 0:
            if os.path.exists(yaml_path):
                print(f"   ✅ YAML file successfully created: {yaml_path}")
                return yaml_path
            else:
                print(f"   ⚠️ Warning: OpenMKM_IO.py ran but YAML file not found")
                if result.stdout:
                    print(f"   Output: {result.stdout[:200]}")
                return None
        else:
            print(f"   ❌ Error running OpenMKM_IO.py (exit code: {result.returncode})")
            if result.stderr:
                print(f"   Error details: {result.stderr[:300]}")
            return None
            
    except subprocess.TimeoutExpired:
        print(f"   ❌ OpenMKM_IO.py timed out after 60 seconds")
        return None
    except Exception as e:
        print(f"   ❌ Unexpected error: {str(e)}")
        return None

################## Section MKM

def get_k(Ea, T=673):
    kb = 8.617333262145E-5  # Boltzmann's constant in eV/K
    kbT = kb * T  # Boltzmann's constant times Temperature (in eV)
    A = kbT / (h / eV)  # Pre-exponential factor in s^-1, converting h to eV·s
    kf = A * math.exp(-Ea / kbT)  # Activation energy in eV, so no conversion needed
    return kf

def get_rate(kf0, kr0):
    # Given constants and conversion factors
    J2eV = 6.24150974E18            # eV/J
    Na = 6.0221415E23               # mol-1
    T = 823                         # Temperature in Kelvin
    h = 6.626068E-34 * J2eV         # Planck's constant in eV*s
    kb = 1.3806503E-23 * J2eV       # Boltzmann's constant in eV/K
    kbT = kb * T                    # Boltzmann's constant times Temperature
    kJ_mol2eV = 0.01036427
    wave2eV = 0.00012398
    FNH30 = 1                       # Inlet molar flow rate (assumed in mol/s)
    p0 = 1                          # Total pressure in atm
    # Function to calculate the partial pressures
    def get_pressures(FNH30, X, p0):
        # Calculating flow rates of NH3, N2, and H2
        FNH3 = FNH30 * (1 - X)
        FN2 = FNH30 * 0.5 * X
        FH2 = FNH30 * 1.5 * X
        
        # Total flow rate
        FT = FNH3 + FN2 + FH2
        
        # Calculating partial pressures
        pNH3 = (FNH3 / FT) * p0
        pN2 = (FN2 / FT) * p0
        pH2 = (FH2 / FT) * p0
        
        return pNH3, pN2, pH2

    # kf0 = [5631186144.388563, 50.01374140854966, 167.3703043694318, 35375.898901179375, 1.4277669985218864e-08, 10257.81417760215]
    # kr0 = [133466475.11618523, 78037.92020523513, 51662039.04911003, 526671805.90369725, 1.6185464638531345e-08, 66391314202.84105]
    
    # Function to get the rates
    def get_rates(theta, kf, kr, X, FNH30, p0):
        pNH3, pN2, pH2 = get_pressures(FNH30, X, p0)
        
        tNH3, tNH2, tNH, tN, tH = theta
        tstar = 1.0 - tNH3 - tNH2 - tNH - tN - tH  # Site balance for tstar
        
        rate = np.zeros(6)
        rate[0] = kf[0] * pNH3 * tstar - kr[0] * tNH3
        rate[1] = kf[1] * tNH3 * tstar - kr[1] * tNH2 * tH
        rate[2] = kf[2] * tNH2 * tstar - kr[2] * tNH * tH
        rate[3] = kf[3] * tNH * tstar - kr[3] * tN * tH
        rate[4] = kf[4] * tN ** 2 - kr[4] * pN2 * tstar
        rate[5] = kf[5] * tH ** 2 - kr[5] * pH2 * tstar
        return rate
    
    # Function to get the ODEs
    def get_odes(rate):
        dt = [0] * 5
        dt[0] = rate[0] - rate[1]                          # d(tNH3)/dt
        dt[1] = rate[1] - rate[2]                          # d(tNH2)/dt
        dt[2] = rate[2] - rate[3]                          # d(tNH)/dt
        dt[3] = rate[3] - 2 * rate[4]                      # d(tN)/dt
        dt[4] = rate[1] + rate[2] + rate[3] - 2 * rate[5]  # d(tH)/dt
        return dt
    
    def combined_odes(initial_state, t, kf, kr, FNH30, p0):
        X = initial_state[-1]
        theta = initial_state[:-1]
        
        rates = get_rates(theta, kf, kr, X, FNH30, p0)
        
        ds_dt = get_odes(rates)
        
        # Example improvement: Calculate dx_dt based on a specific rate
        # Assuming rate[5] corresponds to a key product formation step
        dx_dt = rates[4] / FNH30  # Normalize by inlet flow to represent conversion rate
        
        comb_odes = np.append(ds_dt, dx_dt)
        
        return comb_odes
    
    def solve_combined_odes(kf, kr, X0, FNH30, p0):
        thetas0 = np.zeros(5)
        initial_state = np.append(thetas0, X0)
        start_time = 0
        end_time = 1e6
        
        def combined_odes_wrapper(t, y):
            return combined_odes(y, t, kf, kr, FNH30, p0)
            
        # Solving the ODEs
        solution = solve_ivp(combined_odes_wrapper, [start_time, end_time], initial_state, method='LSODA', dense_output=True)
    
        t = np.linspace(start_time, end_time, 10000) 
        sol = solution.sol(t)
        thetas_solution = sol[:-1, :]
        X_solution = sol[-1, :]
            
        return thetas_solution, X_solution, t 
    
    # Solving for theta0 and rates
    X0 = 0.0
    # print('Running')
    thetas_t, X_t, t = solve_combined_odes(kf0, kr0, X0, FNH30, p0)
    # print('Done')
    rates = get_rates(thetas_t[:,-1], kf0, kr0, X_t[-1],  FNH30 , p0)
    tof = p0 *  X_t[-1]  / 10000 
    print('tof:', tof, 'lg(tof):', math.log10(tof), 'Conversion:', X_t[-1], 'rate_rds:',  min(rates)  )
    
    # print(thetas_t[:,-1],  X_t[-1])
    # print(math.log(p0 /100000 * X_t[-1]  / 1000000 ))
    # print("XNH3:", X_t[-1], )

    return rates